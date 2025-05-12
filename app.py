from flask import Flask, render_template, request, flash, redirect, url_for, send_file, jsonify, make_response, session, Response
import yaml
import subprocess
from datetime import datetime, timedelta
import os
import shutil
import re
import json
import time
import requests
import logging
import io
import csv
import threading
from logging.handlers import RotatingFileHandler
from dataclasses import dataclass, field
import notifications as notifications_module

# Set up logging
def setup_logging():
    """Configure application logging"""
    log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logs')
    if not os.path.exists(log_dir):
        os.makedirs(log_dir)

    log_file = os.path.join(log_dir, 'certifly.log')

    # Create a logger
    logger = logging.getLogger('certifly')
    logger.setLevel(logging.INFO)

    # Create handlers
    file_handler = RotatingFileHandler(log_file, maxBytes=1024*1024*5, backupCount=5)
    console_handler = logging.StreamHandler()

    # Create formatters
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)

    # Add handlers to logger
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger

# Initialize logger
logger = setup_logging()

@dataclass
class CertificateStatus:
    domain: str
    days_remaining: int
    expiry_date: datetime
    status: str  # 'valid', 'warning', 'expired'
    ping_status: str = "unknown"  # 'up', 'down', 'unknown'

    def __post_init__(self):
        # Ensure days_remaining is an integer
        try:
            self.days_remaining = int(self.days_remaining)
        except (ValueError, TypeError):
            logger.warning(f"Invalid days_remaining for {self.domain}: {self.days_remaining}. Setting to -1.")
            self.days_remaining = -1

@dataclass
class DomainStatus:
    name: str
    days_remaining: int
    expiry_date: datetime
    registrar: str
    status: str  # 'valid', 'warning', 'expired', 'error'
    ping_status: str = "unknown"  # 'up', 'down', 'unknown'

    def __post_init__(self):
        # Ensure days_remaining is an integer
        try:
            self.days_remaining = int(self.days_remaining)
        except (ValueError, TypeError):
            logger.warning(f"Invalid days_remaining for {self.name}: {self.days_remaining}. Setting to -1.")
            self.days_remaining = -1

@dataclass
class PingStatus:
    """Ping monitoring status"""
    host: str
    status: str  # 'up', 'down', 'unknown'
    last_checked: datetime = None
    response_time: float = 0.0  # in milliseconds
    response_history: list = field(default_factory=list)  # List of recent response times

# Ping cache - using database cache
PING_CACHE_EXPIRY = 300  # 5 minutes in seconds

def get_cached_ping_data(domain):
    """Get ping data from cache if available and not expired"""
    cache_key = f"ping_{domain}"
    cached_data = db.get_cache(cache_key)

    if cached_data:
        logger.debug(f"Using cached ping data for {domain}")
        return cached_data

    return None

def cache_ping_data(domain, ping_result):
    """Cache ping data for a domain"""
    cache_key = f"ping_{domain}"

    # Prepare data for caching
    cache_data = {
        'timestamp': time.time(),
        'status': ping_result['status'],
        'response_time': ping_result['response_time']
    }

    # Log what we're caching
    logger.debug(f"Caching ping data for {domain}: status={ping_result['status']}, " +
                f"response_time={ping_result['response_time']}")

    # Cache the data
    db.set_cache(cache_key, cache_data, PING_CACHE_EXPIRY)

def check_ping(domain):
    """Check if a domain is reachable via ping and return status and response time with caching"""
    # Validate the domain
    try:
        domain = validate_domain(domain)
    except ValueError as e:
        logger.error(str(e))
        return {
            "status": "invalid",
            "response_time": 0.0,
            "last_checked": datetime.now()
        }

    # Default response for errors
    default_down_response = {
        "status": "down",
        "response_time": 0.0,
        "last_checked": datetime.now()
    }

    default_unknown_response = {
        "status": "unknown",
        "response_time": 0.0,
        "last_checked": datetime.now()
    }

    logger.debug(f"Checking ping for domain: {domain}")

    # Check if we have valid cached data
    cached_data = get_cached_ping_data(domain)
    if cached_data:
        logger.debug(f"Using cached ping data for {domain}")
        return {
            "status": cached_data['status'],
            "response_time": cached_data['response_time'],
            "last_checked": datetime.fromtimestamp(cached_data['timestamp']) if 'timestamp' in cached_data else datetime.now()
        }

    try:
        import platform
        import socket

        # First try a quick socket connection to port 80 or 443
        # This is often faster than ping and works in more environments
        try:
            # Try to resolve the domain first
            socket.gethostbyname(domain)

            # Try to connect to port 443 (HTTPS) first, then 80 (HTTP)
            for port in [443, 80]:
                try:
                    # Measure the actual response time
                    start_time = time.time()

                    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                    sock.settimeout(1)  # 1 second timeout
                    result = sock.connect_ex((domain, port))
                    sock.close()

                    end_time = time.time()

                    if result == 0:  # Connection successful
                        # Calculate actual response time in milliseconds
                        response_time = (end_time - start_time) * 1000  # Convert to ms

                        response = {
                            "status": "up",
                            "response_time": round(response_time, 2),  # Round to 2 decimal places
                            "last_checked": datetime.now()
                        }
                        cache_ping_data(domain, response)
                        # Record ping status in history
                        db.record_ping_status(domain, "up", response_time)
                        return response
                except:
                    continue
        except socket.gaierror:
            # DNS resolution failed, domain likely doesn't exist
            cache_ping_data(domain, default_down_response)
            # Record ping status in history
            db.record_ping_status(domain, "down", 0.0)
            return default_down_response

        # If socket connection failed, fall back to ping
        # Different ping command parameters for Windows vs Unix-like systems
        is_windows = platform.system().lower() == 'windows'
        param = '-n' if is_windows else '-c'
        timeout_param = '-w' if is_windows else '-W'
        timeout_value = '1000' if is_windows else '1'  # Windows uses milliseconds

        # Use a short timeout and only 1 packet with explicit timeout parameter
        command = ['ping', param, '1', timeout_param, timeout_value, domain]

        logger.debug(f"Running ping command: {' '.join(command)}")

        # Run the ping command with a timeout
        result = subprocess.run(
            command,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=2,  # Process timeout as a safety measure
            text=True
        )

        if result.returncode != 0:
            logger.debug(f"Ping failed for {domain} with return code {result.returncode}")
            cache_ping_data(domain, default_down_response)
            # Record ping status in history
            db.record_ping_status(domain, "down", 0.0)
            return default_down_response

        # Extract ping time from output using regex
        pattern = r'time[=<](\d+)ms' if is_windows else r'time=([\d.]+) ms'
        match = re.search(pattern, result.stdout)

        if not match:
            # If we can't extract the time but ping was successful, try to measure it ourselves
            logger.debug(f"Ping successful for {domain} but couldn't extract time from output")

            # Measure the response time manually with a second ping
            try:
                start_time = time.time()
                # Run a quick ping
                subprocess.run(
                    command,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    timeout=2,
                    text=True
                )
                end_time = time.time()

                # Calculate response time in milliseconds
                measured_time = (end_time - start_time) * 1000

                response = {
                    "status": "up",
                    "response_time": round(measured_time, 2),
                    "last_checked": datetime.now()
                }
            except:
                # If that fails too, use a default value
                response = {
                    "status": "up",
                    "response_time": 100.0,  # Default value if we can't measure the actual time
                    "last_checked": datetime.now()
                }

            cache_ping_data(domain, response)
            # Record ping status in history
            db.record_ping_status(domain, "up", response["response_time"])
            return response

        # Return successful ping with extracted time
        response_time = float(match.group(1))
        logger.debug(f"Ping successful for {domain} with response time {response_time}ms")
        response = {
            "status": "up",
            "response_time": round(response_time, 2),  # Round to 2 decimal places
            "last_checked": datetime.now()
        }
        cache_ping_data(domain, response)
        # Record ping status in history
        db.record_ping_status(domain, "up", response_time)
        return response

    except subprocess.TimeoutExpired:
        # If the ping command times out
        logger.warning(f"Ping command timed out for {domain}")
        cache_ping_data(domain, default_down_response)
        # Record ping status in history
        db.record_ping_status(domain, "down", 0.0)
        return default_down_response
    except subprocess.SubprocessError as e:
        # Other subprocess errors
        logger.warning(f"Subprocess error in check_ping for {domain}: {str(e)}")
        cache_ping_data(domain, default_down_response)
        # Record ping status in history
        db.record_ping_status(domain, "down", 0.0)
        return default_down_response
    except Exception as e:
        # Log the specific error for debugging
        logger.error(f"Error in check_ping for {domain}: {str(e)}", exc_info=True)
        # Record ping status in history
        db.record_ping_status(domain, "unknown", None)
        return default_unknown_response

@dataclass
class Alert:
    domain: str
    type: str  # 'ssl' or 'domain'
    status: str
    message: str
    date: str

@dataclass
class EmailSettings:
    smtp_server: str
    smtp_port: int
    smtp_username: str
    smtp_password: str
    notification_email: str
    warning_threshold_days: int

@dataclass
class AppSettings:
    auto_refresh_enabled: bool
    auto_refresh_interval: int
    theme: str
    timezone: str
    warning_threshold_days: int = 10

@dataclass
class NotificationSettings:
    """Notification settings for different platforms"""
    email: dict
    teams: dict
    slack: dict
    discord: dict
    telegram: dict
    webhook: dict
    sms: dict

@dataclass
class Stats:
    total: int
    valid: int
    warning: int
    expired: int
    error: int

@dataclass
class PingStats:
    total: int
    up: int
    down: int
    unknown: int

import auth
import database as db
import secrets
from database import get_user_preference, set_user_preference
from api.uptime_api import uptime_api

app = Flask(__name__, static_folder='static')
app.secret_key = os.environ.get('FLASK_SECRET_KEY', os.urandom(24))

# Register API blueprints
app.register_blueprint(uptime_api)

# Initialize authentication system and create default admin user if needed
auth.initialize_auth()
logger.info("Authentication system initialized")

# Initialize scheduler for background tasks
def init_app_scheduler():
    """Initialize the scheduler for background tasks"""
    if os.environ.get('DISABLE_SCHEDULER') == 'true':
        logger.info("Scheduler disabled by environment variable")
        return

    try:
        import scheduler
        scheduler.init_scheduler(app)
        logger.info("Scheduler initialized")
    except ImportError:
        logger.warning("APScheduler not installed, skipping scheduler initialization")
    except Exception as e:
        logger.error(f"Error initializing scheduler: {e}")

# Initialize scheduler after app is fully configured
if not os.environ.get('FLASK_ENV') == 'test':
    init_app_scheduler()

# CSRF token generation
def generate_csrf_token():
    if '_csrf_token' not in session:
        session['_csrf_token'] = secrets.token_hex(16)
    return session['_csrf_token']

# Add CSRF token to all templates
app.jinja_env.globals['csrf_token'] = generate_csrf_token

# Add current timestamp to all templates to prevent image caching
@app.context_processor
def inject_now():
    return {'now': int(time.time())}

# Add profile picture helper function to templates
@app.context_processor
def inject_profile_picture_helper():
    def profile_picture(profile_image, size=32, font_size=18):
        """
        Generate HTML for a profile picture with fallback to default icon

        Args:
            profile_image: Path to profile image or None
            size: Size of the profile image in pixels
            font_size: Size of the default icon font in pixels

        Returns:
            Dictionary with profile image information
        """
        return {
            'profile_image': profile_image,
            'size': size,
            'font_size': font_size
        }

    return {'profile_picture': profile_picture}

# Function to clean up old temporary profile images
def cleanup_temp_profile_images():
    """Remove temporary profile images older than 1 hour"""
    import os
    import time
    from datetime import datetime, timedelta

    temp_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'temp')
    if not os.path.exists(temp_dir):
        return

    # Get current time
    now = time.time()

    # Check all files in the temp directory
    for filename in os.listdir(temp_dir):
        file_path = os.path.join(temp_dir, filename)

        # Skip if not a file
        if not os.path.isfile(file_path):
            continue

        # Get file modification time
        file_mod_time = os.path.getmtime(file_path)

        # Remove if older than 1 hour
        if now - file_mod_time > 3600:  # 3600 seconds = 1 hour
            try:
                os.remove(file_path)
                logger.info(f"Removed old temporary file: {filename}")
            except Exception as e:
                logger.error(f"Error removing temporary file {filename}: {str(e)}")

# Run cleanup on startup
cleanup_temp_profile_images()

# Custom Jinja2 filter for formatting datetime
@app.template_filter('datetime')
def format_datetime(value, format='%Y-%m-%d %H:%M:%S'):
    """Format a datetime object or timestamp to string with user's timezone"""
    # Import pytz at the top level
    import pytz
    from datetime import timezone

    # Get the current user's timezone preference
    current_user = auth.get_current_user()
    if current_user and 'settings' in current_user and 'timezone' in current_user['settings']:
        timezone_str = current_user['settings']['timezone']
        logger.debug(f"Using user's timezone preference: {timezone_str}")
    else:
        # Fallback to app settings
        app_settings = get_app_settings()
        timezone_str = app_settings.timezone
        logger.debug(f"Using app timezone setting: {timezone_str}")

    # Get the user's timezone
    try:
        user_tz = pytz.timezone(timezone_str)
    except pytz.exceptions.UnknownTimeZoneError:
        logger.warning(f"Unknown timezone {timezone_str}, using UTC")
        user_tz = pytz.UTC

    # Handle timestamp (integer or float)
    if isinstance(value, (int, float)):
        try:
            # Convert timestamp to datetime (assume UTC)
            utc_dt = datetime.fromtimestamp(value, tz=timezone.utc)
            # Log the conversion for debugging
            logger.debug(f"Converting timestamp {value} to UTC datetime: {utc_dt}")
            # Convert to user's timezone
            local_dt = utc_dt.astimezone(user_tz)
            logger.debug(f"Converted to user timezone ({timezone_str}): {local_dt}")
            # Format the datetime
            formatted = local_dt.strftime(format)
            logger.debug(f"Formatted datetime: {formatted}")
            return formatted
        except (ValueError, OverflowError) as e:
            logger.error(f"Error converting timestamp {value}: {e}")
            return ''

    # Handle datetime object
    elif value and isinstance(value, datetime):
        try:
            # Assume the datetime is in UTC if it has no timezone
            if value.tzinfo is None:
                value = value.replace(tzinfo=timezone.utc)
                logger.debug(f"Added UTC timezone to naive datetime: {value}")
            # Convert to user's timezone
            local_dt = value.astimezone(user_tz)
            logger.debug(f"Converted to user timezone ({timezone_str}): {local_dt}")
            # Format the datetime
            formatted = local_dt.strftime(format)
            logger.debug(f"Formatted datetime: {formatted}")
            return formatted
        except Exception as e:
            logger.error(f"Error formatting datetime {value}: {e}")
            return str(value)

    # Handle ISO format string
    elif isinstance(value, str) and value:
        try:
            # Parse ISO format string
            dt = datetime.fromisoformat(value.replace('Z', '+00:00'))
            logger.debug(f"Parsed ISO datetime string: {dt}")
            # Convert to user's timezone
            local_dt = dt.astimezone(user_tz)
            logger.debug(f"Converted to user timezone ({timezone_str}): {local_dt}")
            # Format the datetime
            formatted = local_dt.strftime(format)
            logger.debug(f"Formatted datetime: {formatted}")
            return formatted
        except (ValueError, OverflowError) as e:
            logger.error(f"Error parsing ISO datetime string {value}: {e}")
            return value

    return ''

# Add a filter for relative time (e.g., "2 hours ago")
@app.template_filter('relative_time')
def relative_time(value):
    """Format a timestamp as a relative time string"""
    import pytz
    from datetime import timezone

    # Get the current user's timezone preference
    current_user = auth.get_current_user()
    if current_user and 'settings' in current_user and 'timezone' in current_user['settings']:
        timezone_str = current_user['settings']['timezone']
        logger.debug(f"Using user's timezone preference for relative time: {timezone_str}")
    else:
        # Fallback to app settings
        app_settings = get_app_settings()
        timezone_str = app_settings.timezone
        logger.debug(f"Using app timezone setting for relative time: {timezone_str}")

    # Get the user's timezone
    try:
        user_tz = pytz.timezone(timezone_str)
    except pytz.exceptions.UnknownTimeZoneError:
        user_tz = pytz.UTC

    # Handle timestamp (integer or float)
    if isinstance(value, (int, float)):
        try:
            # Convert timestamp to datetime (assume UTC)
            utc_dt = datetime.fromtimestamp(value, tz=timezone.utc)
            # Convert to user's timezone
            local_dt = utc_dt.astimezone(user_tz)
        except (ValueError, OverflowError) as e:
            logger.error(f"Error converting timestamp {value}: {e}")
            return ''
    # Handle datetime object
    elif value and isinstance(value, datetime):
        try:
            # Assume the datetime is in UTC if it has no timezone
            if value.tzinfo is None:
                value = value.replace(tzinfo=timezone.utc)
            # Convert to user's timezone
            local_dt = value.astimezone(user_tz)
        except Exception as e:
            logger.error(f"Error formatting datetime {value}: {e}")
            return str(value)
    else:
        return ''

    # Calculate the time difference
    now = datetime.now(tz=user_tz)
    diff = now - local_dt

    # Format the relative time
    seconds = diff.total_seconds()
    if seconds < 60:
        return 'just now'
    elif seconds < 3600:
        minutes = int(seconds / 60)
        return f"{minutes} minute{'s' if minutes != 1 else ''} ago"
    elif seconds < 86400:
        hours = int(seconds / 3600)
        return f"{hours} hour{'s' if hours != 1 else ''} ago"
    elif seconds < 604800:
        days = int(seconds / 86400)
        return f"{days} day{'s' if days != 1 else ''} ago"
    elif seconds < 2592000:
        weeks = int(seconds / 604800)
        return f"{weeks} week{'s' if weeks != 1 else ''} ago"
    elif seconds < 31536000:
        months = int(seconds / 2592000)
        return f"{months} month{'s' if months != 1 else ''} ago"
    else:
        years = int(seconds / 31536000)
        return f"{years} year{'s' if years != 1 else ''} ago"

# Add security headers
@app.after_request
def add_security_headers(response):
    """Add security headers to each response"""
    # Content Security Policy - allow Bootstrap and other CDN resources
    response.headers['Content-Security-Policy'] = "default-src 'self'; script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https:; font-src 'self' https://cdn.jsdelivr.net; img-src 'self' data:;"
    # X-Content-Type-Options
    response.headers['X-Content-Type-Options'] = 'nosniff'
    # X-Frame-Options
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    # X-XSS-Protection
    response.headers['X-XSS-Protection'] = '1; mode=block'
    # Referrer-Policy
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    return response

# Authentication routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page"""
    # If user is already logged in, redirect to dashboard
    if auth.get_current_user():
        return redirect(url_for('index'))

    # Clear any existing flash messages when loading the login page
    session.pop('_flashes', None)

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        # Validate input
        if not username or not password:
            flash('Please enter both username and password', 'error')
            return render_template('login.html')

        # Get user from database
        user = db.get_user_by_username(username)
        if not user:
            flash('Invalid username or password', 'error')
            return render_template('login.html')

        # Check if user is active
        if not user['is_active']:
            flash('Your account has been deactivated. Please contact an administrator.', 'error')
            return render_template('login.html')

        # Verify password
        password_parts = user['password_hash'].split(':')
        if len(password_parts) != 2:
            flash('Invalid account configuration. Please contact an administrator.', 'error')
            return render_template('login.html')

        stored_hash, salt = password_parts
        if not auth.verify_password(password, stored_hash, salt):
            flash('Invalid username or password', 'error')
            return render_template('login.html')

        # Create session
        session_token = auth.create_user_session(user['id'])

        # Log the login action
        db.log_user_action(
            user_id=user['id'],
            username=user['username'],
            action_type='login',
            resource_type='session',
            details='User logged in',
            ip_address=request.remote_addr
        )

        # Set session cookie
        response = make_response(redirect(url_for('index')))
        response.set_cookie(
            auth.SESSION_COOKIE_NAME,
            session_token,
            max_age=auth.SESSION_EXPIRY,
            httponly=True,
            secure=request.is_secure,
            samesite='Lax'
        )

        # Check if user has any organizations
        user_orgs = db.get_user_organizations(user['id'])
        if not user_orgs and not user['is_admin']:
            # Set session cookie first, then redirect to create first organization page
            response = make_response(redirect(url_for('create_first_organization')))
            response.set_cookie(
                auth.SESSION_COOKIE_NAME,
                session_token,
                max_age=auth.SESSION_EXPIRY,
                httponly=True,
                secure=request.is_secure,
                samesite='Lax'
            )
            return response

        return response

    return render_template('login.html')

@app.route('/signup')
def signup():
    """Signup page - shows information about registration"""
    # If user is already logged in, redirect to dashboard
    if auth.get_current_user():
        return redirect(url_for('index'))

    return render_template('signup.html')

@app.route('/register', methods=['GET', 'POST'])
def register():
    """Registration page - allows users to create an account"""
    # If user is already logged in, redirect to dashboard
    if auth.get_current_user():
        return redirect(url_for('index'))

    # Allow open registration for now - this can be controlled by a setting later
    # This allows users to register from the signup page

    if request.method == 'POST':
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')

        # Validate input
        if not username or not email or not password or not confirm_password:
            flash('Please fill in all fields', 'error')
            return render_template('register.html')

        if password != confirm_password:
            flash('Passwords do not match', 'error')
            return render_template('register.html')

        # Check if username or email already exists
        if db.get_user_by_username(username):
            flash('Username already exists', 'error')
            return render_template('register.html')

        if db.get_user_by_email(email):
            flash('Email already exists', 'error')
            return render_template('register.html')

        # Hash password
        password_hash, salt = auth.hash_password(password)
        combined_hash = f"{password_hash}:{salt}"

        # Create user
        user_id = db.create_user(username, email, combined_hash)
        if not user_id:
            flash('Error creating user', 'error')
            return render_template('register.html')

        # Log the user creation action
        db.log_user_action(
            user_id=user_id,
            username=username,
            action_type='create',
            resource_type='user',
            resource_id=user_id,
            resource_name=username,
            details=f'User self-registered with email {email}',
            ip_address=request.remote_addr
        )

        flash('Account created successfully. You can now log in.', 'success')
        return redirect(url_for('login'))

    return render_template('register.html')

@app.route('/create_first_organization', methods=['GET', 'POST'])
@auth.login_required
def create_first_organization():
    """Create first organization for new users"""
    user = auth.get_current_user()

    # Check if user already has organizations
    user_orgs = db.get_user_organizations(user['id'])
    if user_orgs:
        flash('You already have organizations', 'info')
        return redirect(url_for('index'))

    if request.method == 'POST':
        name = request.form.get('name', '')
        description = request.form.get('description', '')

        if not name:
            flash('Organization name is required', 'error')
            return redirect(url_for('create_first_organization'))

        # Create organization
        org_id = db.create_organization(name, description)
        if not org_id:
            flash('Failed to create organization', 'error')
            return redirect(url_for('create_first_organization'))

        # Add user to organization as admin
        db.add_user_to_organization(user['id'], org_id, 'admin')

        # Set as current organization
        session['current_organization_id'] = org_id

        flash(f'Organization "{name}" created successfully', 'success')

        # Redirect to WHOIS API key setup
        return redirect(url_for('setup_whois_api_key'))

    return render_template('create_first_organization.html', user=user)

@app.route('/setup_whois_api_key', methods=['GET', 'POST'])
@auth.login_required
def setup_whois_api_key():
    """Setup WHOIS API key after creating first organization"""
    user = auth.get_current_user()

    # Make sure user has a current organization
    if not user.get('current_organization'):
        flash('You need to create or join an organization first', 'error')
        return redirect(url_for('create_first_organization'))

    # Get the current organization ID
    org_id = user['current_organization']['id']

    if request.method == 'POST':
        api_key = request.form.get('whois_api_key', '').strip()
        skip = request.form.get('skip', 'false') == 'true'

        if skip:
            flash('You can add a WHOIS API key later in Settings', 'info')
            return redirect(url_for('index'))

        if not api_key:
            flash('Please enter a valid WHOIS API key or skip this step', 'error')
            return redirect(url_for('setup_whois_api_key'))

        # Save the WHOIS API key for this specific organization
        db.set_organization_setting(org_id, 'whois_api_key', api_key)

        flash('WHOIS API key saved successfully for your organization', 'success')
        return redirect(url_for('index'))

    return render_template('setup_whois_api_key.html', user=user)

@app.route('/logout')
def logout():
    """Logout user"""
    # Get current user before logging out
    current_user = auth.get_current_user()

    # Get session token from cookie
    session_token = request.cookies.get(auth.SESSION_COOKIE_NAME)
    if session_token:
        # Delete session from database
        db.delete_session(session_token)

    # Log the logout action if we have user information
    if current_user:
        db.log_user_action(
            user_id=current_user['id'],
            username=current_user['username'],
            action_type='logout',
            resource_type='session',
            details='User logged out',
            ip_address=request.remote_addr
        )

    # Clear session cookie and all session data
    response = make_response(redirect(url_for('login')))
    response.delete_cookie(auth.SESSION_COOKIE_NAME)

    # Store flash message directly in the response
    flash('You have been logged out', 'success')

    # Clear all session data to prevent flash messages from persisting
    session.clear()

    return response

# Define data directory for persistent storage
# Use environment variable if set, otherwise use default location
DATA_DIR = os.environ.get('CERT_MONITOR_DATA_DIR', os.path.expanduser('~/.cert-monitor'))
CONFIG_FILE = os.path.join(DATA_DIR, "config.yaml")
CACHE_FILE = os.path.join(DATA_DIR, "whois_cache.json")

# Cache settings
CACHE_EXPIRY = 86400  # 24 hours in seconds

# Ensure data directory exists
def ensure_data_dir():
    if not os.path.exists(DATA_DIR):
        try:
            os.makedirs(DATA_DIR, exist_ok=True)
            print(f"Created data directory at {DATA_DIR}")
        except Exception as e:
            print(f"Error creating data directory: {e}")
            # Fall back to local directory if we can't create the data directory
            return False
    return True

# Default settings are defined here for reference
DEFAULT_EMAIL_SETTINGS = {
    'smtp_server': 'smtp.office365.com',
    'smtp_port': 587,
    'smtp_username': '',
    'smtp_password': '',
    'notification_email': '',
    'warning_threshold_days': 10
}

DEFAULT_APP_SETTINGS = {
    'auto_refresh_enabled': False,
    'auto_refresh_interval': 5,
    'theme': 'light',
    'timezone': 'UTC',
    'warning_threshold_days': 10
}

DEFAULT_NOTIFICATIONS = {
    'email': {
        'enabled': False,
        'smtp_server': '',
        'smtp_port': '',
        'smtp_username': '',
        'smtp_password': '',
        'from_email': '',
        'notification_email': '',
        'warning_threshold_days': 10
    },
    'teams': {
        'enabled': False,
        'webhook_url': ''
    },
    'slack': {
        'enabled': False,
        'webhook_url': '',
        'channel': ''
    },
    'discord': {
        'enabled': False,
        'webhook_url': '',
        'username': 'Certifly Bot'
    },
    'telegram': {
        'enabled': False,
        'bot_token': '',
        'chat_id': ''
    },
    'webhook': {
        'enabled': False,
        'webhook_url': '',
        'format': 'json',
        'custom_headers': '{}',
        'custom_fields': '{}'
    },
    'sms': {
        'enabled': False,
        'account_sid': '',
        'auth_token': '',
        'from_number': '',
        'to_number': ''
    }
}

# Simplified compatibility functions for code that still uses YAML configuration
def load_config():
    """Compatibility function that loads configuration from database"""
    # Create a config dictionary with all settings from the database
    config = {
        'ssl_domains': db.get_setting('ssl_domains', []),
        'domain_expiry': db.get_setting('domain_expiry', []),
        'ping_hosts': db.get_setting('ping_hosts', []),
        'email_settings': db.get_setting('email_settings', DEFAULT_EMAIL_SETTINGS),
        'app_settings': db.get_setting('app_settings', DEFAULT_APP_SETTINGS),
        'api_settings': db.get_setting('api_settings', {}),
        'notifications': db.get_setting('notifications', DEFAULT_NOTIFICATIONS),
        'acknowledged_alerts': db.get_setting('acknowledged_alerts', []),
        'deleted_alerts': db.get_setting('deleted_alerts', [])
    }
    return config

def save_config(data):
    """Compatibility function that saves configuration to database"""
    # Save each section of the config to the database
    for key in ['ssl_domains', 'domain_expiry', 'ping_hosts', 'email_settings',
                'app_settings', 'api_settings', 'notifications',
                'acknowledged_alerts', 'deleted_alerts']:
        if key in data:
            db.set_setting(key, data[key])

def get_email_settings() -> EmailSettings:
    """Get email settings from database"""
    settings = db.get_setting('email_settings', DEFAULT_EMAIL_SETTINGS)

    # Handle warning_threshold_days safely
    try:
        warning_threshold = int(settings.get('warning_threshold_days', 10))
    except (ValueError, TypeError):
        logger.warning(f"Invalid warning_threshold_days value: {settings.get('warning_threshold_days')}. Using default value of 10.")
        warning_threshold = 10

    # Handle smtp_port safely
    try:
        smtp_port = int(settings.get('smtp_port', 587))
    except (ValueError, TypeError):
        logger.warning(f"Invalid smtp_port value: {settings.get('smtp_port')}. Using default value of 587.")
        smtp_port = 587

    return EmailSettings(
        smtp_server=settings.get('smtp_server', 'smtp.office365.com'),
        smtp_port=smtp_port,
        smtp_username=settings.get('smtp_username', ''),
        smtp_password=settings.get('smtp_password', ''),
        notification_email=settings.get('notification_email', ''),
        warning_threshold_days=warning_threshold
    )

def get_app_settings() -> AppSettings:
    """Get app settings from database"""
    settings = db.get_setting('app_settings', DEFAULT_APP_SETTINGS)

    # Log the raw settings for debugging
    logger.debug(f"Raw app settings from database: {settings}")

    # Handle auto_refresh_interval safely
    try:
        auto_refresh_interval = int(settings.get('auto_refresh_interval', 5))
    except (ValueError, TypeError):
        logger.warning(f"Invalid auto_refresh_interval value: {settings.get('auto_refresh_interval')}. Using default value of 5.")
        auto_refresh_interval = 5

    # Handle warning_threshold_days safely
    try:
        warning_threshold = int(settings.get('warning_threshold_days', 10))
        logger.debug(f"Retrieved warning_threshold_days from database: {warning_threshold}")
    except (ValueError, TypeError):
        logger.warning(f"Invalid warning_threshold_days value: {settings.get('warning_threshold_days')}. Using default value of 10.")
        warning_threshold = 10

    app_settings = AppSettings(
        auto_refresh_enabled=settings.get('auto_refresh_enabled', False),
        auto_refresh_interval=auto_refresh_interval,
        theme=settings.get('theme', 'light'),
        timezone=settings.get('timezone', 'UTC'),
        warning_threshold_days=warning_threshold
    )

    # Log the final app settings object for debugging
    logger.debug(f"Final app settings object: {app_settings}")

    return app_settings

def get_notification_settings() -> NotificationSettings:
    """Get notification settings for all platforms from database"""
    notifications = db.get_setting('notifications', DEFAULT_NOTIFICATIONS)

    # Ensure all notification channels exist with default values
    if 'email' not in notifications:
        notifications['email'] = DEFAULT_NOTIFICATIONS['email']
    if 'teams' not in notifications:
        notifications['teams'] = DEFAULT_NOTIFICATIONS['teams']
    if 'slack' not in notifications:
        notifications['slack'] = DEFAULT_NOTIFICATIONS['slack']
    if 'discord' not in notifications:
        notifications['discord'] = DEFAULT_NOTIFICATIONS['discord']
    if 'telegram' not in notifications:
        notifications['telegram'] = DEFAULT_NOTIFICATIONS['telegram']
    if 'webhook' not in notifications:
        notifications['webhook'] = DEFAULT_NOTIFICATIONS['webhook']
    if 'sms' not in notifications:
        notifications['sms'] = DEFAULT_NOTIFICATIONS['sms']

    return NotificationSettings(
        email=notifications.get('email', {}),
        teams=notifications.get('teams', {}),
        slack=notifications.get('slack', {}),
        discord=notifications.get('discord', {}),
        telegram=notifications.get('telegram', {}),
        webhook=notifications.get('webhook', {}),
        sms=notifications.get('sms', {})
    )

# WHOIS caching functions using database
def get_cached_whois_data(domain):
    """Get WHOIS data from cache if available and not expired"""
    cache_key = f"whois_{domain}"
    cached_data = db.get_cache(cache_key)

    if cached_data:
        logger.debug(f"Using cached WHOIS data for {domain}")
        return cached_data

    return None

def cache_whois_data(domain, data):
    """Cache WHOIS data for a domain"""
    cache_key = f"whois_{domain}"
    db.set_cache(cache_key, data, CACHE_EXPIRY)

# Email notification function removed - replaced by the notification system in notifications.py

def convert_to_user_timezone(dt, timezone_str='UTC'):
    """Convert a datetime object from UTC to the user's timezone"""
    if dt is None:
        return None

    try:
        import pytz
        # Validate the timezone
        try:
            user_tz = pytz.timezone(timezone_str)
            logger.debug(f"Using timezone: {timezone_str}")
        except pytz.exceptions.UnknownTimeZoneError:
            logger.warning(f"Unknown timezone: {timezone_str}. Using UTC instead.")
            user_tz = pytz.UTC
            timezone_str = 'UTC'

        # Ensure the datetime is timezone-aware (UTC)
        if dt.tzinfo is None:
            logger.debug(f"Making naive datetime {dt} timezone-aware (UTC)")
            dt = dt.replace(tzinfo=pytz.UTC)
        elif dt.tzinfo != pytz.UTC:
            logger.debug(f"Converting datetime {dt} with tzinfo {dt.tzinfo} to UTC")
            dt = dt.astimezone(pytz.UTC)

        # Convert to the user's timezone
        logger.debug(f"Converting datetime {dt} from UTC to {timezone_str}")
        converted_dt = dt.astimezone(user_tz)
        logger.debug(f"Converted datetime: {converted_dt}")

        return converted_dt
    except ImportError:
        # If pytz is not available, return the original datetime
        logger.warning("pytz not available, timezone conversion skipped")
        return dt
    except Exception as e:
        # If there's any error, return the original datetime
        logger.error(f"Error converting timezone: {str(e)}")
        return dt


def get_whois_api_key():
    """Get the WHOIS API key from database or environment variable"""
    # First check environment variable
    api_key = os.environ.get('WHOIS_API_KEY')
    if api_key:
        return api_key

    # Get current user and organization
    user = auth.get_current_user()
    if user and user.get('current_organization'):
        # Get organization-specific API key
        org_id = user['current_organization']['id']
        api_key = db.get_organization_setting(org_id, 'whois_api_key', '')

        if api_key:
            logger.debug(f"WHOIS API key found for organization {org_id}")
            return api_key

    # If no organization-specific key, check global settings as fallback
    api_settings = db.get_setting('api_settings', {})
    api_key = api_settings.get('whois_api_key', '')

    # Log the API key status (without revealing the actual key)
    if api_key:
        logger.debug("WHOIS API key found in global database settings")
    else:
        logger.debug("No WHOIS API key found in settings")

    return api_key

def get_whois_data(domain: str):
    """Get WHOIS data for a domain using a public API with caching"""
    # Check if we have valid cached data
    cached_data = get_cached_whois_data(domain)
    if cached_data:
        logger.debug(f"Using cached WHOIS data for {domain}")
        return cached_data

    logger.info(f"Fetching fresh WHOIS data for {domain}")

    # Get API key
    api_key = get_whois_api_key()
    if not api_key:
        # Try to get API key directly from database as a fallback
        api_settings = db.get_setting('api_settings', {})
        api_key = api_settings.get('whois_api_key', '')

        if api_key:
            logger.info("Retrieved WHOIS API key directly from database")
        else:
            logger.error("No WHOIS API key configured. Please set WHOIS_API_KEY environment variable or configure it in settings.")
            return None

    try:
        # Use WHOIS API to get domain information
        url = f"https://www.whoisxmlapi.com/whoisserver/WhoisService?apiKey={api_key}&domainName={domain}&outputFormat=json"

        # Create a request with a user agent to avoid being blocked
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }

        logger.debug(f"Making WHOIS API request for {domain}")

        # Use requests library for better error handling
        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()  # Raise exception for 4XX/5XX status codes

        data = response.json()
        logger.debug(f"Received WHOIS API response for {domain}")

        # Extract the relevant information
        whois_record = data.get('WhoisRecord', {})
        registrar_data = whois_record.get('registrarName', 'Unknown')

        # Try to get expiry date
        expiry_date_str = None
        registry_data = whois_record.get('registryData', {})
        if registry_data:
            expiry_date_str = registry_data.get('expiresDate')
            logger.debug(f"Found expiry date in registry data for {domain}: {expiry_date_str}")

        if not expiry_date_str and 'registryData' in whois_record:
            # Try to find expiry date in raw text
            raw_text = whois_record.get('rawText', '')
            expiry_matches = re.findall(r'Expir(?:y|ation) Date:\s*(.+)', raw_text, re.IGNORECASE)
            if expiry_matches:
                expiry_date_str = expiry_matches[0].strip()
                logger.debug(f"Found expiry date in raw text for {domain}: {expiry_date_str}")

        result = {
            'registrar': registrar_data,
            'expiry_date_str': expiry_date_str
        }

        # Cache the result
        if expiry_date_str:  # Only cache successful results
            logger.debug(f"Caching WHOIS data for {domain}")
            cache_whois_data(domain, result)
        else:
            logger.warning(f"No expiry date found for {domain}, not caching result")

        return result
    except requests.exceptions.RequestException as e:
        logger.error(f"Error fetching WHOIS data for {domain}: {str(e)}")
        return None
    except (json.JSONDecodeError, ValueError) as e:
        logger.error(f"Error parsing WHOIS data for {domain}: {str(e)}")
        return None
    except Exception as e:
        logger.error(f"Unexpected error fetching WHOIS data for {domain}: {str(e)}", exc_info=True)
        return None

# Domain expiry cache - using database cache
DOMAIN_EXPIRY_CACHE_EXPIRY = 86400  # 24 hours in seconds (same as WHOIS cache)
# Removed global DOMAIN_EXPIRY_CACHE variable - using database cache instead

def get_cached_domain_expiry_data(domain):
    """Get domain expiry data from cache if available and not expired"""
    cache_key = f"domain_expiry_{domain}"
    cached_data = db.get_cache(cache_key)

    if cached_data:
        logger.debug(f"Using cached domain expiry data for {domain}")

        # Convert expiry_date string back to datetime if needed
        if 'expiry_date' in cached_data and isinstance(cached_data['expiry_date'], str):
            try:
                cached_data['expiry_date'] = datetime.fromisoformat(cached_data['expiry_date'])
            except ValueError:
                logger.warning(f"Could not parse expiry_date from cache for {domain}")
                return None

        # Ensure days_remaining is an integer
        if 'days_remaining' in cached_data:
            try:
                cached_data['days_remaining'] = int(cached_data['days_remaining'])
            except (ValueError, TypeError):
                logger.warning(f"Invalid days_remaining in cache for {domain}: {cached_data.get('days_remaining')}. Setting to -1.")
                cached_data['days_remaining'] = -1

        return cached_data

    return None

def cache_domain_expiry_data(domain, domain_status):
    """Cache domain expiry data for a domain"""
    cache_key = f"domain_expiry_{domain}"

    # Validate domain_status object
    if not domain_status:
        logger.error(f"Cannot cache domain expiry data for {domain}: domain_status is None")
        return

    # Ensure days_remaining is an integer
    try:
        days_remaining = int(domain_status.days_remaining)
    except (ValueError, TypeError):
        logger.warning(f"Invalid days_remaining for {domain}: {domain_status.days_remaining}. Using -1.")
        days_remaining = -1

    # Validate expiry_date
    if not isinstance(domain_status.expiry_date, datetime):
        logger.warning(f"Invalid expiry_date for {domain}: {domain_status.expiry_date}. Using current time.")
        expiry_date = datetime.now()
    else:
        expiry_date = domain_status.expiry_date

    # Validate registrar
    registrar = str(domain_status.registrar) if domain_status.registrar else "Unknown"

    # Validate status
    valid_statuses = ['valid', 'warning', 'expired', 'error']
    status = domain_status.status if domain_status.status in valid_statuses else 'error'

    # Convert datetime to string for JSON serialization
    expiry_date_str = expiry_date.isoformat() if expiry_date else None

    # Prepare data for caching
    cache_data = {
        'days_remaining': days_remaining,
        'expiry_date': expiry_date_str,  # Store as ISO format string
        'registrar': registrar,
        'status': status
    }

    # Log what we're caching
    logger.debug(f"Caching domain expiry data for {domain}: days_remaining={days_remaining}, " +
                f"expiry_date={expiry_date_str}, registrar={registrar}, status={status}")

    # Cache the data
    db.set_cache(cache_key, cache_data, DOMAIN_EXPIRY_CACHE_EXPIRY)

def check_domain_expiry(domain: str) -> DomainStatus:
    """Check domain expiry date using whois with caching"""
    logger.debug(f"Checking domain expiry for {domain}")

    # Default values for error case
    default_days_remaining = -1
    default_expiry_date = datetime.now()
    default_registrar = "Unknown"
    default_status = 'error'

    # Step 1: Try to use cached data first
    try:
        cached_data = get_cached_domain_expiry_data(domain)
        if cached_data:
            logger.debug(f"Using cached domain expiry data for {domain}")

            # Safely extract and convert days_remaining to integer
            try:
                days_remaining = int(cached_data.get('days_remaining', default_days_remaining))
            except (ValueError, TypeError):
                logger.warning(f"Invalid days_remaining in cache for {domain}: {cached_data.get('days_remaining')}. Using default.")
                days_remaining = default_days_remaining

            # Get other values with defaults
            expiry_date = cached_data.get('expiry_date', default_expiry_date)
            registrar = cached_data.get('registrar', default_registrar)
            status = cached_data.get('status', default_status)

            # Always get fresh ping status
            ping_result = check_ping(domain)

            # Return domain status from cache
            return DomainStatus(
                name=domain,
                days_remaining=days_remaining,
                expiry_date=expiry_date,
                registrar=registrar,
                status=status,
                ping_status=ping_result.get("status", "unknown")
            )
    except Exception as e:
        logger.error(f"Error using cached domain expiry data for {domain}: {str(e)}", exc_info=True)
        # Continue to fresh check

    # Step 2: Get fresh WHOIS data
    try:
        # Get WHOIS data
        whois_data = get_whois_data(domain)
        if not whois_data:
            logger.error(f"Failed to get WHOIS data for {domain}")
            raise ValueError("No WHOIS data available")

        # Check if we have an expiry date string
        expiry_date_str = whois_data.get('expiry_date_str')
        if not expiry_date_str:
            logger.error(f"No expiry date found in WHOIS data for {domain}")
            raise ValueError("No expiry date in WHOIS data")

        # Try to parse the expiry date
        expiry_date = None
        date_formats = [
            '%Y-%m-%dT%H:%M:%SZ',      # ISO format
            '%Y-%m-%dT%H:%M:%S.%fZ',   # ISO format with microseconds
            '%Y-%m-%dT%H:%M:%S%z',     # ISO format with timezone
            '%Y-%m-%d',                # Simple date
            '%d-%b-%Y',                # 01-Jan-2023
            '%d %b %Y',                # 01 Jan 2023
            '%Y.%m.%d',                # 2023.01.01
            '%d.%m.%Y',                # 01.01.2023
            '%Y/%m/%d',                # 2023/01/01
            '%d/%m/%Y',                # 01/01/2023
            '%b %d %Y',                # Jan 01 2023
            '%B %d %Y',                # January 01 2023
            '%d-%B-%Y',                # 01-January-2023
            '%Y-%m-%d %H:%M:%S',       # 2023-01-01 12:00:00
        ]

        logger.debug(f"Trying to parse expiry date string for {domain}: '{expiry_date_str}'")

        for date_format in date_formats:
            try:
                expiry_date = datetime.strptime(expiry_date_str, date_format)
                logger.debug(f"Successfully parsed expiry date for {domain} using format {date_format}: {expiry_date}")
                break
            except ValueError:
                continue

        if not expiry_date:
            logger.error(f"Could not parse expiry date string for {domain}: '{expiry_date_str}'")
            raise ValueError(f"Could not parse expiry date: {expiry_date_str}")

        # Calculate days remaining
        days_remaining = (expiry_date - datetime.now()).days

        # Ensure days_remaining is an integer
        days_remaining = int(days_remaining)

        # Get registrar info
        registrar = whois_data.get('registrar', default_registrar)

        # Get warning threshold from app settings - always get fresh settings
        app_settings = get_app_settings()
        try:
            warning_threshold = int(app_settings.warning_threshold_days)
            logger.debug(f"Using warning threshold of {warning_threshold} days for domain expiry check of {domain}")
        except (ValueError, TypeError):
            logger.warning(f"Invalid warning_threshold_days: {app_settings.warning_threshold_days}. Using default value of 10.")
            warning_threshold = 10

        # Determine status
        if days_remaining <= 0:
            status = 'expired'
            logger.warning(f"Domain {domain} has expired")
        elif days_remaining <= warning_threshold:
            status = 'warning'
            logger.warning(f"Domain {domain} will expire in {days_remaining} days (warning threshold: {warning_threshold} days)")
        else:
            status = 'valid'
            logger.debug(f"Domain {domain} is valid, will expire in {days_remaining} days (warning threshold: {warning_threshold} days)")

        # Get ping status
        ping_result = check_ping(domain)

        # Create domain status object
        domain_status = DomainStatus(
            name=domain,
            days_remaining=days_remaining,
            expiry_date=expiry_date,
            registrar=registrar,
            status=status,
            ping_status=ping_result.get("status", "unknown")
        )

        # Cache the successful result
        try:
            cache_domain_expiry_data(domain, domain_status)
        except Exception as cache_error:
            logger.error(f"Error caching domain expiry data for {domain}: {str(cache_error)}")
            # Continue even if caching fails

        # Send notifications if domain is in warning or expired state
        if status in ['warning', 'expired']:
            try:
                notification_settings = get_notification_settings()

                # Send notifications to all enabled platforms
                if notification_settings.email.get('enabled', False):
                    notifications_module.send_domain_expiry_notification('email', notification_settings.email, domain_status)

                if notification_settings.teams.get('enabled', False):
                    notifications_module.send_domain_expiry_notification('teams', notification_settings.teams, domain_status)

                if notification_settings.slack.get('enabled', False):
                    notifications_module.send_domain_expiry_notification('slack', notification_settings.slack, domain_status)

                if notification_settings.discord.get('enabled', False):
                    notifications_module.send_domain_expiry_notification('discord', notification_settings.discord, domain_status)

                if notification_settings.telegram.get('enabled', False):
                    notifications_module.send_domain_expiry_notification('telegram', notification_settings.telegram, domain_status)

                if notification_settings.webhook.get('enabled', False):
                    notifications_module.send_domain_expiry_notification('webhook', notification_settings.webhook, domain_status)

                if notification_settings.sms.get('enabled', False):
                    notifications_module.send_domain_expiry_notification('sms', notification_settings.sms, domain_status)
            except Exception as notify_error:
                logger.error(f"Error sending notifications for {domain}: {str(notify_error)}")
                # Continue even if notification fails

        return domain_status

    except Exception as e:
        # Handle any errors in the WHOIS data processing
        logger.error(f"Error checking domain expiry for {domain}: {str(e)}", exc_info=True)

        # Get ping status if possible
        try:
            ping_result = check_ping(domain)
            ping_status = ping_result.get("status", "unknown")
        except Exception:
            ping_status = "unknown"

        # Create error status
        error_status = DomainStatus(
            name=domain,
            days_remaining=default_days_remaining,
            expiry_date=default_expiry_date,
            registrar=f"{default_registrar} (Error: {str(e)[:50]}...)" if len(str(e)) > 50 else f"{default_registrar} (Error: {str(e)})",
            status=default_status,
            ping_status=ping_status
        )

        # Cache the error result with a shorter expiry time (1 hour)
        try:
            cache_key = f"domain_expiry_{domain}"
            # Convert datetime to string for JSON serialization
            expiry_date_str = default_expiry_date.isoformat() if default_expiry_date else None

            cache_data = {
                'days_remaining': default_days_remaining,
                'expiry_date': expiry_date_str,  # Store as ISO format string
                'registrar': f"{default_registrar} (Error retrieving data)",
                'status': default_status
            }
            # Cache for 1 hour instead of 24 hours
            db.set_cache(cache_key, cache_data, 3600)
        except Exception as cache_error:
            logger.error(f"Error caching error status for {domain}: {str(cache_error)}")
            # Continue even if caching fails

        return error_status



# SSL Certificate cache - using database cache
SSL_CACHE_EXPIRY = 3600  # 1 hour in seconds

def get_cached_ssl_data(domain):
    """Get SSL certificate data from cache if available and not expired"""
    cache_key = f"ssl_{domain}"
    cached_data = db.get_cache(cache_key)

    if cached_data:
        logger.debug(f"Using cached SSL data for {domain}")

        # Convert expiry_date string back to datetime if needed
        if 'expiry_date' in cached_data and isinstance(cached_data['expiry_date'], str) and cached_data['expiry_date']:
            try:
                cached_data['expiry_date'] = datetime.fromisoformat(cached_data['expiry_date'])
                logger.debug(f"Successfully converted expiry_date string to datetime for {domain}")
            except ValueError as e:
                logger.warning(f"Could not parse expiry_date from cache for {domain}: {str(e)}")
                return None
            except Exception as e:
                logger.warning(f"Unexpected error parsing expiry_date for {domain}: {str(e)}")
                return None

        return cached_data

    return None

def update_ssl_status(domain, warning_threshold):
    """Update the status of cached SSL certificate data based on the new warning threshold"""
    cache_key = f"ssl_{domain}"
    cached_data = db.get_cache(cache_key)

    if not cached_data:
        logger.debug(f"No cached SSL data found for {domain}")
        return False

    # Get days remaining
    try:
        days_remaining = int(cached_data.get('days_remaining', -1))
    except (ValueError, TypeError):
        logger.warning(f"Invalid days_remaining in SSL cache for {domain}: {cached_data.get('days_remaining')}. Cannot update status.")
        return False

    # Calculate new status based on days remaining and new warning threshold
    old_status = cached_data.get('status')

    # Only update status if it's not an error
    if old_status != 'error':
        if days_remaining <= 0:
            new_status = 'expired'
        elif days_remaining <= warning_threshold:
            new_status = 'warning'
        else:
            new_status = 'valid'

        # Always update the cache with the new status, even if it hasn't changed
        # This ensures the warning threshold is consistently applied
        logger.info(f"Updating SSL status for {domain} from {old_status} to {new_status} (days remaining: {days_remaining}, new threshold: {warning_threshold})")
        cached_data['status'] = new_status
        db.set_cache(cache_key, cached_data, SSL_CACHE_EXPIRY)

        # Return True if the status changed
        return old_status != new_status

    return False

def update_domain_expiry_status(domain, warning_threshold):
    """Update the status of cached domain expiry data based on the new warning threshold"""
    cache_key = f"domain_expiry_{domain}"
    cached_data = db.get_cache(cache_key)

    if not cached_data:
        logger.debug(f"No cached domain expiry data found for {domain}")
        return False

    # Get days remaining
    try:
        days_remaining = int(cached_data.get('days_remaining', -1))
    except (ValueError, TypeError):
        logger.warning(f"Invalid days_remaining in domain expiry cache for {domain}: {cached_data.get('days_remaining')}. Cannot update status.")
        return False

    # Calculate new status based on days remaining and new warning threshold
    old_status = cached_data.get('status')

    # Only update status if it's not an error
    if old_status != 'error':
        if days_remaining <= 0:
            new_status = 'expired'
        elif days_remaining <= warning_threshold:
            new_status = 'warning'
        else:
            new_status = 'valid'

        # Always update the cache with the new status, even if it hasn't changed
        # This ensures the warning threshold is consistently applied
        logger.info(f"Updating domain expiry status for {domain} from {old_status} to {new_status} (days remaining: {days_remaining}, new threshold: {warning_threshold})")
        cached_data['status'] = new_status
        db.set_cache(cache_key, cached_data, DOMAIN_EXPIRY_CACHE_EXPIRY)

        # Return True if the status changed
        return old_status != new_status

    return False

def cache_ssl_data(domain, cert_status):
    """Cache SSL certificate data for a domain"""
    cache_key = f"ssl_{domain}"

    # Convert datetime to string for JSON serialization
    expiry_date_str = cert_status.expiry_date.isoformat() if cert_status.expiry_date else None

    # Prepare data for caching
    cache_data = {
        'days_remaining': cert_status.days_remaining,
        'expiry_date': expiry_date_str,  # Store as ISO format string
        'status': cert_status.status
    }

    # Log what we're caching
    logger.debug(f"Caching SSL data for {domain}: days_remaining={cert_status.days_remaining}, " +
                f"expiry_date={expiry_date_str}, status={cert_status.status}")

    # Cache the data
    db.set_cache(cache_key, cache_data, SSL_CACHE_EXPIRY)

def check_certificate(domain: str) -> CertificateStatus:
    """Check SSL certificate expiry date using Python's ssl module with caching"""
    logger.info(f"Checking SSL certificate for {domain}")

    # Check if we have valid cached data
    cached_data = get_cached_ssl_data(domain)
    if cached_data:
        logger.info(f"Using cached SSL data for {domain}: {cached_data}")

        # Check ping status (always fresh)
        ping_result = check_ping(domain)

        # Ensure the status is not 'unknown' or 'checking'
        status = cached_data['status']
        if status not in ['valid', 'warning', 'expired', 'error']:
            logger.warning(f"Invalid status '{status}' in cached data for {domain}, forcing a fresh check")
            # Clear the cache and continue to perform a fresh check
            db.clear_cache(f"ssl_{domain}")
        else:
            # Return the valid cached data
            return CertificateStatus(
                domain=domain,
                days_remaining=cached_data['days_remaining'],
                expiry_date=cached_data['expiry_date'],
                status=cached_data['status'],
                ping_status=ping_result["status"]
            )

    try:
        import ssl
        import socket
        from datetime import datetime

        logger.info(f"Starting SSL check for {domain}")

        # Create a context with default verification options
        context = ssl.create_default_context()
        # Set minimum TLS version to TLSv1.2 for security
        context.minimum_version = ssl.TLSVersion.TLSv1_2

        logger.info(f"Created SSL context for {domain} with minimum TLS version set to TLSv1.2")

        # Connect to the server with a shorter timeout (2 seconds instead of 5)
        logger.info(f"Attempting to connect to {domain}:443 with 2 second timeout")
        with socket.create_connection((domain, 443), timeout=2) as sock:
            logger.info(f"Successfully connected to {domain}:443")
            # Set a timeout on the SSL handshake as well
            sock.settimeout(2)
            with context.wrap_socket(sock, server_hostname=domain) as ssock:
                logger.info(f"Successfully established SSL connection to {domain}")

                # Get the certificate
                cert = ssock.getpeercert()
                logger.info(f"Retrieved certificate for {domain}: {cert}")

                # Extract expiry date
                expiry_date_tuple = cert['notAfter']
                logger.info(f"Certificate notAfter date: {expiry_date_tuple}")

                # Format: 'May 30 00:00:00 2023 GMT'
                expiry_date = datetime.strptime(expiry_date_tuple, '%b %d %H:%M:%S %Y %Z')
                days_remaining = (expiry_date - datetime.now()).days

                logger.info(f"Certificate for {domain} expires on {expiry_date}, {days_remaining} days remaining")

                # Get warning threshold from app settings - always get fresh settings
                app_settings = get_app_settings()

                # Ensure warning_threshold_days is an integer
                try:
                    warning_threshold = int(app_settings.warning_threshold_days)
                    logger.debug(f"Using warning threshold of {warning_threshold} days for SSL certificate check of {domain}")
                except (ValueError, TypeError):
                    logger.warning(f"Invalid warning_threshold_days: {app_settings.warning_threshold_days}. Using default value of 10.")
                    warning_threshold = 10

                status = 'valid'
                if days_remaining <= 0:
                    status = 'expired'
                    logger.warning(f"Certificate for {domain} has expired")
                elif days_remaining <= warning_threshold:
                    status = 'warning'
                    logger.warning(f"Certificate for {domain} will expire in {days_remaining} days (warning threshold: {warning_threshold} days)")

                # Check ping status
                ping_result = check_ping(domain)

                cert_status = CertificateStatus(
                    domain=domain,
                    days_remaining=days_remaining,
                    expiry_date=expiry_date,
                    status=status,
                    ping_status=ping_result["status"]
                )

                # Cache the successful result
                cache_ssl_data(domain, cert_status)

                # Send notifications if certificate is in warning or expired state
                if status in ['warning', 'expired']:
                    # Get notification settings
                    notification_settings = get_notification_settings()

                    # Check if we should send a notification based on previous notifications
                    should_notify = db.should_send_notification(domain, 'ssl', status)

                    if should_notify:
                        logger.info(f"Sending notifications for {domain} certificate ({status})")
                        notifications_sent = False

                        # Send notifications to all enabled platforms
                        if notification_settings.email.get('enabled', False):
                            logger.debug(f"Sending email notification for {domain}")
                            success, message = notifications_module.send_certificate_expiry_notification('email', notification_settings.email, cert_status)
                            if success:
                                notifications_sent = True
                                logger.info(f"Email notification sent for {domain}: {message}")
                            else:
                                logger.error(f"Failed to send email notification for {domain}: {message}")

                        if notification_settings.teams.get('enabled', False):
                            logger.debug(f"Sending Teams notification for {domain}")
                            success, message = notifications_module.send_certificate_expiry_notification('teams', notification_settings.teams, cert_status)
                            if success:
                                notifications_sent = True
                                logger.info(f"Teams notification sent for {domain}: {message}")
                            else:
                                logger.error(f"Failed to send Teams notification for {domain}: {message}")

                        if notification_settings.slack.get('enabled', False):
                            logger.debug(f"Sending Slack notification for {domain}")
                            success, message = notifications_module.send_certificate_expiry_notification('slack', notification_settings.slack, cert_status)
                            if success:
                                notifications_sent = True
                                logger.info(f"Slack notification sent for {domain}: {message}")
                            else:
                                logger.error(f"Failed to send Slack notification for {domain}: {message}")

                        if notification_settings.discord.get('enabled', False):
                            logger.debug(f"Sending Discord notification for {domain}")
                            success, message = notifications_module.send_certificate_expiry_notification('discord', notification_settings.discord, cert_status)
                            if success:
                                notifications_sent = True
                                logger.info(f"Discord notification sent for {domain}: {message}")
                            else:
                                logger.error(f"Failed to send Discord notification for {domain}: {message}")

                        if notification_settings.telegram.get('enabled', False):
                            logger.debug(f"Sending Telegram notification for {domain}")
                            success, message = notifications_module.send_certificate_expiry_notification('telegram', notification_settings.telegram, cert_status)
                            if success:
                                notifications_sent = True
                                logger.info(f"Telegram notification sent for {domain}: {message}")
                            else:
                                logger.error(f"Failed to send Telegram notification for {domain}: {message}")

                        if notification_settings.webhook.get('enabled', False):
                            logger.debug(f"Sending Webhook notification for {domain}")
                            success, message = notifications_module.send_certificate_expiry_notification('webhook', notification_settings.webhook, cert_status)
                            if success:
                                notifications_sent = True
                                logger.info(f"Webhook notification sent for {domain}: {message}")
                            else:
                                logger.error(f"Failed to send Webhook notification for {domain}: {message}")

                        if notification_settings.sms.get('enabled', False):
                            logger.debug(f"Sending SMS notification for {domain}")
                            success, message = notifications_module.send_certificate_expiry_notification('sms', notification_settings.sms, cert_status)
                            if success:
                                notifications_sent = True
                                logger.info(f"SMS notification sent for {domain}: {message}")
                            else:
                                logger.error(f"Failed to send SMS notification for {domain}: {message}")

                        # Log the notification if any were sent successfully
                        if notifications_sent:
                            db.log_notification_sent(domain, 'ssl', status)
                            logger.info(f"Logged notification for {domain} SSL certificate ({status})")
                    else:
                        logger.info(f"Skipping notification for {domain} SSL certificate ({status}) - already sent recently")

                return cert_status
    except ssl.SSLError as e:
        logger.error(f"SSL Error checking certificate for {domain}: {str(e)}", exc_info=True)
        logger.info(f"Attempting to ping {domain} after SSL error")
        ping_result = check_ping(domain)
        logger.info(f"Ping result for {domain} after SSL error: {ping_result}")

        error_status = CertificateStatus(
            domain=domain,
            days_remaining=-1,
            expiry_date=datetime.now(),
            status='error',
            ping_status=ping_result["status"]
        )

        # Cache the error result to prevent repeated failures
        cache_ssl_data(domain, error_status)

        return error_status
    except socket.gaierror as e:
        logger.error(f"DNS resolution error for {domain}: {str(e)}", exc_info=True)
        logger.info(f"Attempting to ping {domain} after DNS error")
        ping_result = check_ping(domain)
        logger.info(f"Ping result for {domain} after DNS error: {ping_result}")

        error_status = CertificateStatus(
            domain=domain,
            days_remaining=-1,
            expiry_date=datetime.now(),
            status='error',
            ping_status=ping_result["status"]
        )

        # Cache the error result to prevent repeated failures
        cache_ssl_data(domain, error_status)

        return error_status
    except socket.timeout as e:
        logger.error(f"Connection timeout for {domain}: {str(e)}", exc_info=True)
        logger.info(f"Attempting to ping {domain} after timeout")
        ping_result = check_ping(domain)
        logger.info(f"Ping result for {domain} after timeout: {ping_result}")

        error_status = CertificateStatus(
            domain=domain,
            days_remaining=-1,
            expiry_date=datetime.now(),
            status='error',
            ping_status=ping_result["status"]
        )

        # Cache the error result to prevent repeated failures
        cache_ssl_data(domain, error_status)

        return error_status
    except Exception as e:
        logger.error(f"Unexpected error checking certificate for {domain}: {str(e)}", exc_info=True)
        logger.info(f"Attempting to ping {domain} after unexpected error")
        # Even if certificate check fails, try to ping the domain
        ping_result = check_ping(domain)
        logger.info(f"Ping result for {domain} after unexpected error: {ping_result}")

        error_status = CertificateStatus(
            domain=domain,
            days_remaining=-1,
            expiry_date=datetime.now(),
            status='error',
            ping_status=ping_result["status"]
        )

        # Cache the error result to prevent repeated failures
        cache_ssl_data(domain, error_status)

        return error_status

def validate_domain(domain):
    """Validate that the domain is a valid domain name or IP address."""
    domain_regex = re.compile(
        r"^(?:[a-zA-Z0-9-]{1,63}\.)+[a-zA-Z]{2,63}$"  # Domain name
        r"|^(?:\d{1,3}\.){3}\d{1,3}$"  # IPv4 address
    )
    if not domain_regex.match(domain):
        raise ValueError(f"Invalid domain: {domain}")
    return domain
    """Clean a domain name by removing protocol, trailing slashes, and paths"""
    if not domain:
        return ""

    # Clean the domain (remove http://, https://, trailing slashes)
    domain = domain.lower().strip()
    if domain.startswith('http://'):
        domain = domain[7:]
    elif domain.startswith('https://'):
        domain = domain[8:]

    # Remove any path after domain
    domain = domain.split('/')[0]

    return domain

def is_valid_domain(domain):
    """Check if a domain name is valid"""
    if not domain:
        return False

    # Clean the domain first
    domain = clean_domain_name(domain)

    # Simple regex for domain validation
    pattern = r'^([a-zA-Z0-9]([a-zA-Z0-9\-]{0,61}[a-zA-Z0-9])?\.)+[a-zA-Z]{2,}$'

    # Allow IP addresses as well
    ip_pattern = r'^(\d{1,3}\.){3}\d{1,3}$'

    return bool(re.match(pattern, domain) or re.match(ip_pattern, domain))

def get_ssl_stats(certificates):
    """Calculate SSL certificate statistics"""
    total = len(certificates)
    valid = sum(1 for cert in certificates if cert.status == 'valid')
    warning = sum(1 for cert in certificates if cert.status == 'warning')
    expired = sum(1 for cert in certificates if cert.status == 'expired')
    error = sum(1 for cert in certificates if cert.status == 'error')

    return Stats(
        total=total,
        valid=valid,
        warning=warning,
        expired=expired,
        error=error
    )

def get_domain_stats(domains):
    """Calculate domain expiry statistics"""
    total = len(domains)
    valid = sum(1 for domain in domains if domain.status == 'valid')
    warning = sum(1 for domain in domains if domain.status == 'warning')
    expired = sum(1 for domain in domains if domain.status == 'expired')
    error = sum(1 for domain in domains if domain.status == 'error')

    return Stats(
        total=total,
        valid=valid,
        warning=warning,
        expired=expired,
        error=error
    )

def get_ping_hosts():
    """Get all ping hosts from config"""
    config = load_config()
    return config.get('ping_hosts', [])

def check_ping_hosts():
    """Check all ping hosts and return their status"""
    hosts = get_ping_hosts()
    results = []

    for host_entry in hosts:
        host = host_entry.get('host', '')
        if not host:
            continue

        ping_result = check_ping(host)

        # Get response history from config
        response_history = host_entry.get('response_history', [])

        # Add current response time to history if up
        if ping_result["status"] == "up":
            if 'response_history' not in host_entry:
                host_entry['response_history'] = []

            host_entry['response_history'].append(ping_result["response_time"])
            # Keep only the last 20 entries
            host_entry['response_history'] = host_entry['response_history'][-20:]
            response_history = host_entry['response_history']

        results.append(PingStatus(
            host=host,
            status=ping_result["status"],
            last_checked=datetime.now(),
            response_time=ping_result["response_time"],
            response_history=response_history
        ))

    return results

def get_ping_stats(ping_results):
    """Calculate ping statistics"""
    total = len(ping_results)
    up = sum(1 for result in ping_results if result['status'] == 'up')
    down = sum(1 for result in ping_results if result['status'] == 'down')
    unknown = sum(1 for result in ping_results if result['status'] == 'unknown')

    return PingStats(
        total=total,
        up=up,
        down=down,
        unknown=unknown
    )

def get_recent_alerts(certificates, domains, limit=5):
    """Get recent alerts from both SSL certificates and domain expiry"""
    alerts = []

    # Add SSL certificate alerts
    for cert in certificates:
        if cert.status in ['warning', 'expired', 'error']:
            message = ""
            if cert.status == 'warning':
                message = f"Certificate will expire in {cert.days_remaining} days"
            elif cert.status == 'expired':
                message = "Certificate has expired"
            else:
                message = "Error checking certificate"

            alerts.append(Alert(
                domain=cert.domain,
                type='ssl',
                status=cert.status,
                message=message,
                date=datetime.now().strftime('%Y-%m-%d')
            ))

    # Add domain expiry alerts
    for domain in domains:
        if domain.status in ['warning', 'expired', 'error']:
            message = ""
            if domain.status == 'warning':
                message = f"Domain will expire in {domain.days_remaining} days"
            elif domain.status == 'expired':
                message = "Domain has expired"
            else:
                message = "Error checking domain"

            alerts.append(Alert(
                domain=domain.name,
                type='domain',
                status=domain.status,
                message=message,
                date=datetime.now().strftime('%Y-%m-%d')
            ))

    # Sort by status severity (expired, warning, error) and limit
    status_priority = {'expired': 0, 'warning': 1, 'error': 2}
    alerts.sort(key=lambda x: status_priority.get(x.status, 3))

    return alerts[:limit]

def generate_alert_id(alert_type, domain, status):
    """Generate a unique ID for an alert based on its properties"""
    return f"{alert_type.lower()}:{domain}:{status.lower()}"

def get_alert_details(alert_id):
    """Get details for an alert by its ID"""
    config = load_config()

    # Get alert details
    alert_details = None

    # Check SSL certificate alerts
    for entry in config.get('ssl_domains', []):
        domain_name = entry.get('url')
        cert_status = check_certificate(domain_name)

        if cert_status.status in ['warning', 'expired', 'error']:
            temp_alert_id = generate_alert_id('SSL', domain_name, cert_status.status)
            if temp_alert_id == alert_id:
                alert_details = {
                    'id': alert_id,
                    'type': 'SSL',
                    'domain': domain_name,
                    'status': cert_status.status,
                    'message': f'SSL certificate for {domain_name} will expire in {cert_status.days_remaining} days' if cert_status.status == 'warning' else
                              f'SSL certificate for {domain_name} has expired' if cert_status.status == 'expired' else
                              f'Error checking SSL certificate for {domain_name}'
                }
                break

    # If not found, check domain expiry alerts
    if not alert_details:
        for entry in config.get('domain_expiry', []):
            domain_name = entry.get('name')
            domain_status = check_domain_expiry(domain_name)

            if domain_status.status in ['warning', 'expired', 'error']:
                temp_alert_id = generate_alert_id('Domain', domain_name, domain_status.status)
                if temp_alert_id == alert_id:
                    alert_details = {
                        'id': alert_id,
                        'type': 'Domain',
                        'domain': domain_name,
                        'status': domain_status.status,
                        'message': f'Domain {domain_name} will expire in {domain_status.days_remaining} days' if domain_status.status == 'warning' else
                                  f'Domain {domain_name} has expired' if domain_status.status == 'expired' else
                                  f'Error checking expiry for domain {domain_name}'
                    }
                    break

    # If still not found, check ping alerts
    if not alert_details:
        for entry in config.get('ping_hosts', []):
            domain_name = entry.get('host')
            ping_result = check_ping(domain_name)

            if ping_result['status'] == 'down':
                temp_alert_id = generate_alert_id('Ping', domain_name, 'down')
                if temp_alert_id == alert_id:
                    alert_details = {
                        'id': alert_id,
                        'type': 'Ping',
                        'domain': domain_name,
                        'status': 'down',
                        'message': f'Host {domain_name} is down'
                    }
                    break

    return alert_details

@app.route('/alerts')
@auth.login_required
def alerts():
    """Alerts page showing all system alerts"""
    start_time = time.time()
    logger.debug("Starting alerts page load")

    # Get current user and organization
    user = auth.get_current_user()
    current_org = user.get('current_organization')

    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    # Get all alerts from the system
    alerts = []

    # Get SSL certificate alerts
    config = load_config()
    acknowledged_alerts = config.get('acknowledged_alerts', [])
    deleted_alerts = config.get('deleted_alerts', [])
    current_time = datetime.now().strftime('%Y-%m-%d %H:%M')

    # Get domains from database for current organization only
    domains_from_db = db.get_domains_by_organization(current_org['id'])

    # Process each domain
    for domain in domains_from_db:
        domain_name = domain['name']
        # Check if domain is monitored for SSL
        if domain['ssl_monitored']:
            try:
                cert_status = check_certificate(domain_name)
                logger.debug(f"SSL check for {domain_name}: status={cert_status.status}, days_remaining={cert_status.days_remaining}")

                if cert_status.status == 'warning':
                    alert_id = generate_alert_id('SSL', domain_name, 'warning')
                    # Skip if this alert has been deleted
                    if alert_id in deleted_alerts:
                        continue
                    is_acknowledged = alert_id in acknowledged_alerts
                    alerts.append({
                        'id': alert_id,
                        'type': 'SSL',
                        'icon': 'shield-exclamation',
                        'message': f'SSL certificate for {domain_name} will expire in {cert_status.days_remaining} days',
                        'time': current_time,
                        'domain': domain_name,
                        'acknowledged': is_acknowledged
                    })
                elif cert_status.status == 'expired':
                    alert_id = generate_alert_id('SSL', domain_name, 'expired')
                    # Skip if this alert has been deleted
                    if alert_id in deleted_alerts:
                        continue
                    is_acknowledged = alert_id in acknowledged_alerts
                    alerts.append({
                        'id': alert_id,
                        'type': 'SSL',
                        'icon': 'shield-x',
                        'message': f'SSL certificate for {domain_name} has expired',
                        'time': current_time,
                        'domain': domain_name,
                        'acknowledged': is_acknowledged
                    })
                elif cert_status.status == 'error':
                    alert_id = generate_alert_id('SSL', domain_name, 'error')
                    # Skip if this alert has been deleted
                    if alert_id in deleted_alerts:
                        continue
                    is_acknowledged = alert_id in acknowledged_alerts
                    alerts.append({
                        'id': alert_id,
                        'type': 'Error',
                        'icon': 'exclamation-triangle',
                        'message': f'Error checking SSL certificate for {domain_name}',
                        'time': current_time,
                        'domain': domain_name,
                        'acknowledged': is_acknowledged
                    })
            except Exception as e:
                logger.error(f"Error checking SSL certificate for {domain_name}: {str(e)}")
                # Create an error alert for this domain
                alert_id = generate_alert_id('SSL', domain_name, 'error')
                # Skip if this alert has been deleted
                if alert_id in deleted_alerts:
                    continue
                is_acknowledged = alert_id in acknowledged_alerts
                alerts.append({
                    'id': alert_id,
                    'type': 'Error',
                    'icon': 'exclamation-triangle',
                    'message': f'Error checking SSL certificate for {domain_name}',
                    'time': current_time,
                    'domain': domain_name,
                    'acknowledged': is_acknowledged
                })

        # Check if domain is monitored for expiry
        if domain['expiry_monitored']:
            try:
                domain_status = check_domain_expiry(domain_name)
                logger.debug(f"Domain expiry check for {domain_name}: status={domain_status.status}, days_remaining={domain_status.days_remaining}")

                if domain_status.status == 'warning':
                    alert_id = generate_alert_id('Domain', domain_name, 'warning')
                    # Skip if this alert has been deleted
                    if alert_id in deleted_alerts:
                        continue
                    is_acknowledged = alert_id in acknowledged_alerts
                    alerts.append({
                        'id': alert_id,
                        'type': 'Domain',
                        'icon': 'calendar-exclamation',
                        'message': f'Domain {domain_name} will expire in {domain_status.days_remaining} days',
                        'time': current_time,
                        'domain': domain_name,
                        'acknowledged': is_acknowledged
                    })
                elif domain_status.status == 'expired':
                    alert_id = generate_alert_id('Domain', domain_name, 'expired')
                    # Skip if this alert has been deleted
                    if alert_id in deleted_alerts:
                        continue
                    is_acknowledged = alert_id in acknowledged_alerts
                    alerts.append({
                        'id': alert_id,
                        'type': 'Domain',
                        'icon': 'calendar-x',
                        'message': f'Domain {domain_name} has expired',
                        'time': current_time,
                        'domain': domain_name,
                        'acknowledged': is_acknowledged
                    })
                elif domain_status.status == 'error':
                    alert_id = generate_alert_id('Domain', domain_name, 'error')
                    # Skip if this alert has been deleted
                    if alert_id in deleted_alerts:
                        continue
                    is_acknowledged = alert_id in acknowledged_alerts
                    alerts.append({
                        'id': alert_id,
                        'type': 'Error',
                        'icon': 'exclamation-triangle',
                        'message': f'Error checking expiry for domain {domain_name}',
                        'time': current_time,
                        'domain': domain_name,
                        'acknowledged': is_acknowledged
                    })
            except Exception as e:
                logger.error(f"Error checking domain expiry for {domain_name}: {str(e)}")
                # Create an error alert for this domain
                alert_id = generate_alert_id('Domain', domain_name, 'error')
                # Skip if this alert has been deleted
                if alert_id in deleted_alerts:
                    continue
                is_acknowledged = alert_id in acknowledged_alerts
                alerts.append({
                    'id': alert_id,
                    'type': 'Error',
                    'icon': 'exclamation-triangle',
                    'message': f'Error checking expiry for domain {domain_name}',
                    'time': current_time,
                    'domain': domain_name,
                    'acknowledged': is_acknowledged
                })

        # Check if domain is monitored for ping
        if domain['ping_monitored']:
            try:
                # Always do a fresh ping check
                ping_result = check_ping(domain_name)
                ping_status = ping_result['status']
                logger.debug(f"Ping check for {domain_name}: status={ping_status}")

                if ping_status == 'down':
                    alert_id = generate_alert_id('Ping', domain_name, 'down')
                    # Skip if this alert has been deleted
                    if alert_id in deleted_alerts:
                        continue
                    is_acknowledged = alert_id in acknowledged_alerts
                    alerts.append({
                        'id': alert_id,
                        'type': 'Ping',
                        'icon': 'wifi-off',
                        'message': f'Host {domain_name} is down',
                        'time': current_time,
                        'domain': domain_name,
                        'acknowledged': is_acknowledged
                    })
            except Exception as e:
                logger.error(f"Error checking ping for {domain_name}: {str(e)}")
                # Create an error alert for this domain
                alert_id = generate_alert_id('Ping', domain_name, 'error')
                # Skip if this alert has been deleted
                if alert_id in deleted_alerts:
                    continue
                is_acknowledged = alert_id in acknowledged_alerts
                alerts.append({
                    'id': alert_id,
                    'type': 'Error',
                    'icon': 'exclamation-triangle',
                    'message': f'Error checking ping for {domain_name}',
                    'time': current_time,
                    'domain': domain_name,
                    'acknowledged': is_acknowledged
                })

    # Sort alerts by acknowledgment status (unacknowledged first) and then by time (newest first)
    alerts.sort(key=lambda x: (x['acknowledged'], x['time']))

    end_time = time.time()
    logger.debug(f"Alerts page loaded in {end_time - start_time:.2f} seconds")
    logger.debug(f"Number of alerts found: {len(alerts)}")
    for alert in alerts:
        logger.debug(f"Alert: {alert['id']} - {alert['message']} - Acknowledged: {alert['acknowledged']}")

    return render_template('alerts.html', alerts=alerts)

@app.route('/acknowledge_alert/<alert_id>', methods=['POST'])
def acknowledge_alert(alert_id):
    """Acknowledge an alert"""
    config = load_config()
    success = False
    message = ""

    if 'acknowledged_alerts' not in config:
        config['acknowledged_alerts'] = []

    if alert_id not in config['acknowledged_alerts']:
        # Get alert details before acknowledging
        alert_details = None

        # Get SSL certificate alerts
        for entry in config.get('ssl_domains', []):
            domain_name = entry.get('url')
            cert_status = check_certificate(domain_name)

            if cert_status.status in ['warning', 'expired', 'error']:
                temp_alert_id = generate_alert_id('SSL', domain_name, cert_status.status)
                if temp_alert_id == alert_id:
                    alert_details = {
                        'id': alert_id,
                        'type': 'SSL',
                        'domain': domain_name,
                        'status': cert_status.status,
                        'message': f'SSL certificate for {domain_name} will expire in {cert_status.days_remaining} days' if cert_status.status == 'warning' else
                                  f'SSL certificate for {domain_name} has expired' if cert_status.status == 'expired' else
                                  f'Error checking SSL certificate for {domain_name}'
                    }
                    break

        # If not found, check domain expiry alerts
        if not alert_details:
            for entry in config.get('domain_expiry', []):
                domain_name = entry.get('name')
                domain_status = check_domain_expiry(domain_name)

                if domain_status.status in ['warning', 'expired', 'error']:
                    temp_alert_id = generate_alert_id('Domain', domain_name, domain_status.status)
                    if temp_alert_id == alert_id:
                        alert_details = {
                            'id': alert_id,
                            'type': 'Domain',
                            'domain': domain_name,
                            'status': domain_status.status,
                            'message': f'Domain {domain_name} will expire in {domain_status.days_remaining} days' if domain_status.status == 'warning' else
                                      f'Domain {domain_name} has expired' if domain_status.status == 'expired' else
                                      f'Error checking expiry for domain {domain_name}'
                        }
                        break

        # If still not found, check ping alerts
        if not alert_details:
            for entry in config.get('ping_hosts', []):
                domain_name = entry.get('host')
                ping_result = check_ping(domain_name)

                if ping_result['status'] == 'down':
                    temp_alert_id = generate_alert_id('Ping', domain_name, 'down')
                    if temp_alert_id == alert_id:
                        alert_details = {
                            'id': alert_id,
                            'type': 'Ping',
                            'domain': domain_name,
                            'status': 'down',
                            'message': f'Host {domain_name} is down'
                        }
                        break

        if alert_details:
            # Add to acknowledged alerts
            config['acknowledged_alerts'].append(alert_id)
            save_config(config)

            # Record in alert history
            import database as db
            current_user = auth.get_current_user()
            user_id = current_user['id'] if current_user else None
            username = current_user['username'] if current_user else 'System'

            # Get current organization
            current_org = current_user.get('current_organization') if current_user else None
            organization_id = current_org['id'] if current_org else None

            db.add_alert_history(
                alert_id=alert_id,
                domain_name=alert_details['domain'],
                alert_type=alert_details['type'].lower(),
                status=alert_details.get('status', 'warning'),
                message=alert_details['message'],
                action='acknowledged',
                user_id=user_id,
                username=username
            )

            # Log the alert acknowledgment action
            if user_id:
                db.log_user_action(
                    user_id=user_id,
                    username=username,
                    action_type='acknowledge',
                    resource_type='alert',
                    resource_id=alert_id,
                    resource_name=alert_details['domain'],
                    details=f"Acknowledged {alert_details['type']} alert: {alert_details['message']}",
                    ip_address=request.remote_addr,
                    organization_id=organization_id
                )

            success = True
            message = f"Alert '{alert_details['message']}' acknowledged"
        else:
            message = "Alert not found"
    else:
        success = True
        message = "Alert already acknowledged"

    # Check if this is an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({
            'success': success,
            'message': message
        })

    # For regular form submissions, flash the message and redirect
    if success:
        flash(message, 'success')
    else:
        flash(message, 'error')

    return redirect(url_for('alerts'))

@app.route('/health')
def health_check():
    return jsonify({'status': 'healthy'}), 200

@app.route('/unacknowledge_alert/<alert_id>', methods=['POST'])
def unacknowledge_alert(alert_id):
    """Remove acknowledgment from an alert"""
    config = load_config()
    success = False
    message = ""

    if 'acknowledged_alerts' in config and alert_id in config['acknowledged_alerts']:
        # Get alert details before unacknowledging
        alert_details = None

        # Get SSL certificate alerts
        for entry in config.get('ssl_domains', []):
            domain_name = entry.get('url')
            cert_status = check_certificate(domain_name)

            if cert_status.status in ['warning', 'expired', 'error']:
                temp_alert_id = generate_alert_id('SSL', domain_name, cert_status.status)
                if temp_alert_id == alert_id:
                    alert_details = {
                        'id': alert_id,
                        'type': 'SSL',
                        'domain': domain_name,
                        'status': cert_status.status,
                        'message': f'SSL certificate for {domain_name} will expire in {cert_status.days_remaining} days' if cert_status.status == 'warning' else
                                  f'SSL certificate for {domain_name} has expired' if cert_status.status == 'expired' else
                                  f'Error checking SSL certificate for {domain_name}'
                    }
                    break

        # If not found, check domain expiry alerts
        if not alert_details:
            for entry in config.get('domain_expiry', []):
                domain_name = entry.get('name')
                domain_status = check_domain_expiry(domain_name)

                if domain_status.status in ['warning', 'expired', 'error']:
                    temp_alert_id = generate_alert_id('Domain', domain_name, domain_status.status)
                    if temp_alert_id == alert_id:
                        alert_details = {
                            'id': alert_id,
                            'type': 'Domain',
                            'domain': domain_name,
                            'status': domain_status.status,
                            'message': f'Domain {domain_name} will expire in {domain_status.days_remaining} days' if domain_status.status == 'warning' else
                                      f'Domain {domain_name} has expired' if domain_status.status == 'expired' else
                                      f'Error checking expiry for domain {domain_name}'
                        }
                        break

        # If still not found, check ping alerts
        if not alert_details:
            for entry in config.get('ping_hosts', []):
                domain_name = entry.get('host')
                ping_result = check_ping(domain_name)

                if ping_result['status'] == 'down':
                    temp_alert_id = generate_alert_id('Ping', domain_name, 'down')
                    if temp_alert_id == alert_id:
                        alert_details = {
                            'id': alert_id,
                            'type': 'Ping',
                            'domain': domain_name,
                            'status': 'down',
                            'message': f'Host {domain_name} is down'
                        }
                        break

        # Remove from acknowledged alerts
        config['acknowledged_alerts'].remove(alert_id)
        save_config(config)

        # Record in alert history
        if alert_details:
            import database as db
            current_user = auth.get_current_user()
            user_id = current_user['id'] if current_user else None
            username = current_user['username'] if current_user else 'System'

            db.add_alert_history(
                alert_id=alert_id,
                domain_name=alert_details['domain'],
                alert_type=alert_details['type'].lower(),
                status=alert_details.get('status', 'warning'),
                message=alert_details['message'],
                action='unacknowledged',
                user_id=user_id,
                username=username
            )

        success = True
        message = 'Alert acknowledgment removed'
    else:
        message = 'Alert was not acknowledged'

    # Check if this is an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({
            'success': success,
            'message': message
        })

    # For regular form submissions, flash the message and redirect
    if success:
        flash(message, 'success')
    else:
        flash(message, 'info')

    return redirect(url_for('alerts'))

@app.route('/delete_alert/<alert_id>', methods=['POST'])
def delete_alert(alert_id):
    """Delete an acknowledged alert"""
    config = load_config()
    success = False
    message = ""

    # Only allow deletion of acknowledged alerts
    if 'acknowledged_alerts' in config and alert_id in config['acknowledged_alerts']:
        # Get alert details before deleting
        alert_details = None

        # Get SSL certificate alerts
        for entry in config.get('ssl_domains', []):
            domain_name = entry.get('url')
            cert_status = check_certificate(domain_name)

            if cert_status.status in ['warning', 'expired', 'error']:
                temp_alert_id = generate_alert_id('SSL', domain_name, cert_status.status)
                if temp_alert_id == alert_id:
                    alert_details = {
                        'id': alert_id,
                        'type': 'SSL',
                        'domain': domain_name,
                        'status': cert_status.status,
                        'message': f'SSL certificate for {domain_name} will expire in {cert_status.days_remaining} days' if cert_status.status == 'warning' else
                                  f'SSL certificate for {domain_name} has expired' if cert_status.status == 'expired' else
                                  f'Error checking SSL certificate for {domain_name}'
                    }
                    break

        # If not found, check domain expiry alerts
        if not alert_details:
            for entry in config.get('domain_expiry', []):
                domain_name = entry.get('name')
                domain_status = check_domain_expiry(domain_name)

                if domain_status.status in ['warning', 'expired', 'error']:
                    temp_alert_id = generate_alert_id('Domain', domain_name, domain_status.status)
                    if temp_alert_id == alert_id:
                        alert_details = {
                            'id': alert_id,
                            'type': 'Domain',
                            'domain': domain_name,
                            'status': domain_status.status,
                            'message': f'Domain {domain_name} will expire in {domain_status.days_remaining} days' if domain_status.status == 'warning' else
                                      f'Domain {domain_name} has expired' if domain_status.status == 'expired' else
                                      f'Error checking expiry for domain {domain_name}'
                        }
                        break

        # If still not found, check ping alerts
        if not alert_details:
            for entry in config.get('ping_hosts', []):
                domain_name = entry.get('host')
                ping_result = check_ping(domain_name)

                if ping_result['status'] == 'down':
                    temp_alert_id = generate_alert_id('Ping', domain_name, 'down')
                    if temp_alert_id == alert_id:
                        alert_details = {
                            'id': alert_id,
                            'type': 'Ping',
                            'domain': domain_name,
                            'status': 'down',
                            'message': f'Host {domain_name} is down'
                        }
                        break

        # Remove from acknowledged alerts
        config['acknowledged_alerts'].remove(alert_id)

        # Add to deleted alerts list
        if 'deleted_alerts' not in config:
            config['deleted_alerts'] = []

        if alert_id not in config['deleted_alerts']:
            config['deleted_alerts'].append(alert_id)

        save_config(config)

        # Record in alert history
        if alert_details:
            import database as db
            current_user = auth.get_current_user()
            user_id = current_user['id'] if current_user else None
            username = current_user['username'] if current_user else 'System'

            db.add_alert_history(
                alert_id=alert_id,
                domain_name=alert_details['domain'],
                alert_type=alert_details['type'].lower(),
                status=alert_details.get('status', 'warning'),
                message=alert_details['message'],
                action='deleted',
                user_id=user_id,
                username=username
            )

        success = True
        message = 'Alert deleted successfully'
    else:
        message = 'Only acknowledged alerts can be deleted'

    # Check if this is an AJAX request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({
            'success': success,
            'message': message
        })

    # For regular form submissions, flash the message and redirect
    if success:
        flash(message, 'success')
    else:
        flash(message, 'warning')

    return redirect(url_for('alerts'))

@app.route('/acknowledge_all_alerts', methods=['POST'])
def acknowledge_all_alerts():
    """Acknowledge multiple alerts at once"""
    alert_ids = request.form.getlist('alert_ids')

    if not alert_ids:
        flash('No alerts selected', 'warning')
        return redirect(url_for('alerts'))

    config = load_config()

    if 'acknowledged_alerts' not in config:
        config['acknowledged_alerts'] = []

    # Add all alert IDs that aren't already acknowledged
    added_count = 0
    for alert_id in alert_ids:
        if alert_id not in config['acknowledged_alerts']:
            config['acknowledged_alerts'].append(alert_id)
            added_count += 1

    if added_count > 0:
        save_config(config)
        flash(f'Acknowledged {added_count} alerts', 'success')
    else:
        flash('All selected alerts were already acknowledged', 'info')

    return redirect(url_for('alerts'))

@app.route('/bulk_acknowledge_alerts', methods=['POST'])
@auth.login_required
def bulk_acknowledge_alerts():
    """Acknowledge multiple selected alerts"""
    alert_ids = request.form.getlist('alert_ids')
    if not alert_ids:
        flash('No alerts selected', 'error')
        return redirect(url_for('alerts'))

    config = load_config()
    if 'acknowledged_alerts' not in config:
        config['acknowledged_alerts'] = []

    # Get current user
    current_user = auth.get_current_user()
    username = current_user.get('username', 'Unknown')
    user_id = current_user.get('id')

    # Acknowledge each alert
    acknowledged_count = 0
    for alert_id in alert_ids:
        if alert_id not in config['acknowledged_alerts']:
            config['acknowledged_alerts'].append(alert_id)
            acknowledged_count += 1

            # Record in alert history
            alert_details = get_alert_details(alert_id)
            if alert_details:
                db.add_alert_history(
                    alert_id=alert_id,
                    domain_name=alert_details.get('domain', 'Unknown'),
                    alert_type=alert_details.get('type', 'Unknown').lower(),
                    status=alert_details.get('status', 'Unknown'),
                    message=alert_details.get('message', 'Unknown'),
                    action='acknowledged',
                    user_id=user_id,
                    username=username
                )

    # Save config
    save_config(config)

    if acknowledged_count > 0:
        flash(f'Successfully acknowledged {acknowledged_count} alert(s)', 'success')
    else:
        flash('No alerts were acknowledged', 'info')

    return redirect(url_for('alerts'))

@app.route('/bulk_unacknowledge_alerts', methods=['POST'])
@auth.login_required
def bulk_unacknowledge_alerts():
    """Remove acknowledgment from multiple selected alerts"""
    alert_ids = request.form.getlist('alert_ids')
    if not alert_ids:
        flash('No alerts selected', 'error')
        return redirect(url_for('alerts'))

    config = load_config()
    if 'acknowledged_alerts' not in config:
        flash('No acknowledged alerts found', 'error')
        return redirect(url_for('alerts'))

    # Get current user
    current_user = auth.get_current_user()
    username = current_user.get('username', 'Unknown')
    user_id = current_user.get('id')

    # Unacknowledge each alert
    unacknowledged_count = 0
    for alert_id in alert_ids:
        if alert_id in config['acknowledged_alerts']:
            config['acknowledged_alerts'].remove(alert_id)
            unacknowledged_count += 1

            # Record in alert history
            alert_details = get_alert_details(alert_id)
            if alert_details:
                db.add_alert_history(
                    alert_id=alert_id,
                    domain_name=alert_details.get('domain', 'Unknown'),
                    alert_type=alert_details.get('type', 'Unknown').lower(),
                    status=alert_details.get('status', 'Unknown'),
                    message=alert_details.get('message', 'Unknown'),
                    action='unacknowledged',
                    user_id=user_id,
                    username=username
                )

    # Save config
    save_config(config)

    if unacknowledged_count > 0:
        flash(f'Successfully unacknowledged {unacknowledged_count} alert(s)', 'success')
    else:
        flash('No alerts were unacknowledged', 'info')

    return redirect(url_for('alerts'))

@app.route('/bulk_delete_alerts', methods=['POST'])
@auth.login_required
def bulk_delete_alerts():
    """Delete multiple selected alerts"""
    alert_ids = request.form.getlist('alert_ids')
    if not alert_ids:
        flash('No alerts selected', 'error')
        return redirect(url_for('alerts'))

    config = load_config()
    if 'deleted_alerts' not in config:
        config['deleted_alerts'] = []

    # Get current user
    current_user = auth.get_current_user()
    username = current_user.get('username', 'Unknown')
    user_id = current_user.get('id')

    # Delete each alert
    deleted_count = 0
    for alert_id in alert_ids:
        # Only delete acknowledged alerts
        if alert_id in config.get('acknowledged_alerts', []):
            # Move from acknowledged to deleted
            if alert_id not in config['deleted_alerts']:
                config['deleted_alerts'].append(alert_id)
                deleted_count += 1

                # Remove from acknowledged
                config['acknowledged_alerts'].remove(alert_id)

                # Record in alert history
                alert_details = get_alert_details(alert_id)
                if alert_details:
                    db.add_alert_history(
                        alert_id=alert_id,
                        domain_name=alert_details.get('domain', 'Unknown'),
                        alert_type=alert_details.get('type', 'Unknown').lower(),
                        status=alert_details.get('status', 'Unknown'),
                        message=alert_details.get('message', 'Unknown'),
                        action='deleted',
                        user_id=user_id,
                        username=username
                    )

    # Save config
    save_config(config)

    if deleted_count > 0:
        flash(f'Successfully deleted {deleted_count} alert(s)', 'success')
    else:
        flash('No alerts were deleted', 'info')

    return redirect(url_for('alerts'))

@app.route('/delete_all_acknowledged_alerts', methods=['POST'])
def delete_all_acknowledged_alerts():
    """Delete all acknowledged alerts"""
    config = load_config()

    if 'acknowledged_alerts' not in config or not config['acknowledged_alerts']:
        flash('No acknowledged alerts to delete', 'info')
        return redirect(url_for('alerts'))

    # Count the number of alerts being deleted
    deleted_count = len(config['acknowledged_alerts'])

    # Ensure deleted_alerts list exists
    if 'deleted_alerts' not in config:
        config['deleted_alerts'] = []

    # Get alert details for all acknowledged alerts
    alert_details_map = {}

    # Check SSL certificate alerts
    for entry in config.get('ssl_domains', []):
        domain_name = entry.get('url')
        cert_status = check_certificate(domain_name)

        if cert_status.status in ['warning', 'expired', 'error']:
            alert_id = generate_alert_id('SSL', domain_name, cert_status.status)
            if alert_id in config['acknowledged_alerts']:
                alert_details_map[alert_id] = {
                    'id': alert_id,
                    'type': 'SSL',
                    'domain': domain_name,
                    'status': cert_status.status,
                    'message': f'SSL certificate for {domain_name} will expire in {cert_status.days_remaining} days' if cert_status.status == 'warning' else
                              f'SSL certificate for {domain_name} has expired' if cert_status.status == 'expired' else
                              f'Error checking SSL certificate for {domain_name}'
                }

    # Check domain expiry alerts
    for entry in config.get('domain_expiry', []):
        domain_name = entry.get('name')
        domain_status = check_domain_expiry(domain_name)

        if domain_status.status in ['warning', 'expired', 'error']:
            alert_id = generate_alert_id('Domain', domain_name, domain_status.status)
            if alert_id in config['acknowledged_alerts']:
                alert_details_map[alert_id] = {
                    'id': alert_id,
                    'type': 'Domain',
                    'domain': domain_name,
                    'status': domain_status.status,
                    'message': f'Domain {domain_name} will expire in {domain_status.days_remaining} days' if domain_status.status == 'warning' else
                              f'Domain {domain_name} has expired' if domain_status.status == 'expired' else
                              f'Error checking expiry for domain {domain_name}'
                }

    # Check ping alerts
    for entry in config.get('ping_hosts', []):
        domain_name = entry.get('host')
        ping_result = check_ping(domain_name)

        if ping_result['status'] == 'down':
            alert_id = generate_alert_id('Ping', domain_name, 'down')
            if alert_id in config['acknowledged_alerts']:
                alert_details_map[alert_id] = {
                    'id': alert_id,
                    'type': 'Ping',
                    'domain': domain_name,
                    'status': 'down',
                    'message': f'Host {domain_name} is down'
                }

    # Move all acknowledged alerts to deleted_alerts list
    for alert_id in config['acknowledged_alerts']:
        if alert_id not in config['deleted_alerts']:
            config['deleted_alerts'].append(alert_id)

    # Record in alert history
    import database as db
    current_user = auth.get_current_user()
    user_id = current_user['id'] if current_user else None
    username = current_user['username'] if current_user else 'System'

    for alert_id, alert_details in alert_details_map.items():
        db.add_alert_history(
            alert_id=alert_id,
            domain_name=alert_details['domain'],
            alert_type=alert_details['type'].lower(),
            status=alert_details.get('status', 'warning'),
            message=alert_details['message'],
            action='deleted',
            user_id=user_id,
            username=username
        )

    # Clear the acknowledged alerts list
    config['acknowledged_alerts'] = []
    save_config(config)

    flash(f'Successfully deleted {deleted_count} acknowledged alerts', 'success')
    return redirect(url_for('alerts'))

@app.route('/restore_alert/<alert_id>', methods=['POST'])
@auth.login_required
def restore_alert(alert_id):
    """Restore a single deleted alert"""
    config = load_config()

    if 'deleted_alerts' not in config or alert_id not in config['deleted_alerts']:
        flash('Alert not found or already restored', 'warning')
        return redirect(url_for('archived_alerts'))

    # Get alert details
    alert_details = None

    # Check SSL certificate alerts
    for entry in config.get('ssl_domains', []):
        domain_name = entry.get('url')
        cert_status = check_certificate(domain_name)

        if cert_status.status in ['warning', 'expired', 'error']:
            temp_alert_id = generate_alert_id('SSL', domain_name, cert_status.status)
            if temp_alert_id == alert_id:
                alert_details = {
                    'id': alert_id,
                    'type': 'SSL',
                    'domain': domain_name,
                    'status': cert_status.status,
                    'message': f'SSL certificate for {domain_name} will expire in {cert_status.days_remaining} days' if cert_status.status == 'warning' else
                              f'SSL certificate for {domain_name} has expired' if cert_status.status == 'expired' else
                              f'Error checking SSL certificate for {domain_name}'
                }
                break

    # If not found, check domain expiry alerts
    if not alert_details:
        for entry in config.get('domain_expiry', []):
            domain_name = entry.get('name')
            domain_status = check_domain_expiry(domain_name)

            if domain_status.status in ['warning', 'expired', 'error']:
                temp_alert_id = generate_alert_id('Domain', domain_name, domain_status.status)
                if temp_alert_id == alert_id:
                    alert_details = {
                        'id': alert_id,
                        'type': 'Domain',
                        'domain': domain_name,
                        'status': domain_status.status,
                        'message': f'Domain {domain_name} will expire in {domain_status.days_remaining} days' if domain_status.status == 'warning' else
                                  f'Domain {domain_name} has expired' if domain_status.status == 'expired' else
                                  f'Error checking expiry for domain {domain_name}'
                    }
                    break

    # If still not found, check ping alerts
    if not alert_details:
        for entry in config.get('ping_hosts', []):
            domain_name = entry.get('host')
            ping_result = check_ping(domain_name)

            if ping_result['status'] == 'down':
                temp_alert_id = generate_alert_id('Ping', domain_name, 'down')
                if temp_alert_id == alert_id:
                    alert_details = {
                        'id': alert_id,
                        'type': 'Ping',
                        'domain': domain_name,
                        'status': 'down',
                        'message': f'Host {domain_name} is down'
                    }
                    break

    # Remove from deleted alerts list
    config['deleted_alerts'].remove(alert_id)
    save_config(config)

    # Record in alert history
    if alert_details:
        import database as db
        current_user = auth.get_current_user()
        user_id = current_user['id'] if current_user else None
        username = current_user['username'] if current_user else 'System'

        db.add_alert_history(
            alert_id=alert_id,
            domain_name=alert_details['domain'],
            alert_type=alert_details['type'].lower(),
            status=alert_details.get('status', 'warning'),
            message=alert_details['message'],
            action='restored',
            user_id=user_id,
            username=username
        )

    flash('Alert restored successfully', 'success')
    return redirect(url_for('archived_alerts'))

@app.route('/restore_deleted_alerts', methods=['POST'])
@auth.login_required
def restore_deleted_alerts():
    """Restore all deleted alerts"""
    config = load_config()

    if 'deleted_alerts' not in config or not config['deleted_alerts']:
        flash('No deleted alerts to restore', 'info')
        return redirect(url_for('archived_alerts'))

    # Count the number of alerts being restored
    restored_count = len(config['deleted_alerts'])

    # Get alert details for all deleted alerts
    alert_details_map = {}

    # Check SSL certificate alerts
    for entry in config.get('ssl_domains', []):
        domain_name = entry.get('url')
        cert_status = check_certificate(domain_name)

        if cert_status.status in ['warning', 'expired', 'error']:
            alert_id = generate_alert_id('SSL', domain_name, cert_status.status)
            if alert_id in config['deleted_alerts']:
                alert_details_map[alert_id] = {
                    'id': alert_id,
                    'type': 'SSL',
                    'domain': domain_name,
                    'status': cert_status.status,
                    'message': f'SSL certificate for {domain_name} will expire in {cert_status.days_remaining} days' if cert_status.status == 'warning' else
                              f'SSL certificate for {domain_name} has expired' if cert_status.status == 'expired' else
                              f'Error checking SSL certificate for {domain_name}'
                }

    # Check domain expiry alerts
    for entry in config.get('domain_expiry', []):
        domain_name = entry.get('name')
        domain_status = check_domain_expiry(domain_name)

        if domain_status.status in ['warning', 'expired', 'error']:
            alert_id = generate_alert_id('Domain', domain_name, domain_status.status)
            if alert_id in config['deleted_alerts']:
                alert_details_map[alert_id] = {
                    'id': alert_id,
                    'type': 'Domain',
                    'domain': domain_name,
                    'status': domain_status.status,
                    'message': f'Domain {domain_name} will expire in {domain_status.days_remaining} days' if domain_status.status == 'warning' else
                              f'Domain {domain_name} has expired' if domain_status.status == 'expired' else
                              f'Error checking expiry for domain {domain_name}'
                }

    # Check ping alerts
    for entry in config.get('ping_hosts', []):
        domain_name = entry.get('host')
        ping_result = check_ping(domain_name)

        if ping_result['status'] == 'down':
            alert_id = generate_alert_id('Ping', domain_name, 'down')
            if alert_id in config['deleted_alerts']:
                alert_details_map[alert_id] = {
                    'id': alert_id,
                    'type': 'Ping',
                    'domain': domain_name,
                    'status': 'down',
                    'message': f'Host {domain_name} is down'
                }

    # Record in alert history
    import database as db
    current_user = auth.get_current_user()
    user_id = current_user['id'] if current_user else None
    username = current_user['username'] if current_user else 'System'

    for alert_id, alert_details in alert_details_map.items():
        db.add_alert_history(
            alert_id=alert_id,
            domain_name=alert_details['domain'],
            alert_type=alert_details['type'].lower(),
            status=alert_details.get('status', 'warning'),
            message=alert_details['message'],
            action='restored',
            user_id=user_id,
            username=username
        )

    # Clear the deleted alerts list
    config['deleted_alerts'] = []
    save_config(config)

    flash(f'Successfully restored {restored_count} deleted alerts', 'success')
    return redirect(url_for('archived_alerts'))

@app.route('/archived_alerts')
@auth.login_required
def archived_alerts():
    """Archived Alerts page showing history of all alerts"""
    start_time = time.time()
    logger.debug("Starting archived alerts page load")

    # Get pagination parameters
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)

    # Calculate offset
    offset = (page - 1) * per_page

    # Get alert history from database
    import database as db
    alert_history = db.get_alert_history(limit=per_page, offset=offset)
    total_alerts = db.get_alert_history_count()

    # Calculate pagination info
    total_pages = (total_alerts + per_page - 1) // per_page
    has_prev = page > 1
    has_next = page < total_pages

    end_time = time.time()
    logger.debug(f"Archived alerts page loaded in {end_time - start_time:.2f} seconds")

    return render_template(
        'archived_alerts.html',
        alert_history=alert_history,
        page=page,
        per_page=per_page,
        total_alerts=total_alerts,
        total_pages=total_pages,
        has_prev=has_prev,
        has_next=has_next
    )

def get_alerts_count():
    """Count the number of active unacknowledged alerts in the system"""
    # Get current user and organization
    user = auth.get_current_user()
    if not user:
        return 0

    current_org = user.get('current_organization')
    if not current_org:
        return 0

    config = load_config()
    alerts_count = 0
    acknowledged_alerts = config.get('acknowledged_alerts', [])
    deleted_alerts = config.get('deleted_alerts', [])

    # Get domains from database for current organization only
    domains_from_db = db.get_domains_by_organization(current_org['id'])

    # Process each domain
    for domain in domains_from_db:
        domain_name = domain['name']

        # Check if domain is monitored for SSL
        if domain['ssl_monitored']:
            cert_status = check_certificate(domain_name)
            if cert_status.status in ['warning', 'expired', 'error']:
                alert_id = generate_alert_id('SSL', domain_name, cert_status.status)
                if alert_id not in acknowledged_alerts and alert_id not in deleted_alerts:
                    alerts_count += 1

        # Check if domain is monitored for expiry
        if domain['expiry_monitored']:
            domain_status = check_domain_expiry(domain_name)
            if domain_status.status in ['warning', 'expired', 'error']:
                alert_id = generate_alert_id('Domain', domain_name, domain_status.status)
                if alert_id not in acknowledged_alerts and alert_id not in deleted_alerts:
                    alerts_count += 1

        # Check if domain is monitored for ping
        if domain['ping_monitored']:
            ping_result = check_ping(domain_name)
            if ping_result['status'] == 'down':
                alert_id = generate_alert_id('Ping', domain_name, 'down')
                if alert_id not in acknowledged_alerts and alert_id not in deleted_alerts:
                    alerts_count += 1

    return alerts_count

@app.context_processor
def inject_alerts_count():
    """Inject alerts count into all templates"""
    return {'alerts_count': get_alerts_count()}

@app.context_processor
def inject_user():
    """Inject current user into all templates"""
    try:
        current_user = auth.get_current_user()
        return {'user': current_user}
    except:
        return {'user': None}

# Test alert route
@app.route('/test_alert')
@auth.login_required
def test_alert():
    """Test page for alert notifications"""
    user = auth.get_current_user()
    return render_template('test_alert.html', user=user)

@app.route('/')
@auth.login_required
def index():
    """Dashboard page showing overview of all monitoring"""
    start_time = time.time()
    logger.debug("Starting dashboard page load")

    # Clear any "You have been logged out" messages that might have persisted
    if '_flashes' in session:
        flashes = session.get('_flashes', [])
        session['_flashes'] = [(category, message) for category, message in flashes
                              if message != 'You have been logged out']

    # Get current user
    current_user = auth.get_current_user()

    # Get user's organizations
    user_orgs = current_user.get('organizations', [])

    # If user has no organizations, redirect to create first organization
    if not user_orgs and not current_user['is_admin']:
        flash("You need to create an organization before you can use Certifly", 'info')
        return redirect(url_for('create_first_organization'))

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    # Get domains for current organization
    domains_from_db = db.get_domains_by_organization(current_org['id'])

    config = load_config()
    ssl_domains = {}
    domain_expiry = {}
    ping_hosts = {}

    # Initialize all_domains list
    all_domains = []

    # Process domains from database
    for domain in domains_from_db:
        domain_name = domain['name']
        domain_id = domain['id']

        # Check if domain is monitored for SSL
        ssl_status = None
        if domain['ssl_monitored']:
            ssl_status = check_certificate(domain_name)
            ssl_domains[domain_name] = ssl_status

        # Check if domain is monitored for expiry
        domain_status = None
        if domain['expiry_monitored']:
            domain_status = check_domain_expiry(domain_name)
            domain_expiry[domain_name] = domain_status

        # Check if domain is monitored for ping
        ping_status = 'unknown'
        if domain['ping_monitored']:
            ping_result = check_ping(domain_name)
            ping_status = ping_result['status']
            ping_hosts[domain_name] = ping_result

        # Determine if this domain has any issues
        has_issues = False
        if ssl_status and ssl_status.status in ['warning', 'expired', 'error']:
            has_issues = True
        if domain_status and domain_status.status in ['warning', 'expired', 'error']:
            has_issues = True
        if ping_status == 'down':
            has_issues = True

        # Determine expiry status for display
        expiry_status = 'unknown'
        days_until_expiry = 0
        if domain_status:
            expiry_status = domain_status.status
            days_until_expiry = domain_status.days_remaining

        # Get uptime statistics
        uptime_percentage = db.calculate_uptime_percentage(domain_name)
        if uptime_percentage is not None:
            uptime_percentage = round(uptime_percentage, 1)

        # Get ping history for uptime segments
        ping_history = db.get_ping_history(domain_name, hours=12)

        # Create uptime segments (12 segments representing the last 12 hours)
        uptime_segments = []

        # For newly added domains, we want to show only the most recent segment as the current status
        # and all other segments as unknown
        if len(ping_history) <= 2:  # Changed from 1 to 2 to handle cases with 2 entries
            # If there's minimal history (0, 1, or 2 entries), use current status only for the most recent segment
            uptime_segments = ['unknown'] * 11 + [ping_status]
        else:
            # Group history into 12 hourly segments
            current_time = datetime.now()
            for i in range(12):
                segment_start = current_time - timedelta(hours=12-i)
                segment_end = current_time - timedelta(hours=11-i)

                # Find ping entries in this segment
                segment_entries = [entry for entry in ping_history
                                  if entry['checked_at'] and
                                  segment_start <= datetime.fromtimestamp(entry['checked_at']) < segment_end]

                # Determine segment status
                if segment_entries:
                    # If any entry is down, the segment is down
                    if any(entry['status'] == 'down' for entry in segment_entries):
                        uptime_segments.append('down')
                    else:
                        uptime_segments.append('up')
                else:
                    # For the most recent segment (last hour), use current status if no data
                    if i == 11:  # Last segment (most recent hour)
                        uptime_segments.append(ping_status)
                    else:
                        uptime_segments.append('unknown')

        # Add domain to the list
        all_domains.append({
            'id': domain_id,
            'name': domain_name,
            'ssl_status': ssl_status,
            'domain_status': domain_status,
            'has_issues': has_issues,
            'expiry_status': expiry_status,
            'days_until_expiry': days_until_expiry,
            'health_status': ping_status,
            'uptime_percentage': uptime_percentage,
            'uptime_segments': uptime_segments
        })



    # Sort domains with issues first
    all_domains.sort(key=lambda x: (0 if x['has_issues'] else 1, x['name']))

    # Count active alerts (only unacknowledged ones)
    active_alerts = []
    acknowledged_alerts = config.get('acknowledged_alerts', [])
    deleted_alerts = config.get('deleted_alerts', [])

    # SSL certificate alerts
    for cert in ssl_domains.values():
        if cert.status in ['warning', 'expired', 'error']:
            alert_id = generate_alert_id('SSL', cert.domain, cert.status)
            if alert_id not in acknowledged_alerts and alert_id not in deleted_alerts:
                active_alerts.append(cert)

    # Domain expiry alerts
    for domain in domain_expiry.values():
        if domain.status in ['warning', 'expired', 'error']:
            alert_id = generate_alert_id('Domain', domain.name, domain.status)
            if alert_id not in acknowledged_alerts and alert_id not in deleted_alerts:
                active_alerts.append(domain)

    # Ping alerts
    for domain_name, host_data in ping_hosts.items():
        if host_data.get('status') == 'down':
            alert_id = generate_alert_id('Ping', domain_name, 'down')
            if alert_id not in acknowledged_alerts and alert_id not in deleted_alerts:
                # Create a simple object for ping alerts
                class PingAlert:
                    def __init__(self, domain, status):
                        self.domain = domain
                        self.status = status

                active_alerts.append(PingAlert(domain_name, 'down'))

    end_time = time.time()
    logger.debug(f"Dashboard page loaded in {end_time - start_time:.2f} seconds")

    return render_template('index.html',
                         domains=all_domains,
                         alerts=active_alerts,
                         user=current_user)

# Email settings route removed - functionality moved to notifications page

@app.route('/ssl_certificates')
@auth.login_required
def ssl_certificates():
    """SSL Certificate monitoring page"""
    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    # Get domains for current organization that have SSL monitoring enabled
    domains = db.get_domains_by_organization(current_org['id'])
    certificates = []

    for domain in domains:
        if domain['ssl_monitored']:
            # Check if this domain was recently added (within the last 10 seconds)
            # If so, clear the cache to ensure we get fresh data
            domain_created_at = domain.get('created_at', 0)
            current_time = time.time()

            if current_time - domain_created_at < 10:  # If domain was created in the last 10 seconds
                logger.info(f"Clearing cache for newly added domain: {domain['name']}")
                db.clear_cache(f"ssl_{domain['name']}")

            # Check the certificate
            cert_status = check_certificate(domain['name'])

            # If status is unknown or checking, force a fresh check
            if cert_status.status not in ['valid', 'warning', 'expired', 'error']:
                logger.info(f"Forcing fresh check for domain with unknown status: {domain['name']}")
                db.clear_cache(f"ssl_{domain['name']}")
                cert_status = check_certificate(domain['name'])

            certificates.append(cert_status)

    return render_template('ssl_certificates.html', certificates=certificates, user=current_user)

@app.route('/domain_expiry')
@auth.login_required
def domain_expiry():
    """Domain expiry monitoring page"""
    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    # Get domains for current organization that have domain expiry monitoring enabled
    domains_from_db = db.get_domains_by_organization(current_org['id'])
    domains = []

    for domain in domains_from_db:
        if domain['expiry_monitored']:
            domain_status = check_domain_expiry(domain['name'])
            domains.append(domain_status)

    return render_template('domain_expiry.html', domains=domains, user=current_user)

@app.route('/ping_monitoring')
@auth.login_required
def ping_monitoring():
    """Ping monitoring page"""
    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    # Get domains for current organization that have ping monitoring enabled
    domains_from_db = db.get_domains_by_organization(current_org['id'])
    ping_results = []

    for domain in domains_from_db:
        if domain['ping_monitored']:
            ping_result = check_ping(domain['name'])
            # Create a new dictionary with all the ping result data plus the host
            result_with_host = {
                'host': domain['name'],
                'status': ping_result['status'],
                'response_time': ping_result['response_time'],
                'last_checked': ping_result['last_checked']
            }
            ping_results.append(result_with_host)

    ping_stats = get_ping_stats(ping_results)
    return render_template('ping_monitoring.html', ping_results=ping_results, ping_stats=ping_stats, user=current_user)

@app.route('/domain/<int:domain_id>')
@auth.login_required
def domain_details(domain_id):
    """Domain details page showing comprehensive information about a specific domain"""
    try:
        start_time = time.time()
        logger.debug(f"Starting domain details page load for domain ID {domain_id}")

        # Log the function name to verify it's being called
        logger.info(f"domain_details function called with domain_id={domain_id}")

        # Get current user
        current_user = auth.get_current_user()

        # Get current organization
        current_org = current_user.get('current_organization')
        if not current_org:
            flash("You don't have access to any organizations", 'error')
            return redirect(url_for('profile'))

        # Get domain by ID
        domain = db.get_domain(domain_id)
        if not domain:
            flash('Domain not found', 'error')
            return redirect(url_for('index'))

        # Check if domain belongs to current organization
        if domain['organization_id'] != current_org['id']:
            flash("You don't have permission to view this domain", 'error')
            return redirect(url_for('index'))

        domain_to_get = domain['name']

        # Determine which monitoring services are enabled for this domain
        monitors = []

        # Check if domain is monitored for SSL
        ssl_status = None
        if domain['ssl_monitored']:
            monitors.append('ssl')
            ssl_status = check_certificate(domain_to_get)

        # Check if domain is monitored for expiry
        domain_status = None
        if domain['expiry_monitored']:
            monitors.append('expiry')
            domain_status = check_domain_expiry(domain_to_get)

        # Check if domain is monitored for ping
        ping_status = "unknown"
        ping_response_time = 0.0
        if domain['ping_monitored']:
            monitors.append('ping')
            # If we already have ping status from SSL or domain check, use that
            if ssl_status:
                ping_status = ssl_status.ping_status
            elif domain_status:
                ping_status = domain_status.ping_status
            else:
                # Only do a separate ping check if we don't already have the status
                ping_result = check_ping(domain_to_get)
                ping_status = ping_result["status"]
                ping_response_time = ping_result["response_time"]

        # Get uptime statistics
        uptime_percentage = db.calculate_uptime_percentage(domain_to_get)
        if uptime_percentage is not None:
            uptime_percentage = round(uptime_percentage, 1)

        # Get ping history for uptime segments
        ping_history = db.get_ping_history(domain_to_get, hours=12)

        # Create uptime segments (12 segments representing the last 12 hours)
        uptime_segments = []

        # For newly added domains, we want to show only the most recent segment as the current status
        # and all other segments as unknown
        if len(ping_history) <= 2:  # Changed from 1 to 2 to handle cases with 2 entries
            # If there's minimal history (0, 1, or 2 entries), use current status only for the most recent segment
            uptime_segments = ['unknown'] * 11 + [ping_status]
        else:
            # Group history into 12 hourly segments
            current_time = datetime.now()

            for i in range(12):
                segment_start = current_time - timedelta(hours=12-i)
                segment_end = current_time - timedelta(hours=11-i)

                # Find all checks in this segment
                segment_checks = [check for check in ping_history
                                if check['checked_at'] and
                                segment_start <= datetime.fromtimestamp(check['checked_at']) < segment_end]

                if segment_checks:
                    # If any check is down, the segment is down
                    if any(check['status'] == 'down' for check in segment_checks):
                        uptime_segments.append('down')
                    else:
                        uptime_segments.append('up')
                else:
                    # For the most recent segment (last hour), use current status if no data
                    if i == 11:  # Last segment (most recent hour)
                        uptime_segments.append(ping_status)
                    else:
                        uptime_segments.append('unknown')

        # Get timeframe parameter for the template
        timeframe_hours = request.args.get('timeframe', '24')
        try:
            timeframe_hours = int(timeframe_hours)
            if timeframe_hours not in [24, 168, 720]:  # 24h, 7d, 30d
                timeframe_hours = 24
        except ValueError:
            timeframe_hours = 24

        # Get 24-hour uptime percentage
        uptime_24h = db.calculate_uptime_percentage(domain_to_get, hours=24)
        if uptime_24h is not None:
            uptime_24h = round(uptime_24h, 1)

        # Get 7-day uptime percentage
        uptime_7d = db.calculate_uptime_percentage(domain_to_get, hours=24*7)
        if uptime_7d is not None:
            uptime_7d = round(uptime_7d, 1)

        # Get 30-day uptime percentage
        uptime_30d = db.calculate_uptime_percentage(domain_to_get, hours=24*30)
        if uptime_30d is not None:
            uptime_30d = round(uptime_30d, 1)

        # Determine if this domain has any issues
        has_issues = False
        if ssl_status and ssl_status.status in ['warning', 'expired', 'error']:
            has_issues = True
        if domain_status and domain_status.status in ['warning', 'expired', 'error']:
            has_issues = True
        if ping_status == 'down':
            has_issues = True

        # Determine expiry status for display
        expiry_status = 'unknown'
        days_until_expiry = 0
        if domain_status:
            expiry_status = domain_status.status
            # Ensure days_until_expiry is an integer
            if domain_status.days_remaining is not None:
                days_until_expiry = domain_status.days_remaining
            else:
                days_until_expiry = 0

        # Get user's timezone setting
        app_settings = get_app_settings()
        user_timezone = app_settings.timezone

        # Convert dates to user's timezone if needed
        if ssl_status and ssl_status.expiry_date:
            ssl_status.expiry_date = convert_to_user_timezone(ssl_status.expiry_date, user_timezone)

        if domain_status and domain_status.expiry_date:
            domain_status.expiry_date = convert_to_user_timezone(domain_status.expiry_date, user_timezone)

        # Perform a fresh ping check to ensure we have the latest data
        fresh_ping_result = check_ping(domain_to_get)
        ping_status = fresh_ping_result.get('status', 'unknown')
        ping_response_time = fresh_ping_result.get('response_time', 0)

        # Create domain data object with safe defaults for None values
        domain_data = {
            'id': domain_id,
            'name': domain_to_get,
            'monitors': monitors,
            'ssl_status': ssl_status if ssl_status is not None else {
                'status': 'unknown',
                'days_remaining': 0,
                'expiry_date': None
            },
            'domain_status': domain_status if domain_status is not None else {
                'status': 'unknown',
                'days_remaining': 0,
                'expiry_date': None,
                'registrar': 'Unknown'
            },
            'has_issues': has_issues,
            'expiry_status': expiry_status,
            'days_until_expiry': days_until_expiry,
            'health_status': ping_status,
            'response_time': ping_response_time,
            'timeframe_hours': timeframe_hours,
            'uptime_percentage': uptime_percentage,
            'uptime_24h': uptime_24h,
            'uptime_7d': uptime_7d,
            'uptime_30d': uptime_30d,
            'uptime_segments': uptime_segments
        }

        end_time = time.time()
        logger.debug(f"Domain details page for {domain_to_get} loaded in {end_time - start_time:.2f} seconds")

        # Get user's timezone setting
        app_settings = get_app_settings()
        user_timezone = app_settings.timezone

        return render_template('domain_details.html',
                              domain=domain_data,
                              user=current_user,
                              user_timezone=user_timezone)
    except Exception as e:
        logger.error(f"Error loading domain details page: {str(e)}", exc_info=True)
        flash(f'Error loading domain details: {str(e)}', 'error')
        return redirect(url_for('index'))

@app.route('/add_ssl_domain', methods=['POST'])
def add_ssl_domain():
    domain = request.form.get('domain')

    if not domain:
        flash('Domain name is required', 'error')
        return redirect(url_for('ssl_certificates'))

    # Clean the domain
    domain = clean_domain_name(domain)
    if not domain:
        flash('Invalid domain format', 'error')
        return redirect(url_for('ssl_certificates'))

    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    # Check if domain already exists in this organization
    existing_domain = db.get_domain_by_name_and_org(domain, current_org['id'])
    if existing_domain:
        # If domain exists, update it to enable SSL monitoring
        db.update_domain(existing_domain['id'], domain, True, existing_domain['expiry_monitored'], existing_domain['ping_monitored'])

        # Log the domain update action
        db.log_user_action(
            user_id=current_user['id'],
            username=current_user['username'],
            action_type='update',
            resource_type='domain',
            resource_id=existing_domain['id'],
            resource_name=domain,
            details=f'Updated domain to include SSL monitoring',
            ip_address=request.remote_addr,
            organization_id=current_org['id']
        )

        flash(f'Domain {domain} updated to include SSL monitoring', 'success')
    else:
        # Add new domain with SSL monitoring enabled and auto-logging
        domain_id = db.add_domain(
            name=domain,
            organization_id=current_org['id'],
            ssl_monitored=True,
            expiry_monitored=False,
            ping_monitored=False,
            user_id=current_user['id'],
            username=current_user['username'],
            ip_address=request.remote_addr,
            auto_log=True
        )

        if domain_id:
            logger.info(f"Successfully added domain {domain} with ID: {domain_id}")

            # Immediately check the SSL certificate status
            try:
                logger.info(f"Performing initial SSL certificate check for {domain}")
                # Clear any existing cache to ensure a fresh check
                db.clear_cache(f"ssl_{domain}")

                # Force a synchronous check of the certificate and store the result
                cert_status = check_certificate(domain)
                logger.info(f"Initial SSL certificate check completed for {domain}: {cert_status.status}")

                # Ensure the result is properly cached
                cache_ssl_data(domain, cert_status)

                # Record the SSL check in the database
                try:
                    db.record_ssl_check(
                        domain_name=domain,
                        status=cert_status.status,
                        expiry_date=cert_status.expiry_date,
                        issuer="Initial check",
                        subject=domain
                    )
                    logger.info(f"Recorded initial SSL check for {domain} in database")
                except Exception as record_error:
                    logger.error(f"Error recording initial SSL check: {str(record_error)}", exc_info=True)

                flash(f'Domain {domain} added successfully for SSL monitoring', 'success')
            except Exception as e:
                logger.error(f"Error during initial SSL certificate check for {domain}: {str(e)}", exc_info=True)
                flash(f'Domain {domain} added successfully, but initial SSL check failed. Please refresh the page.', 'warning')
        else:
            logger.error(f"Failed to add domain {domain}")
            flash(f'Error adding domain {domain}', 'error')

    # Return a JavaScript response that will refresh the page
    response = make_response("""
    <html>
    <head>
        <title>Redirecting...</title>
        <script>
            // Wait a moment to ensure the certificate check completes
            setTimeout(function() {
                window.location.href = '/ssl_certificates';
            }, 3000);
        </script>
    </head>
    <body>
        <p>Processing your request... You will be redirected automatically.</p>
    </body>
    </html>
    """)
    return response

@app.route('/remove_ssl_domain/<domain>')
def remove_ssl_domain(domain):
    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    # Get domain by name and organization
    existing_domain = db.get_domain_by_name_and_org(domain, current_org['id'])
    if existing_domain:
        # If domain has other monitoring enabled, just disable SSL monitoring
        if existing_domain['expiry_monitored'] or existing_domain['ping_monitored']:
            db.update_domain(existing_domain['id'], domain, False, existing_domain['expiry_monitored'], existing_domain['ping_monitored'])

            # Log the domain update action
            db.log_user_action(
                user_id=current_user['id'],
                username=current_user['username'],
                action_type='update',
                resource_type='domain',
                resource_id=existing_domain['id'],
                resource_name=domain,
                details=f'Removed domain from SSL monitoring',
                ip_address=request.remote_addr,
                organization_id=current_org['id']
            )

            flash(f'Domain {domain} removed from SSL monitoring', 'success')
        else:
            # If no other monitoring is enabled, delete the domain
            db.delete_domain(existing_domain['id'])

            # Log the domain deletion action
            db.log_user_action(
                user_id=current_user['id'],
                username=current_user['username'],
                action_type='delete',
                resource_type='domain',
                resource_id=existing_domain['id'],
                resource_name=domain,
                details=f'Removed domain from all monitoring',
                ip_address=request.remote_addr,
                organization_id=current_org['id']
            )

            flash(f'Domain {domain} removed from all monitoring', 'success')
    else:
        flash(f'Domain {domain} not found', 'error')

    return redirect(url_for('ssl_certificates'))

@app.route('/refresh_ssl_certificate/<domain>')
def refresh_ssl_certificate(domain):
    """Refresh SSL certificate data for a specific domain"""
    try:
        domain = validate_domain(domain)
    except ValueError as e:
        logger.error(str(e))
        return jsonify({
            'success': False,
            'message': str(e)
        }), 400
    try:
        cert_status = check_certificate(domain)
        return jsonify({
            'success': True,
            'data': {
                'status': cert_status.status,
                'days_remaining': cert_status.days_remaining,
                'expiry_date': cert_status.expiry_date.strftime('%Y-%m-%d')
            }
        })
    except Exception as e:
        logger.error(f"Error refreshing SSL certificate for {domain}: {str(e)}")
        return jsonify({
            'success': False,
            'message': f"Error refreshing certificate: {str(e)}"
        }), 500

@app.route('/api/ssl/<domain>/refresh', methods=['POST'])
def api_refresh_ssl_certificate(domain):
    """API endpoint to refresh SSL certificate data for a specific domain"""
    try:
        # Clear SSL cache for this domain to force a fresh check
        db.clear_cache(f"ssl_{domain}")

        cert_status = check_certificate(domain)
        return jsonify({
            'success': True,
            'data': {
                'status': cert_status.status,
                'days_remaining': cert_status.days_remaining,
                'expiry_date': cert_status.expiry_date.strftime('%Y-%m-%d')
            }
        })
    except Exception as e:
        logger.error(f"Error refreshing SSL certificate for {domain}: {str(e)}")
        return jsonify({
            'success': False,
            'error': f"Error refreshing certificate: {str(e)}"
        }), 500

@app.route('/bulk_import_ssl', methods=['POST'])
def bulk_import_ssl():
    domains_text = request.form.get('domains')
    if not domains_text:
        flash('No domains provided', 'error')
        return redirect(url_for('ssl_certificates'))

    # Split by newlines, commas, or spaces
    domains = re.split(r'[\n,\s]+', domains_text)
    domains = [d.strip() for d in domains if d.strip()]

    if not domains:
        flash('No valid domains found', 'error')
        return redirect(url_for('ssl_certificates'))

    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    added_count = 0
    updated_count = 0
    skipped_count = 0

    for domain in domains:
        # Clean the domain
        domain = clean_domain_name(domain)
        if not domain:
            continue

        # Check if domain already exists in this organization
        existing_domain = db.get_domain_by_name_and_org(domain, current_org['id'])
        if existing_domain:
            # If domain exists and SSL monitoring is not enabled, update it
            if not existing_domain['ssl_monitored']:
                db.update_domain(existing_domain['id'], domain, True, existing_domain['expiry_monitored'], existing_domain['ping_monitored'])
                updated_count += 1
            else:
                skipped_count += 1
        else:
            # Add new domain with SSL monitoring enabled and auto-logging
            domain_id = db.add_domain(
                name=domain,
                organization_id=current_org['id'],
                ssl_monitored=True,
                expiry_monitored=False,
                ping_monitored=False,
                user_id=current_user['id'],
                username=current_user['username'],
                ip_address=request.remote_addr,
                auto_log=True
            )

            if domain_id:
                logger.info(f"Successfully added domain {domain} with ID: {domain_id} (bulk)")

                # Immediately check the SSL certificate status in the background
                try:
                    # Clear any existing cache to ensure a fresh check
                    db.clear_cache(f"ssl_{domain}")

                    # Force a synchronous check of the certificate and store the result
                    cert_status = check_certificate(domain)
                    logger.info(f"Initial SSL certificate check completed for {domain} (bulk): {cert_status.status}")

                    # Ensure the result is properly cached
                    cache_ssl_data(domain, cert_status)

                    # Record the SSL check in the database
                    try:
                        db.record_ssl_check(
                            domain_name=domain,
                            status=cert_status.status,
                            expiry_date=cert_status.expiry_date,
                            issuer="Initial check (bulk)",
                            subject=domain
                        )
                        logger.info(f"Recorded initial SSL check for {domain} (bulk) in database")
                    except Exception as record_error:
                        logger.error(f"Error recording initial SSL check for bulk import: {str(record_error)}", exc_info=True)
                except Exception as e:
                    logger.error(f"Error during initial SSL certificate check for {domain} (bulk): {str(e)}", exc_info=True)

                added_count += 1
            else:
                logger.error(f"Failed to add domain {domain} (bulk)")
                skipped_count += 1

    if added_count > 0 or updated_count > 0:
        message = []
        if added_count > 0:
            message.append(f'Added {added_count} new domains')
        if updated_count > 0:
            message.append(f'Updated {updated_count} existing domains')
        if skipped_count > 0:
            message.append(f'Skipped {skipped_count} domains')

        flash(', '.join(message) + ' for SSL monitoring', 'success')
    else:
        flash(f'No domains added or updated, skipped {skipped_count} domains', 'warning')

    # Return a JavaScript response that will refresh the page
    response = make_response("""
    <html>
    <head>
        <title>Redirecting...</title>
        <script>
            // Wait a moment to ensure the certificate checks complete
            setTimeout(function() {
                window.location.href = '/ssl_certificates';
            }, 5000);
        </script>
    </head>
    <body>
        <p>Processing your request... You will be redirected automatically.</p>
    </body>
    </html>
    """)
    return response

@app.route('/export_ssl_domains')
def export_ssl_domains():
    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    # Get domains for current organization that have SSL monitoring enabled
    domains_from_db = db.get_domains_by_organization(current_org['id'])
    domains = [domain['name'] for domain in domains_from_db if domain['ssl_monitored']]
    domains_text = '\n'.join(domains)

    response = app.response_class(
        response=domains_text,
        status=200,
        mimetype='text/plain'
    )
    response.headers["Content-Disposition"] = "attachment; filename=ssl_domains.txt"
    return response

@app.route('/bulk_refresh_ssl', methods=['POST'])
@auth.login_required
def bulk_refresh_ssl():
    """Refresh SSL certificate information for multiple domains"""
    domains = request.form.getlist('domains')

    if not domains:
        flash('No domains selected', 'error')
        return redirect(url_for('ssl_certificates'))

    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    # Refresh SSL certificate information for each domain
    success_count = 0
    error_count = 0

    for domain in domains:
        try:
            # Check if domain belongs to current organization
            domain_obj = db.get_domain_by_name_and_org(domain, current_org['id'])
            if not domain_obj:
                error_count += 1
                continue

            # Refresh SSL certificate information
            # Use the same function that's used for individual domain refresh
            ssl_status = check_certificate(domain)
            if ssl_status and hasattr(ssl_status, 'status'):
                success_count += 1
            else:
                error_count += 1
        except Exception as e:
            logger.error(f"Error refreshing SSL certificate for {domain}: {str(e)}")
            error_count += 1

    if success_count > 0:
        flash(f'Successfully refreshed SSL certificate information for {success_count} domain(s)', 'success')

    if error_count > 0:
        flash(f'Failed to refresh SSL certificate information for {error_count} domain(s)', 'error')

    return redirect(url_for('ssl_certificates'))

@app.route('/bulk_delete_ssl', methods=['POST'])
@auth.login_required
def bulk_delete_ssl():
    """Remove multiple domains from SSL certificate monitoring"""
    domains = request.form.getlist('domains')

    if not domains:
        flash('No domains selected', 'error')
        return redirect(url_for('ssl_certificates'))

    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    # Remove domains from SSL monitoring
    success_count = 0
    error_count = 0

    for domain in domains:
        try:
            # Check if domain belongs to current organization
            domain_obj = db.get_domain_by_name_and_org(domain, current_org['id'])
            if not domain_obj:
                error_count += 1
                continue

            # Update domain to disable SSL monitoring
            db.update_domain(
                domain_obj['id'],
                domain,
                False,  # ssl_monitored = False
                domain_obj['expiry_monitored'],
                domain_obj['ping_monitored']
            )

            # Log the domain update action
            db.log_user_action(
                user_id=current_user['id'],
                username=current_user['username'],
                action_type='update',
                resource_type='domain',
                resource_id=domain_obj['id'],
                resource_name=domain,
                details=f'Removed domain from SSL monitoring',
                ip_address=request.remote_addr,
                organization_id=current_org['id']
            )

            success_count += 1
        except Exception as e:
            logger.error(f"Error removing SSL monitoring for {domain}: {str(e)}")
            error_count += 1

    if success_count > 0:
        flash(f'Successfully removed {success_count} domain(s) from SSL monitoring', 'success')

    if error_count > 0:
        flash(f'Failed to remove {error_count} domain(s) from SSL monitoring', 'error')

    return redirect(url_for('ssl_certificates'))

@app.route('/add_expiry_domain', methods=['POST'])
def add_expiry_domain():
    domain = request.form.get('domain')

    if not domain:
        flash('Domain name is required', 'error')
        return redirect(url_for('domain_expiry'))

    # Clean the domain
    domain = clean_domain_name(domain)
    if not domain:
        flash('Invalid domain format', 'error')
        return redirect(url_for('domain_expiry'))

    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    # Check if domain already exists in this organization
    existing_domain = db.get_domain_by_name_and_org(domain, current_org['id'])
    if existing_domain:
        # If domain exists, update it to enable expiry monitoring
        db.update_domain(existing_domain['id'], domain, existing_domain['ssl_monitored'], True, existing_domain['ping_monitored'])

        # Log the domain update action
        db.log_user_action(
            user_id=current_user['id'],
            username=current_user['username'],
            action_type='update',
            resource_type='domain',
            resource_id=existing_domain['id'],
            resource_name=domain,
            details=f'Updated domain to include expiry monitoring',
            ip_address=request.remote_addr,
            organization_id=current_org['id']
        )

        flash(f'Domain {domain} updated to include expiry monitoring', 'success')
    else:
        # Add new domain with expiry monitoring enabled and auto-logging
        domain_id = db.add_domain(
            name=domain,
            organization_id=current_org['id'],
            ssl_monitored=False,
            expiry_monitored=True,
            ping_monitored=False,
            user_id=current_user['id'],
            username=current_user['username'],
            ip_address=request.remote_addr,
            auto_log=True
        )

        if domain_id:
            logger.info(f"Successfully added domain {domain} with ID: {domain_id}")

            # Immediately check the domain expiry status
            try:
                logger.info(f"Performing initial domain expiry check for {domain}")
                # Clear any existing cache to ensure a fresh check
                db.clear_cache(f"domain_expiry_{domain}")
                # Check the domain expiry
                check_domain_expiry(domain)
                logger.info(f"Initial domain expiry check completed for {domain}")
                flash(f'Domain {domain} added successfully for expiry monitoring', 'success')
            except Exception as e:
                logger.error(f"Error during initial domain expiry check for {domain}: {str(e)}", exc_info=True)
                flash(f'Domain {domain} added successfully, but initial expiry check failed. Please refresh the page.', 'warning')
        else:
            logger.error(f"Failed to add domain {domain}")
            flash(f'Error adding domain {domain}', 'error')

    return redirect(url_for('domain_expiry'))

@app.route('/remove_expiry_domain/<domain>')
def remove_expiry_domain(domain):
    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    # Get domain by name and organization
    existing_domain = db.get_domain_by_name_and_org(domain, current_org['id'])
    if existing_domain:
        # If domain has other monitoring enabled, just disable expiry monitoring
        if existing_domain['ssl_monitored'] or existing_domain['ping_monitored']:
            db.update_domain(existing_domain['id'], domain, existing_domain['ssl_monitored'], False, existing_domain['ping_monitored'])

            # Log the domain update action
            db.log_user_action(
                user_id=current_user['id'],
                username=current_user['username'],
                action_type='update',
                resource_type='domain',
                resource_id=existing_domain['id'],
                resource_name=domain,
                details=f'Removed domain from expiry monitoring',
                ip_address=request.remote_addr,
                organization_id=current_org['id']
            )

            flash(f'Domain {domain} removed from expiry monitoring', 'success')
        else:
            # If no other monitoring is enabled, delete the domain
            db.delete_domain(existing_domain['id'])

            # Log the domain deletion action
            db.log_user_action(
                user_id=current_user['id'],
                username=current_user['username'],
                action_type='delete',
                resource_type='domain',
                resource_id=existing_domain['id'],
                resource_name=domain,
                details=f'Removed domain from all monitoring',
                ip_address=request.remote_addr,
                organization_id=current_org['id']
            )

            flash(f'Domain {domain} removed from all monitoring', 'success')
    else:
        flash(f'Domain {domain} not found', 'error')

    return redirect(url_for('domain_expiry'))

@app.route('/api/expiry/<domain>/refresh', methods=['POST'])
def api_refresh_domain_expiry(domain):
    """API endpoint to refresh domain expiry data for a specific domain"""
    try:
        # Clear domain expiry cache for this domain to force a fresh check
        db.clear_cache(f"domain_expiry_{domain}")

        domain_status = check_domain_expiry(domain)
        return jsonify({
            'success': True,
            'data': {
                'status': domain_status.status,
                'days_remaining': domain_status.days_remaining,
                'expiry_date': domain_status.expiry_date.strftime('%Y-%m-%d'),
                'registrar': domain_status.registrar
            }
        })
    except Exception as e:
        logger.error(f"Error refreshing domain expiry for {domain}: {str(e)}")
        return jsonify({
            'success': False,
            'error': f"Error refreshing domain expiry: {str(e)}"
        }), 500



@app.route('/bulk_refresh_expiry', methods=['POST'])
@auth.login_required
def bulk_refresh_expiry():
    """Refresh domain expiry information for multiple domains"""
    domains = request.form.getlist('domains')

    if not domains:
        flash('No domains selected', 'error')
        return redirect(url_for('domain_expiry'))

    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    # Refresh domain expiry information for each domain
    success_count = 0
    error_count = 0

    for domain in domains:
        try:
            # Check if domain belongs to current organization
            domain_obj = db.get_domain_by_name_and_org(domain, current_org['id'])
            if not domain_obj:
                error_count += 1
                continue

            # Refresh domain expiry information
            domain_status = check_domain_expiry(domain)
            if domain_status:
                success_count += 1
            else:
                error_count += 1
        except Exception as e:
            logger.error(f"Error refreshing domain expiry for {domain}: {str(e)}")
            error_count += 1

    if success_count > 0:
        flash(f'Successfully refreshed domain expiry information for {success_count} domain(s)', 'success')

    if error_count > 0:
        flash(f'Failed to refresh domain expiry information for {error_count} domain(s)', 'error')

    return redirect(url_for('domain_expiry'))

@app.route('/bulk_delete_expiry', methods=['POST'])
@auth.login_required
def bulk_delete_expiry():
    """Remove multiple domains from domain expiry monitoring"""
    domains = request.form.getlist('domains')

    if not domains:
        flash('No domains selected', 'error')
        return redirect(url_for('domain_expiry'))

    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    # Remove domains from expiry monitoring
    success_count = 0
    error_count = 0

    for domain in domains:
        try:
            # Check if domain belongs to current organization
            domain_obj = db.get_domain_by_name_and_org(domain, current_org['id'])
            if not domain_obj:
                error_count += 1
                continue

            # Update domain to disable expiry monitoring
            db.update_domain(
                domain_obj['id'],
                domain,
                domain_obj['ssl_monitored'],
                False,  # expiry_monitored = False
                domain_obj['ping_monitored']
            )

            # Log the domain update action
            db.log_user_action(
                user_id=current_user['id'],
                username=current_user['username'],
                action_type='update',
                resource_type='domain',
                resource_id=domain_obj['id'],
                resource_name=domain,
                details=f'Removed domain from expiry monitoring',
                ip_address=request.remote_addr,
                organization_id=current_org['id']
            )

            success_count += 1
        except Exception as e:
            logger.error(f"Error removing expiry monitoring for {domain}: {str(e)}")
            error_count += 1

    if success_count > 0:
        flash(f'Successfully removed {success_count} domain(s) from expiry monitoring', 'success')

    if error_count > 0:
        flash(f'Failed to remove {error_count} domain(s) from expiry monitoring', 'error')

    return redirect(url_for('domain_expiry'))

@app.route('/bulk_import_expiry', methods=['POST'])
def bulk_import_expiry():
    domains_text = request.form.get('domains')
    if not domains_text:
        flash('No domains provided', 'error')
        return redirect(url_for('domain_expiry'))

    # Split by newlines, commas, or spaces
    domains = re.split(r'[\n,\s]+', domains_text)
    domains = [d.strip() for d in domains if d.strip()]

    if not domains:
        flash('No valid domains found', 'error')
        return redirect(url_for('domain_expiry'))

    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    added_count = 0
    updated_count = 0
    skipped_count = 0

    for domain in domains:
        # Clean the domain
        domain = clean_domain_name(domain)
        if not domain:
            continue

        # Check if domain already exists in this organization
        existing_domain = db.get_domain_by_name_and_org(domain, current_org['id'])
        if existing_domain:
            # If domain exists and expiry monitoring is not enabled, update it
            if not existing_domain['expiry_monitored']:
                db.update_domain(existing_domain['id'], domain, existing_domain['ssl_monitored'], True, existing_domain['ping_monitored'])
                updated_count += 1
            else:
                skipped_count += 1
        else:
            # Add new domain with expiry monitoring enabled and auto-logging
            domain_id = db.add_domain(
                name=domain,
                organization_id=current_org['id'],
                ssl_monitored=False,
                expiry_monitored=True,
                ping_monitored=False,
                user_id=current_user['id'],
                username=current_user['username'],
                ip_address=request.remote_addr,
                auto_log=True
            )

            if domain_id:
                logger.info(f"Successfully added domain {domain} with ID: {domain_id} (bulk expiry)")

                # Immediately check the domain expiry status in the background
                try:
                    # Clear any existing cache to ensure a fresh check
                    db.clear_cache(f"domain_expiry_{domain}")
                    # Check the domain expiry (this will cache the result)
                    check_domain_expiry(domain)
                    logger.info(f"Initial domain expiry check completed for {domain} (bulk)")
                except Exception as e:
                    logger.error(f"Error during initial domain expiry check for {domain} (bulk): {str(e)}", exc_info=True)

                added_count += 1
            else:
                logger.error(f"Failed to add domain {domain} (bulk expiry)")
                skipped_count += 1

    if added_count > 0 or updated_count > 0:
        message = []
        if added_count > 0:
            message.append(f'Added {added_count} new domains')
        if updated_count > 0:
            message.append(f'Updated {updated_count} existing domains')
        if skipped_count > 0:
            message.append(f'Skipped {skipped_count} domains')

        flash(', '.join(message) + ' for expiry monitoring', 'success')
    else:
        flash(f'No domains added or updated, skipped {skipped_count} domains', 'warning')

    return redirect(url_for('domain_expiry'))

@app.route('/export_expiry_domains')
def export_expiry_domains():
    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    # Get domains for current organization that have expiry monitoring enabled
    domains_from_db = db.get_domains_by_organization(current_org['id'])
    domains = [domain['name'] for domain in domains_from_db if domain['expiry_monitored']]
    domains_text = '\n'.join(domains)

    response = app.response_class(
        response=domains_text,
        status=200,
        mimetype='text/plain'
    )
    response.headers["Content-Disposition"] = "attachment; filename=domain_expiry.txt"
    return response

@app.route('/add_ping_host', methods=['POST'])
def add_ping_host():
    host = request.form.get('host')

    if not host:
        flash('Host is required', 'error')
        return redirect(url_for('ping_monitoring'))

    # Clean the host
    host = clean_domain_name(host)
    if not host:
        flash('Invalid host format', 'error')
        return redirect(url_for('ping_monitoring'))

    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    # Check if host already exists in this organization
    existing_domain = db.get_domain_by_name_and_org(host, current_org['id'])
    if existing_domain:
        # If domain exists, update it to enable ping monitoring
        db.update_domain(existing_domain['id'], host, existing_domain['ssl_monitored'], existing_domain['expiry_monitored'], True)

        # Log the domain update action
        db.log_user_action(
            user_id=current_user['id'],
            username=current_user['username'],
            action_type='update',
            resource_type='domain',
            resource_id=existing_domain['id'],
            resource_name=host,
            details=f'Updated host to include ping monitoring',
            ip_address=request.remote_addr,
            organization_id=current_org['id']
        )

        flash(f'Host {host} updated to include ping monitoring', 'success')
    else:
        # Add new domain with ping monitoring enabled and auto-logging
        domain_id = db.add_domain(
            name=host,
            organization_id=current_org['id'],
            ssl_monitored=False,
            expiry_monitored=False,
            ping_monitored=True,
            user_id=current_user['id'],
            username=current_user['username'],
            ip_address=request.remote_addr,
            auto_log=True
        )

        if domain_id:
            logger.info(f"Successfully added host {host} with ID: {domain_id}")

            # Immediately check the ping status
            try:
                logger.info(f"Performing initial ping check for {host}")
                # Check the ping status
                ping_result = check_ping(host)
                logger.info(f"Initial ping check completed for {host}: {ping_result['status']}")
                flash(f'Host {host} added successfully for ping monitoring', 'success')
            except Exception as e:
                logger.error(f"Error during initial ping check for {host}: {str(e)}", exc_info=True)
                flash(f'Host {host} added successfully, but initial ping check failed. Please refresh the page.', 'warning')
        else:
            logger.error(f"Failed to add host {host}")
            flash(f'Error adding host {host}', 'error')

    return redirect(url_for('ping_monitoring'))

@app.route('/remove_ping_host/<host>')
def remove_ping_host(host):
    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    # Get domain by name and organization
    existing_domain = db.get_domain_by_name_and_org(host, current_org['id'])
    if existing_domain:
        # If domain has other monitoring enabled, just disable ping monitoring
        if existing_domain['ssl_monitored'] or existing_domain['expiry_monitored']:
            db.update_domain(existing_domain['id'], host, existing_domain['ssl_monitored'], existing_domain['expiry_monitored'], False)

            # Log the domain update action
            db.log_user_action(
                user_id=current_user['id'],
                username=current_user['username'],
                action_type='update',
                resource_type='domain',
                resource_id=existing_domain['id'],
                resource_name=host,
                details=f'Removed host from ping monitoring',
                ip_address=request.remote_addr,
                organization_id=current_org['id']
            )

            flash(f'Host {host} removed from ping monitoring', 'success')
        else:
            # If no other monitoring is enabled, delete the domain
            db.delete_domain(existing_domain['id'])

            # Log the domain deletion action
            db.log_user_action(
                user_id=current_user['id'],
                username=current_user['username'],
                action_type='delete',
                resource_type='domain',
                resource_id=existing_domain['id'],
                resource_name=host,
                details=f'Removed host from all monitoring',
                ip_address=request.remote_addr,
                organization_id=current_org['id']
            )

            flash(f'Host {host} removed from all monitoring', 'success')
    else:
        flash(f'Host {host} not found', 'error')

    return redirect(url_for('ping_monitoring'))

@app.route('/api/ping/<host>/refresh', methods=['POST'])
def api_refresh_ping_host(host):
    """API endpoint to refresh ping status for a specific host"""
    try:
        # Clear ping cache for this host to force a fresh check
        db.clear_cache(f"ping_{host}")

        ping_result = check_ping(host)

        # Get uptime percentage
        uptime_percentage = db.calculate_uptime_percentage(host)
        if uptime_percentage is not None:
            uptime_percentage = round(uptime_percentage, 1)

        # Get ping history for uptime segments
        ping_history = db.get_ping_history(host, hours=12)

        # Create uptime segments (12 segments representing the last 12 hours)
        uptime_segments = []
        if ping_history:
            # Group history into 12 hourly segments
            current_time = datetime.now()

            for i in range(12):
                segment_start = current_time - timedelta(hours=12-i)
                segment_end = current_time - timedelta(hours=11-i)

                # Find all checks in this segment
                segment_checks = [check for check in ping_history
                                if check['checked_at'] and
                                segment_start <= datetime.fromtimestamp(check['checked_at']) < segment_end]

                if segment_checks:
                    # If any check is down, the segment is down
                    if any(check['status'] == 'down' for check in segment_checks):
                        uptime_segments.append('down')
                    else:
                        uptime_segments.append('up')
                else:
                    # No checks in this segment
                    uptime_segments.append('unknown')
        else:
            # If no history, use current status only for the most recent segment (last hour)
            uptime_segments = ['unknown'] * 11 + [ping_result["status"]]

        # Get response history from config
        config = load_config()
        ping_hosts = config.get('ping_hosts', [])
        response_history = []

        for ping_host in ping_hosts:
            if ping_host.get('host') == host:
                # Initialize response history if it doesn't exist
                if 'response_history' not in ping_host:
                    ping_host['response_history'] = []

                # Add new response time to history (keep last 20 entries)
                if ping_result["status"] == "up":
                    ping_host['response_history'].append(ping_result["response_time"])
                    # Keep only the last 20 entries
                    ping_host['response_history'] = ping_host['response_history'][-20:]

                response_history = ping_host.get('response_history', [])

                # Save the updated config
                save_config(config)
                break

        return jsonify({
            'success': True,
            'data': {
                'host': host,
                'ping_status': ping_result["status"],
                'response_time': ping_result["response_time"],
                'response_history': response_history,
                'uptime_percentage': uptime_percentage,
                'uptime_segments': uptime_segments,
                'last_checked': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
        })
    except Exception as e:
        logger.error(f"Error refreshing ping status for {host}: {str(e)}")
        return jsonify({
            'success': False,
            'error': f"Error refreshing ping status: {str(e)}"
        }), 500



@app.route('/bulk_refresh_ping', methods=['POST'])
@auth.login_required
def bulk_refresh_ping():
    """Refresh ping status for multiple hosts"""
    hosts = request.form.getlist('hosts')

    if not hosts:
        flash('No hosts selected', 'error')
        return redirect(url_for('ping_monitoring'))

    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    # Refresh ping status for each host
    success_count = 0
    error_count = 0

    for host in hosts:
        try:
            # Check if host belongs to current organization
            host_obj = db.get_ping_host_by_name_and_org(host, current_org['id'])
            if not host_obj:
                error_count += 1
                continue

            # Refresh ping status
            ping_result = check_ping(host)
            if ping_result:
                success_count += 1
            else:
                error_count += 1
        except Exception as e:
            logger.error(f"Error refreshing ping status for {host}: {str(e)}")
            error_count += 1

    if success_count > 0:
        flash(f'Successfully refreshed ping status for {success_count} host(s)', 'success')

    if error_count > 0:
        flash(f'Failed to refresh ping status for {error_count} host(s)', 'error')

    return redirect(url_for('ping_monitoring'))

@app.route('/bulk_delete_ping', methods=['POST'])
@auth.login_required
def bulk_delete_ping():
    """Remove multiple hosts from ping monitoring"""
    hosts = request.form.getlist('hosts')

    if not hosts:
        flash('No hosts selected', 'error')
        return redirect(url_for('ping_monitoring'))

    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    # Remove hosts from ping monitoring
    success_count = 0
    error_count = 0

    for host in hosts:
        try:
            # Check if host belongs to current organization
            host_obj = db.get_domain_by_name_and_org(host, current_org['id'])
            if not host_obj:
                error_count += 1
                continue

            # Update domain to disable ping monitoring
            db.update_domain(
                host_obj['id'],
                host,
                host_obj['ssl_monitored'],
                host_obj['expiry_monitored'],
                False  # ping_monitored = False
            )

            # Log the domain update action
            db.log_user_action(
                user_id=current_user['id'],
                username=current_user['username'],
                action_type='update',
                resource_type='domain',
                resource_id=host_obj['id'],
                resource_name=host,
                details=f'Removed host from ping monitoring',
                ip_address=request.remote_addr,
                organization_id=current_org['id']
            )

            success_count += 1
        except Exception as e:
            logger.error(f"Error removing ping monitoring for {host}: {str(e)}")
            error_count += 1

    if success_count > 0:
        flash(f'Successfully removed {success_count} host(s) from ping monitoring', 'success')

    if error_count > 0:
        flash(f'Failed to remove {error_count} host(s) from ping monitoring', 'error')

    return redirect(url_for('ping_monitoring'))

@app.route('/bulk_import_ping', methods=['POST'])
def bulk_import_ping():
    hosts_text = request.form.get('hosts')
    if not hosts_text:
        flash('No hosts provided', 'error')
        return redirect(url_for('ping_monitoring'))

    # Split by newlines, commas, or spaces
    hosts = re.split(r'[\n,\s]+', hosts_text)
    hosts = [h.strip() for h in hosts if h.strip()]

    if not hosts:
        flash('No valid hosts found', 'error')
        return redirect(url_for('ping_monitoring'))

    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    added_count = 0
    updated_count = 0
    skipped_count = 0

    for host in hosts:
        # Clean the host
        host = clean_domain_name(host)
        if not host:
            continue

        # Check if host already exists in this organization
        existing_domain = db.get_domain_by_name_and_org(host, current_org['id'])
        if existing_domain:
            # If domain exists and ping monitoring is not enabled, update it
            if not existing_domain['ping_monitored']:
                db.update_domain(existing_domain['id'], host, existing_domain['ssl_monitored'], existing_domain['expiry_monitored'], True)

                # Log the domain update action
                db.log_user_action(
                    user_id=current_user['id'],
                    username=current_user['username'],
                    action_type='update',
                    resource_type='domain',
                    resource_id=existing_domain['id'],
                    resource_name=host,
                    details=f'Updated host to enable ping monitoring',
                    ip_address=request.remote_addr,
                    organization_id=current_org['id']
                )

                updated_count += 1
            else:
                skipped_count += 1
        else:
            # Add new domain with ping monitoring enabled and auto-logging
            domain_id = db.add_domain(
                name=host,
                organization_id=current_org['id'],
                ssl_monitored=False,
                expiry_monitored=False,
                ping_monitored=True,
                user_id=current_user['id'],
                username=current_user['username'],
                ip_address=request.remote_addr,
                auto_log=True
            )

            if domain_id:
                logger.info(f"Successfully added host {host} with ID: {domain_id} (bulk ping)")

                # Immediately check the ping status in the background
                try:
                    # Check the ping status
                    ping_result = check_ping(host)
                    logger.info(f"Initial ping check completed for {host} (bulk): {ping_result['status']}")
                except Exception as e:
                    logger.error(f"Error during initial ping check for {host} (bulk): {str(e)}", exc_info=True)

                added_count += 1
            else:
                logger.error(f"Failed to add host {host} (bulk ping)")
                skipped_count += 1

    if added_count > 0 or updated_count > 0:
        message = []
        if added_count > 0:
            message.append(f'Added {added_count} new hosts')
        if updated_count > 0:
            message.append(f'Updated {updated_count} existing hosts')
        if skipped_count > 0:
            message.append(f'Skipped {skipped_count} hosts')

        flash(', '.join(message) + ' for ping monitoring', 'success')
    else:
        flash(f'No hosts added or updated, skipped {skipped_count} hosts', 'warning')

    return redirect(url_for('ping_monitoring'))

@app.route('/export_ping_hosts')
def export_ping_hosts():
    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    # Get domains for current organization that have ping monitoring enabled
    domains_from_db = db.get_domains_by_organization(current_org['id'])
    hosts = [domain['name'] for domain in domains_from_db if domain['ping_monitored']]
    hosts_text = '\n'.join(hosts)

    response = app.response_class(
        response=hosts_text,
        status=200,
        mimetype='text/plain'
    )
    response.headers["Content-Disposition"] = "attachment; filename=ping_hosts.txt"
    return response

@app.route('/api/ping/<host>/response_history')
def api_ping_response_history(host):
    """API endpoint to get ping response time history for a specific host"""
    try:
        # Get timeframe from query parameter
        timeframe_hours = request.args.get('timeframe', '24')
        try:
            timeframe_hours = int(timeframe_hours)
            if timeframe_hours not in [24, 168, 720]:  # 24h, 7d, 30d
                timeframe_hours = 24
        except ValueError:
            timeframe_hours = 24

        # Perform a fresh ping check to ensure we have the latest data
        fresh_ping_result = check_ping(host)
        logger.debug(f"Fresh ping check for {host}: {fresh_ping_result}")

        # Get response time history
        response_history = db.get_ping_response_history(host, hours=timeframe_hours)

        # Add the fresh ping result to the response history if it was successful
        if fresh_ping_result.get('status') == 'up':
            # Create a new entry with the current timestamp
            fresh_entry = {
                'timestamp': int(time.time() * 1000),  # Current time in milliseconds
                'response_time': fresh_ping_result.get('response_time', 0),
                'formatted_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'iso_time': datetime.now().isoformat() + 'Z'  # Add 'Z' to indicate UTC
            }

            # Add the fresh entry to the response history
            response_history.append(fresh_entry)

            # Sort the response history by timestamp
            response_history.sort(key=lambda x: x.get('timestamp', 0))

        return jsonify({
            'success': True,
            'data': response_history
        })
    except Exception as e:
        logger.error(f"Error getting ping response history for {host}: {str(e)}")
        return jsonify({
            'success': False,
            'error': f"Error getting ping response history: {str(e)}"
        }), 500



@app.route('/app_settings', methods=['GET', 'POST'])
@auth.login_required
def app_settings():
    """Application settings page"""
    # Get current user and organization
    user = auth.get_current_user()
    current_org = user.get('current_organization')

    # We've removed the refresh button, but keeping this code in case it's needed in the future
    if request.args.get('refresh') == 'true':
        # Get app settings
        settings = get_app_settings()
        warning_threshold = settings.warning_threshold_days

        # Get all domains
        all_domains = []
        config = load_config()

        # Add SSL domains
        for entry in config.get('ssl_domains', []):
            domain = entry.get('url')
            if domain and domain not in all_domains:
                all_domains.append(domain)

        # Add domain expiry domains
        for entry in config.get('domain_expiry', []):
            domain = entry.get('name')
            if domain and domain not in all_domains:
                all_domains.append(domain)

        # Update status for all domains based on current warning threshold
        domains_updated = 0
        for domain in all_domains:
            # Update SSL certificate status
            ssl_updated = update_ssl_status(domain, warning_threshold)
            # Update domain expiry status
            domain_updated = update_domain_expiry_status(domain, warning_threshold)

            if ssl_updated or domain_updated:
                domains_updated += 1

        if domains_updated > 0:
            flash(f'Status updated for {domains_updated} domains using current warning threshold of {warning_threshold} days.', 'success')
        else:
            flash(f'No status changes needed with current warning threshold of {warning_threshold} days.', 'info')

        return redirect(url_for('app_settings'))

    if request.method == 'POST':
        form_type = request.form.get('form_type')
        config = load_config()

        if form_type == 'monitoring_settings':
            # Update app settings
            if 'app_settings' not in config:
                config['app_settings'] = {}

            # Handle warning threshold safely
            try:
                warning_threshold = int(request.form.get('warning_threshold', 10))
            except (ValueError, TypeError):
                logger.warning(f"Invalid warning_threshold value: {request.form.get('warning_threshold')}. Using default value of 10.")
                warning_threshold = 10

            # Handle auto refresh interval safely
            try:
                auto_refresh_interval = int(request.form.get('auto_refresh_interval', 5))
            except (ValueError, TypeError):
                logger.warning(f"Invalid auto_refresh_interval value: {request.form.get('auto_refresh_interval')}. Using default value of 5.")
                auto_refresh_interval = 5

            # Get timezone setting
            timezone = request.form.get('timezone', 'UTC')
            # Validate timezone (use UTC as fallback if invalid)
            try:
                import pytz
                if timezone not in pytz.all_timezones:
                    logger.warning(f"Invalid timezone: {timezone}. Using default value of UTC.")
                    timezone = 'UTC'
            except ImportError:
                # If pytz is not available, accept any timezone value
                logger.warning("pytz not available, timezone validation skipped")

            # Check if warning threshold has changed
            old_warning_threshold = config['app_settings'].get('warning_threshold_days', 10)
            threshold_changed = int(old_warning_threshold) != warning_threshold

            if threshold_changed:
                logger.info(f"Warning threshold changed from {old_warning_threshold} to {warning_threshold}. Clearing all certificate and domain expiry caches.")

            # Update app settings
            app_settings = {
                'warning_threshold_days': warning_threshold,
                'auto_refresh_enabled': 'auto_refresh_enabled' in request.form,
                'auto_refresh_interval': auto_refresh_interval,
                'timezone': timezone,
                'theme': config['app_settings'].get('theme', 'light')  # Preserve existing theme
            }

            # Save app settings directly to database
            db.set_setting('app_settings', app_settings)

            # Update the config object for compatibility
            config['app_settings'] = app_settings

            # Also update the email notification settings to use the same warning threshold
            if 'notifications' in config and 'email' in config['notifications']:
                config['notifications']['email']['warning_threshold_days'] = warning_threshold

            # Always update the status of all cached SSL and domain expiry data when settings are saved
            # Get all domains
            all_domains = []

            # Add SSL domains
            for entry in config.get('ssl_domains', []):
                domain = entry.get('url')
                if domain and domain not in all_domains:
                    all_domains.append(domain)

            # Add domain expiry domains
            for entry in config.get('domain_expiry', []):
                domain = entry.get('name')
                if domain and domain not in all_domains:
                    all_domains.append(domain)

            # Update status for all domains based on new warning threshold
            domains_updated = 0
            for domain in all_domains:
                # Update SSL certificate status
                ssl_updated = update_ssl_status(domain, warning_threshold)
                # Update domain expiry status
                domain_updated = update_domain_expiry_status(domain, warning_threshold)

                if ssl_updated or domain_updated:
                    domains_updated += 1

            if threshold_changed:
                if domains_updated > 0:
                    flash(f'Warning threshold updated to {warning_threshold} days. Status updated for {domains_updated} domains.', 'success')
                else:
                    flash(f'Warning threshold updated to {warning_threshold} days.', 'success')
            else:
                flash('Settings saved successfully.', 'success')

            # Get WHOIS API key
            whois_api_key = request.form.get('whois_api_key', '').strip()

            # Save the WHOIS API key for the current organization
            if current_org:
                org_id = current_org['id']
                db.set_organization_setting(org_id, 'whois_api_key', whois_api_key)

                # Log the API key status (without revealing the actual key)
                if whois_api_key:
                    logger.info(f"WHOIS API key updated successfully for organization {org_id}")
                    # Removed success notification for WHOIS API key
                else:
                    logger.warning(f"WHOIS API key was cleared or not provided for organization {org_id}")
                    flash('Warning: No WHOIS API key provided. Domain expiry monitoring will not work correctly.', 'warning')
            else:
                # Fallback to global settings if no organization is selected
                if 'api_settings' not in config:
                    config['api_settings'] = {}

                config['api_settings']['whois_api_key'] = whois_api_key
                db.set_setting('api_settings', {'whois_api_key': whois_api_key})

                if whois_api_key:
                    logger.info("WHOIS API key updated successfully in global settings")
                    # Removed success notification for WHOIS API key
                else:
                    logger.warning("WHOIS API key was cleared or not provided in global settings")
                    flash('Warning: No WHOIS API key provided. Domain expiry monitoring will not work correctly.', 'warning')

            save_config(config)
            flash('Settings updated successfully', 'success')

        return redirect(url_for('app_settings'))

    # Get app settings directly from the database to ensure we have the latest values
    settings = get_app_settings()

    # Use the warning threshold from app settings
    warning_threshold = settings.warning_threshold_days

    # Log the warning threshold for debugging
    logger.info(f"Retrieved warning threshold from app settings: {warning_threshold}")

    # Get WHOIS API key - first try organization-specific, then fall back to global
    whois_api_key = ''
    if current_org:
        org_id = current_org['id']
        whois_api_key = db.get_organization_setting(org_id, 'whois_api_key', '')

    # If no organization-specific key, check global settings
    if not whois_api_key:
        api_settings = db.get_setting('api_settings', {})
        whois_api_key = api_settings.get('whois_api_key', '')

    # Settings for the template
    # Get timezone display name
    timezone_display = settings.timezone
    try:
        import pytz
        if settings.timezone in pytz.all_timezones:
            # Find the city/region name from the timezone
            timezone_parts = settings.timezone.split('/')
            if len(timezone_parts) > 1:
                city = timezone_parts[-1].replace('_', ' ')
                # Get UTC offset
                tz = pytz.timezone(settings.timezone)
                utc_offset = datetime.now(tz).strftime('%z')
                # Format offset as +/-HH:MM
                if utc_offset:
                    offset_hours = utc_offset[:3]
                    offset_minutes = utc_offset[3:]
                    formatted_offset = f"{offset_hours}:{offset_minutes}"
                    timezone_display = f"{city} (UTC {formatted_offset})"
                else:
                    timezone_display = f"{city} (UTC)"
            # Special case for UTC
            elif settings.timezone == 'UTC':
                timezone_display = "UTC (+00:00)"
    except (ImportError, Exception) as e:
        logger.warning(f"Error formatting timezone display name: {str(e)}")

    combined_settings = {
        'auto_refresh_enabled': settings.auto_refresh_enabled,
        'auto_refresh_interval': settings.auto_refresh_interval,
        'theme': settings.theme,
        'timezone': settings.timezone,
        'timezone_display': timezone_display,
        'warning_threshold_days': warning_threshold,
        'whois_api_key': whois_api_key
    }

    return render_template('app_settings.html',
                          settings=combined_settings,
                          data_dir=DATA_DIR,
                          config_file=CONFIG_FILE,
                          current_org=current_org)

@app.route('/backup_config')
def backup_config():
    """Download a backup of the current configuration"""
    if not os.path.exists(CONFIG_FILE):
        flash('No configuration file found to backup', 'error')
        return redirect(url_for('app_settings'))

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return send_file(CONFIG_FILE,
                    mimetype='application/x-yaml',
                    as_attachment=True,
                    download_name=f'certifly_backup_{timestamp}.yaml')

@app.route('/clear_whois_cache')
@auth.login_required
def clear_whois_cache():
    """Clear the WHOIS cache and domain expiry cache"""
    try:
        # Clear all WHOIS cache entries from database
        db.clear_cache_by_prefix('whois_')

        # Clear domain expiry cache from database
        db.clear_cache_by_prefix('domain_expiry_')

        # Remove the cache file if it exists
        cache_file = os.path.join(DATA_DIR, "domain_expiry_cache.json")
        if os.path.exists(cache_file):
            try:
                os.remove(cache_file)
                logger.info("Domain expiry cache file removed")
            except OSError as e:
                logger.error(f"Error removing domain expiry cache file: {e}")

        # Check if WHOIS API key is configured
        api_key = get_whois_api_key()
        if not api_key:
            flash('Warning: No WHOIS API key configured. Please add a WHOIS API key in settings.', 'warning')
        else:
            flash('WHOIS and domain expiry caches cleared successfully', 'success')
    except Exception as e:
        flash(f'Error clearing caches: {str(e)}', 'error')

    return redirect(url_for('app_settings'))

@app.route('/clear_ssl_cache/<domain>')
@auth.login_required
def clear_ssl_cache(domain):
    """Clear the SSL cache for a specific domain"""
    try:
        # Clear SSL cache for this domain
        cache_key = f"ssl_{domain}"
        db.clear_cache(cache_key)
        logger.info(f"SSL cache cleared for domain: {domain}")

        # Force a fresh check
        cert_status = check_certificate(domain)
        logger.info(f"Fresh SSL check for {domain} after clearing cache: {cert_status.status}")

        flash(f'SSL cache for {domain} cleared successfully. New status: {cert_status.status}', 'success')
    except Exception as e:
        logger.error(f"Error clearing SSL cache for {domain}: {str(e)}", exc_info=True)
        flash(f'Error clearing SSL cache: {str(e)}', 'error')

    # Redirect back to the referring page or to the SSL certificates page
    referrer = request.referrer
    if referrer:
        from urllib.parse import urlparse
        referrer = referrer.replace('\\', '/')
        parsed_url = urlparse(referrer)
        # Define a list of trusted domains
        trusted_domains = ['yourtrustedomain.com']
        # Allow only relative URLs or URLs from trusted domains
        if not parsed_url.netloc or parsed_url.netloc in trusted_domains:
            return redirect(referrer)
    return redirect(url_for('ssl_certificates'))

@app.route('/clear_all_ssl_cache')
@auth.login_required
def clear_all_ssl_cache():
    """Clear SSL cache for all domains"""
    try:
        # Clear all SSL cache entries
        db.clear_cache_by_prefix("ssl_")
        logger.info("Cleared SSL cache for all domains")
        flash('Successfully cleared SSL cache for all domains. Please refresh the domains to check their status.', 'success')
    except Exception as e:
        logger.error(f"Error clearing SSL cache: {str(e)}", exc_info=True)
        flash(f'Failed to clear SSL cache: {str(e)}', 'error')

    return redirect(url_for('ssl_certificates'))

@app.route('/test_whois_api')
@auth.login_required
def test_whois_api():
    """Test the WHOIS API configuration"""
    api_key = get_whois_api_key()
    if not api_key:
        flash('No WHOIS API key configured. Please add a WHOIS API key in settings.', 'error')
        return redirect(url_for('app_settings'))

    # Test the API with a known domain
    test_domain = "google.com"
    try:
        url = f"https://www.whoisxmlapi.com/whoisserver/WhoisService?apiKey={api_key}&domainName={test_domain}&outputFormat=json"
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }

        response = requests.get(url, headers=headers, timeout=10)
        response.raise_for_status()

        data = response.json()
        if 'WhoisRecord' in data:
            flash(f'WHOIS API test successful! API key is working correctly.', 'success')
        else:
            flash(f'WHOIS API test failed. Received unexpected response: {data}', 'error')
    except Exception as e:
        flash(f'WHOIS API test failed: {str(e)}', 'error')

    return redirect(url_for('app_settings'))

# Reports functionality
@app.route('/reports')
@auth.login_required
def reports():
    """Reports page"""
    # Get current user
    user = auth.get_current_user()

    # Get all organizations the user has access to
    organizations = []
    if user.get('is_admin'):
        organizations = db.get_all_organizations()
    else:
        organizations = db.get_user_organizations(user.get('id'))

    # Get all domains the user has access to
    domains = []
    current_org = user.get('current_organization')
    if current_org:
        domains = db.get_domains_by_organization(current_org.get('id'))

    return render_template('reports.html',
                          organizations=organizations,
                          domains=domains)

@app.route('/generate_report', methods=['POST'])
@auth.login_required
def generate_report():
    """Generate a report based on form data"""
    # Get current user
    user = auth.get_current_user()

    # Get form data
    report_type = request.form.get('report_type', 'all_domains')
    time_range = int(request.form.get('time_range', 30))
    organization_id = request.form.get('organization_id', 'all')
    domain_ids = request.form.getlist('domain_ids')

    # Get status filters
    ssl_statuses = request.form.getlist('ssl_status[]') or ['valid', 'warning', 'expired', 'error']
    domain_statuses = request.form.getlist('domain_status[]') or ['valid', 'warning', 'expired', 'error']
    ping_statuses = request.form.getlist('ping_status[]') or ['up', 'down', 'unknown']

    # Get include options
    include_charts = request.form.get('include_charts') != 'off'  # Default to true if not specified
    include_tables = request.form.get('include_tables') != 'off'  # Default to true if not specified
    include_alerts = request.form.get('include_alerts') != 'off'  # Default to true if not specified

    # Get domains based on selection
    domains = []
    if organization_id == 'all':
        # Get all domains the user has access to
        if user.get('is_admin'):
            domains = db.get_all_domains()
        else:
            # Get domains from all organizations the user belongs to
            user_orgs = db.get_user_organizations(user.get('id'))
            for org in user_orgs:
                domains.extend(db.get_domains_by_organization(org.get('id')))
    else:
        # Get domains for the selected organization
        domains = db.get_domains_by_organization(organization_id)

    # Filter domains if specific ones were selected
    if domain_ids and 'all' not in domain_ids:
        domains = [d for d in domains if str(d.id) in domain_ids]

    # Generate report data based on report type
    report_data = {}

    if report_type == 'ssl_status':
        report_data = generate_ssl_report(domains, time_range, ssl_statuses, include_charts, include_tables, include_alerts)
    elif report_type == 'domain_expiry':
        report_data = generate_domain_expiry_report(domains, time_range, domain_statuses, include_charts, include_tables, include_alerts)
    elif report_type == 'ping_uptime':
        report_data = generate_ping_report(domains, time_range, ping_statuses, include_charts, include_tables, include_alerts)
    else:  # all_domains
        report_data = generate_all_domains_report(domains, time_range, include_charts, include_tables, include_alerts)

    # Return JSON response
    return jsonify(report_data)

@app.route('/export_report', methods=['POST'])
@auth.login_required
def export_report():
    """Export a report in the specified format"""
    # Get current user
    user = auth.get_current_user()

    # Get form data
    report_type = request.form.get('report_type', 'all_domains')
    time_range = int(request.form.get('time_range', 30))
    organization_id = request.form.get('organization_id', 'all')
    domain_ids = request.form.getlist('domain_ids')
    export_format = request.form.get('export_format', 'csv')

    # Get status filters
    ssl_statuses = request.form.getlist('ssl_status[]') or ['valid', 'warning', 'expired', 'error']
    domain_statuses = request.form.getlist('domain_status[]') or ['valid', 'warning', 'expired', 'error']
    ping_statuses = request.form.getlist('ping_status[]') or ['up', 'down', 'unknown']

    # Get include options
    include_charts = request.form.get('include_charts') != 'off'  # Default to true if not specified
    include_tables = request.form.get('include_tables') != 'off'  # Default to true if not specified
    include_alerts = request.form.get('include_alerts') != 'off'  # Default to true if not specified

    # Get domains based on selection (same logic as generate_report)
    domains = []
    if organization_id == 'all':
        if user.get('is_admin'):
            domains = db.get_all_domains()
        else:
            user_orgs = db.get_user_organizations(user.get('id'))
            for org in user_orgs:
                domains.extend(db.get_domains_by_organization(org.get('id')))
    else:
        domains = db.get_domains_by_organization(organization_id)

    # Filter domains if specific ones were selected
    if domain_ids and 'all' not in domain_ids:
        domains = [d for d in domains if str(d.id) in domain_ids]

    # Generate report data
    if report_type == 'ssl_status':
        report_data = generate_ssl_report(domains, time_range, ssl_statuses, include_charts, include_tables, include_alerts)
    elif report_type == 'domain_expiry':
        report_data = generate_domain_expiry_report(domains, time_range, domain_statuses, include_charts, include_tables, include_alerts)
    elif report_type == 'ping_uptime':
        report_data = generate_ping_report(domains, time_range, ping_statuses, include_charts, include_tables, include_alerts)
    else:  # all_domains
        report_data = generate_all_domains_report(domains, time_range, include_charts, include_tables, include_alerts)

    # Export the report in the specified format
    if export_format == 'csv':
        return export_report_csv(report_data)
    elif export_format == 'pdf':
        return export_report_pdf(report_data)
    elif export_format == 'excel':
        return export_report_excel(report_data)
    else:
        flash('Invalid export format', 'error')
        return redirect(url_for('reports'))

# Helper functions for report generation
def generate_ssl_report(domains, time_range, status_filters=None, include_charts=True, include_tables=True, include_alerts=True):
    """Generate SSL certificate status report"""
    # Set default status filters if not provided
    if status_filters is None:
        status_filters = ['valid', 'warning', 'expired', 'error']

    # Get SSL certificate status for each domain
    ssl_statuses = []
    for domain in domains:
        try:
            # Handle both object and dictionary access
            domain_name = domain.get('name') if isinstance(domain, dict) else domain.name
            ssl_status = check_certificate(domain_name)
            # Only include domains with statuses in the filter
            if ssl_status.status in status_filters:
                ssl_statuses.append(ssl_status)
        except Exception as e:
            domain_name = domain.get('name') if isinstance(domain, dict) else getattr(domain, 'name', 'unknown')
            logger.error(f"Error checking SSL for {domain_name}: {str(e)}")

    # Count domains by status
    valid_count = sum(1 for s in ssl_statuses if s.status == 'valid')
    warning_count = sum(1 for s in ssl_statuses if s.status == 'warning')
    expired_count = sum(1 for s in ssl_statuses if s.status == 'expired')
    error_count = sum(1 for s in ssl_statuses if s.status == 'error')

    # Create table data if tables are included
    table_data = []
    if include_tables:
        for status in ssl_statuses:
            table_data.append({
                'Domain': status.domain,
                'Status': status.status.capitalize(),
                'Expiry Date': status.expiry_date.strftime('%Y-%m-%d') if status.expiry_date else 'Unknown',
                'Days Remaining': str(status.days_remaining) if status.days_remaining >= 0 else 'Expired' if status.days_remaining < 0 else 'Unknown'
            })

    # Get alerts related to SSL certificates if alerts are included
    alerts_data = []
    if include_alerts:
        alerts = db.get_alerts_by_type('ssl_certificate', time_range)
        for alert in alerts:
            alerts_data.append({
                'Date': alert['created_at'].strftime('%Y-%m-%d %H:%M'),
                'Domain': alert['domain'],
                'Alert Type': 'SSL Certificate',
                'Message': alert['message'],
                'Status': 'Acknowledged' if alert['acknowledged'] else 'Active'
            })

    # Create chart data if charts are included
    status_chart = {}
    trend_chart = {}
    if include_charts:
        status_chart = {
            'labels': ['Valid', 'Warning', 'Expired', 'Error'],
            'values': [valid_count, warning_count, expired_count, error_count]
        }

        # Create trend data (mock data for now)
        # In a real implementation, this would use historical data from the database
        trend_chart = {
            'labels': [f'Day {i}' for i in range(1, time_range + 1)],
            'datasets': [
                {
                    'label': 'Valid',
                    'data': [valid_count] * time_range,
                    'borderColor': '#28a745',
                    'backgroundColor': 'rgba(40, 167, 69, 0.1)'
                },
                {
                    'label': 'Warning',
                    'data': [warning_count] * time_range,
                    'borderColor': '#ffc107',
                    'backgroundColor': 'rgba(255, 193, 7, 0.1)'
                },
                {
                    'label': 'Expired',
                    'data': [expired_count] * time_range,
                    'borderColor': '#dc3545',
                    'backgroundColor': 'rgba(220, 53, 69, 0.1)'
                }
            ]
        }

    # Return the complete report data
    return {
        'title': 'SSL Certificate Status Report',
        'summary': {
            'total': len(ssl_statuses),
            'healthy': valid_count,
            'warning': warning_count,
            'critical': expired_count + error_count
        },
        'charts': {
            'status': status_chart,
            'trend': trend_chart
        },
        'table': table_data,
        'alerts': alerts_data
    }

def generate_domain_expiry_report(domains, time_range, status_filters=None, include_charts=True, include_tables=True, include_alerts=True):
    """Generate domain expiry status report"""
    # Set default status filters if not provided
    if status_filters is None:
        status_filters = ['valid', 'warning', 'expired', 'error']

    # Get domain expiry status for each domain
    domain_statuses = []
    for domain in domains:
        try:
            # Handle both object and dictionary access
            domain_name = domain.get('name') if isinstance(domain, dict) else domain.name
            domain_status = check_domain_expiry(domain_name)
            # Only include domains with statuses in the filter
            if domain_status.status in status_filters:
                domain_statuses.append(domain_status)
        except Exception as e:
            domain_name = domain.get('name') if isinstance(domain, dict) else getattr(domain, 'name', 'unknown')
            logger.error(f"Error checking domain expiry for {domain_name}: {str(e)}")

    # Count domains by status
    valid_count = sum(1 for s in domain_statuses if s.status == 'valid')
    warning_count = sum(1 for s in domain_statuses if s.status == 'warning')
    expired_count = sum(1 for s in domain_statuses if s.status == 'expired')
    error_count = sum(1 for s in domain_statuses if s.status == 'error')

    # Create table data if tables are included
    table_data = []
    if include_tables:
        for status in domain_statuses:
            table_data.append({
                'Domain': status.name,
                'Status': status.status.capitalize(),
                'Expiry Date': status.expiry_date.strftime('%Y-%m-%d') if status.expiry_date else 'Unknown',
                'Days Remaining': str(status.days_remaining) if status.days_remaining >= 0 else 'Expired' if status.days_remaining < 0 else 'Unknown',
                'Registrar': status.registrar
            })

    # Get alerts related to domain expiry if alerts are included
    alerts_data = []
    if include_alerts:
        alerts = db.get_alerts_by_type('domain_expiry', time_range)
        for alert in alerts:
            alerts_data.append({
                'Date': alert['created_at'].strftime('%Y-%m-%d %H:%M'),
                'Domain': alert['domain'],
                'Alert Type': 'Domain Expiry',
                'Message': alert['message'],
                'Status': 'Acknowledged' if alert['acknowledged'] else 'Active'
            })

    # Create chart data if charts are included
    status_chart = {}
    trend_chart = {}
    if include_charts:
        status_chart = {
            'labels': ['Valid', 'Warning', 'Expired', 'Error'],
            'values': [valid_count, warning_count, expired_count, error_count]
        }

        # Create trend data (mock data for now)
        trend_chart = {
            'labels': [f'Day {i}' for i in range(1, time_range + 1)],
            'datasets': [
                {
                    'label': 'Valid',
                    'data': [valid_count] * time_range,
                    'borderColor': '#28a745',
                    'backgroundColor': 'rgba(40, 167, 69, 0.1)'
                },
                {
                    'label': 'Warning',
                    'data': [warning_count] * time_range,
                    'borderColor': '#ffc107',
                    'backgroundColor': 'rgba(255, 193, 7, 0.1)'
                },
                {
                    'label': 'Expired',
                    'data': [expired_count] * time_range,
                    'borderColor': '#dc3545',
                    'backgroundColor': 'rgba(220, 53, 69, 0.1)'
                }
            ]
        }

    # Return the complete report data
    return {
        'title': 'Domain Expiry Status Report',
        'summary': {
            'total': len(domain_statuses),
            'healthy': valid_count,
            'warning': warning_count,
            'critical': expired_count + error_count
        },
        'charts': {
            'status': status_chart,
            'trend': trend_chart
        },
        'table': table_data,
        'alerts': alerts_data
    }

def generate_ping_report(domains, time_range, status_filters=None, include_charts=True, include_tables=True, include_alerts=True):
    """Generate ping uptime report"""
    # Set default status filters if not provided
    if status_filters is None:
        status_filters = ['up', 'down', 'unknown']

    # Get ping status for each domain
    ping_statuses = []
    for domain in domains:
        try:
            # Handle both object and dictionary access
            domain_name = domain.get('name') if isinstance(domain, dict) else domain.name
            ping_result = check_ping(domain_name)
            # Only include domains with statuses in the filter
            if ping_result['status'] in status_filters:
                ping_statuses.append({
                    'domain': domain_name,
                    'status': ping_result['status'],
                    'response_time': ping_result.get('response_time', 0)
                })
        except Exception as e:
            domain_name = domain.get('name') if isinstance(domain, dict) else getattr(domain, 'name', 'unknown')
            logger.error(f"Error checking ping for {domain_name}: {str(e)}")
            if 'unknown' in status_filters:
                ping_statuses.append({
                    'domain': domain_name,
                    'status': 'unknown',
                    'response_time': 0
                })

    # Count domains by status
    up_count = sum(1 for s in ping_statuses if s['status'] == 'up')
    down_count = sum(1 for s in ping_statuses if s['status'] == 'down')
    unknown_count = sum(1 for s in ping_statuses if s['status'] == 'unknown')

    # Create table data if tables are included
    table_data = []
    if include_tables:
        for status in ping_statuses:
            table_data.append({
                'Domain': status['domain'],
                'Status': status['status'].capitalize(),
                'Response Time': f"{status['response_time']} ms" if status['response_time'] > 0 else 'N/A',
                'Last Checked': datetime.now().strftime('%Y-%m-%d %H:%M')
            })

    # Get alerts related to ping monitoring if alerts are included
    alerts_data = []
    if include_alerts:
        alerts = db.get_alerts_by_type('ping', time_range)
        for alert in alerts:
            alerts_data.append({
                'Date': alert['created_at'].strftime('%Y-%m-%d %H:%M'),
                'Domain': alert['domain'],
                'Alert Type': 'Ping Monitoring',
                'Message': alert['message'],
                'Status': 'Acknowledged' if alert['acknowledged'] else 'Active'
            })

    # Create chart data if charts are included
    status_chart = {}
    trend_chart = {}
    if include_charts:
        status_chart = {
            'labels': ['Up', 'Down', 'Unknown'],
            'values': [up_count, down_count, unknown_count]
        }

        # Create trend data (mock data for now)
        trend_chart = {
            'labels': [f'Day {i}' for i in range(1, time_range + 1)],
            'datasets': [
                {
                    'label': 'Uptime %',
                    'data': [round((up_count / len(ping_statuses)) * 100) if len(ping_statuses) > 0 else 0] * time_range,
                    'borderColor': '#28a745',
                    'backgroundColor': 'rgba(40, 167, 69, 0.1)'
                }
            ]
        }

    # Return the complete report data
    return {
        'title': 'Ping Uptime Report',
        'summary': {
            'total': len(ping_statuses),
            'healthy': up_count,
            'warning': unknown_count,
            'critical': down_count
        },
        'charts': {
            'status': status_chart,
            'trend': trend_chart
        },
        'table': table_data,
        'alerts': alerts_data
    }

def generate_all_domains_report(domains, time_range, include_charts=True, include_tables=True, include_alerts=True):
    """Generate comprehensive report for all domains"""
    # Get status for each domain
    domain_data = []
    for domain in domains:
        try:
            # Handle both object and dictionary access
            domain_name = domain.get('name') if isinstance(domain, dict) else domain.name
            ssl_status = check_certificate(domain_name)
            domain_expiry = check_domain_expiry(domain_name)
            ping_result = check_ping(domain_name)

            # Determine overall status
            if ssl_status.status == 'expired' or domain_expiry.status == 'expired' or ping_result['status'] == 'down':
                overall_status = 'critical'
            elif ssl_status.status == 'warning' or domain_expiry.status == 'warning':
                overall_status = 'warning'
            elif ssl_status.status == 'error' or domain_expiry.status == 'error' or ping_result['status'] == 'unknown':
                overall_status = 'warning'
            else:
                overall_status = 'healthy'

            domain_data.append({
                'domain': domain_name,
                'ssl_status': ssl_status,
                'domain_expiry': domain_expiry,
                'ping_status': ping_result,
                'overall_status': overall_status
            })
        except Exception as e:
            domain_name = domain.get('name') if isinstance(domain, dict) else getattr(domain, 'name', 'unknown')
            logger.error(f"Error checking domain {domain_name}: {str(e)}")

    # Count domains by overall status
    healthy_count = sum(1 for d in domain_data if d['overall_status'] == 'healthy')
    warning_count = sum(1 for d in domain_data if d['overall_status'] == 'warning')
    critical_count = sum(1 for d in domain_data if d['overall_status'] == 'critical')

    # Create table data if tables are included
    table_data = []
    if include_tables:
        for data in domain_data:
            ssl_days = data['ssl_status'].days_remaining if data['ssl_status'].days_remaining >= 0 else 'Expired'
            domain_days = data['domain_expiry'].days_remaining if data['domain_expiry'].days_remaining >= 0 else 'Expired'

            table_data.append({
                'Domain': data['domain'],
                'Status': data['overall_status'].capitalize(),
                'SSL Expiry': f"{ssl_days} days" if isinstance(ssl_days, int) else ssl_days,
                'Domain Expiry': f"{domain_days} days" if isinstance(domain_days, int) else domain_days,
                'Ping Status': data['ping_status']['status'].capitalize()
            })

    # Get all alerts if alerts are included
    alerts_data = []
    if include_alerts:
        alerts = db.get_alerts(time_range)
        for alert in alerts:
            alerts_data.append({
                'Date': alert['created_at'].strftime('%Y-%m-%d %H:%M'),
                'Domain': alert['domain'],
                'Alert Type': alert['alert_type'].replace('_', ' ').title(),
                'Message': alert['message'],
                'Status': 'Acknowledged' if alert['acknowledged'] else 'Active'
            })

    # Create chart data if charts are included
    status_chart = {}
    trend_chart = {}
    if include_charts:
        status_chart = {
            'labels': ['Healthy', 'Warning', 'Critical'],
            'values': [healthy_count, warning_count, critical_count]
        }

        # Create trend data (mock data for now)
        trend_chart = {
            'labels': [f'Day {i}' for i in range(1, time_range + 1)],
            'datasets': [
                {
                    'label': 'Healthy',
                    'data': [healthy_count] * time_range,
                    'borderColor': '#28a745',
                    'backgroundColor': 'rgba(40, 167, 69, 0.1)'
                },
                {
                    'label': 'Warning',
                    'data': [warning_count] * time_range,
                    'borderColor': '#ffc107',
                    'backgroundColor': 'rgba(255, 193, 7, 0.1)'
                },
                {
                    'label': 'Critical',
                    'data': [critical_count] * time_range,
                    'borderColor': '#dc3545',
                    'backgroundColor': 'rgba(220, 53, 69, 0.1)'
                }
            ]
        }

    # Return the complete report data
    return {
        'title': 'All Domains Status Report',
        'summary': {
            'total': len(domain_data),
            'healthy': healthy_count,
            'warning': warning_count,
            'critical': critical_count
        },
        'charts': {
            'status': status_chart,
            'trend': trend_chart
        },
        'table': table_data,
        'alerts': alerts_data
    }

def export_report_csv(report_data):
    """Export report as CSV file"""
    # Create CSV content
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow(['Report: ' + report_data['title']])
    writer.writerow(['Generated: ' + datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow([])

    # Write summary
    writer.writerow(['Summary'])
    writer.writerow(['Total Domains', report_data['summary']['total']])
    writer.writerow(['Healthy', report_data['summary']['healthy']])
    writer.writerow(['Warning', report_data['summary']['warning']])
    writer.writerow(['Critical', report_data['summary']['critical']])
    writer.writerow([])

    # Write table data
    if report_data['table']:
        writer.writerow(report_data['table'][0].keys())
        for row in report_data['table']:
            writer.writerow(row.values())

    # Create response
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-disposition": f"attachment; filename={report_data['title'].replace(' ', '_')}.csv"}
    )

def export_report_excel(report_data):
    """Export report as Excel file"""
    # Use openpyxl to create an Excel file
    import io
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

    # Add title
    ws['A1'] = 'Report: ' + report_data['title']
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')

    # Add generation timestamp
    ws['A2'] = 'Generated: ' + datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws.merge_cells('A2:B2')

    # Add summary header
    ws['A4'] = 'Summary'
    ws['A4'].font = Font(bold=True)

    # Add summary data
    ws['A5'] = 'Total Domains'
    ws['B5'] = report_data['summary']['total']

    ws['A6'] = 'Healthy'
    ws['B6'] = report_data['summary']['healthy']

    ws['A7'] = 'Warning'
    ws['B7'] = report_data['summary']['warning']

    ws['A8'] = 'Critical'
    ws['B8'] = report_data['summary']['critical']

    # Add table data if available
    if report_data['table']:
        # Add a header row for the table
        row_num = 10
        ws['A' + str(row_num)] = 'Detailed Data'
        ws['A' + str(row_num)].font = Font(bold=True)
        ws.merge_cells(f'A{row_num}:B{row_num}')
        row_num += 1

        # Get the keys from the first row as headers
        headers = list(report_data['table'][0].keys())

        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Write data rows
        for row_data in report_data['table']:
            row_num += 1
            for col_num, key in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num).value = row_data.get(key, '')

    # Save to a BytesIO object
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    filename = f"{report_data['title'].replace(' ', '_')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-disposition": f"attachment; filename={filename}"}
    )

def export_report_pdf(report_data):
    """Export report as PDF file"""
    # Use reportlab to create a PDF file
    import io
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    # Create a buffer for the PDF
    buffer = io.BytesIO()

    # Create the PDF document
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()

    # Create custom styles
    title_style = ParagraphStyle(
        'Title',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=12
    )

    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=10
    )

    normal_style = styles['Normal']

    # Create the content elements
    elements = []

    # Add title
    elements.append(Paragraph(f"Report: {report_data['title']}", title_style))
    elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    elements.append(Spacer(1, 12))

    # Add summary
    elements.append(Paragraph("Summary", subtitle_style))

    summary_data = [
        ["Total Domains", str(report_data['summary']['total'])],
        ["Healthy", str(report_data['summary']['healthy'])],
        ["Warning", str(report_data['summary']['warning'])],
        ["Critical", str(report_data['summary']['critical'])]
    ]

    summary_table = Table(summary_data, colWidths=[200, 100])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 12))

    # Add table data if available
    if report_data['table']:
        elements.append(Paragraph("Detailed Data", subtitle_style))

        # Get headers from the first row
        headers = list(report_data['table'][0].keys())

        # Prepare table data
        table_data = [headers]  # First row is headers

        # Add data rows
        for row in report_data['table']:
            table_data.append([str(row.get(key, '')) for key in headers])

        # Create the table
        col_widths = [100] * len(headers)  # Adjust column widths as needed
        detail_table = Table(table_data, colWidths=col_widths)

        # Style the table
        table_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]
        detail_table.setStyle(TableStyle(table_style))

        elements.append(detail_table)

    # Build the PDF
    doc.build(elements)

    # Get the value from the buffer
    buffer.seek(0)

    # Create response
    filename = f"{report_data['title'].replace(' ', '_')}.pdf"
    return Response(
        buffer.getvalue(),
        mimetype="application/pdf",
        headers={"Content-disposition": f"attachment; filename={filename}"}
    )

@app.route('/restore_config', methods=['POST'])
def restore_config():
    """Restore configuration from a backup file"""
    if 'backup_file' not in request.files:
        flash('No backup file provided', 'error')
        return redirect(url_for('app_settings'))

    backup_file = request.files['backup_file']
    if backup_file.filename == '':
        flash('No backup file selected', 'error')
        return redirect(url_for('app_settings'))

    # Validate YAML format
    try:
        backup_content = backup_file.read()
        backup_file.seek(0)  # Reset file pointer
        yaml_content = yaml.safe_load(backup_content)

        if not isinstance(yaml_content, dict):
            flash('Invalid backup file format', 'error')
            return redirect(url_for('app_settings'))

        # Create a backup of the current config before restoring
        if os.path.exists(CONFIG_FILE):
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            pre_restore_backup = os.path.join(DATA_DIR, f"config_pre_restore_{timestamp}.yaml")
            shutil.copy2(CONFIG_FILE, pre_restore_backup)

        # Save the uploaded file as the new config
        backup_file.save(CONFIG_FILE)
        flash('Configuration restored successfully', 'success')
    except Exception as e:
        flash(f'Error restoring backup: {str(e)}', 'error')

    return redirect(url_for('app_settings'))

@app.route('/add_ssl_from_dashboard', methods=['POST'])
def add_ssl_from_dashboard():
    """Add a new domain to SSL certificate monitoring from the dashboard"""
    if request.method == 'POST':
        domain = request.form.get('domain', '').strip()

        if not domain:
            flash('Please enter a domain name', 'error')
            return redirect(url_for('index'))

        # Clean the domain
        domain = clean_domain_name(domain)
        if not domain:
            flash('Invalid domain format', 'error')
            return redirect(url_for('index'))

        # Check if domain already exists
        config = load_config()
        if 'ssl_domains' not in config:
            config['ssl_domains'] = []

        for entry in config.get('ssl_domains', []):
            if entry.get('url') == domain:
                flash(f'Domain {domain} is already being monitored for SSL', 'error')
                return redirect(url_for('index'))

        # Add domain to config
        config['ssl_domains'].append({'url': domain})
        save_config(config)

        flash(f'Domain {domain} added to SSL monitoring successfully', 'success')
        return redirect(url_for('index'))

@app.route('/add_domain_from_dashboard', methods=['POST'])
def add_domain_from_dashboard():
    """Add a new domain to monitoring from the dashboard"""
    if request.method == 'POST':
        domain = request.form.get('domain', '').strip()
        monitor_ssl = 'monitor_ssl' in request.form
        monitor_expiry = 'monitor_expiry' in request.form
        monitor_ping = 'monitor_ping' in request.form

        if not domain:
            flash('Please enter a domain name', 'error')
            return redirect(url_for('index'))

        # Clean the domain
        domain = clean_domain_name(domain)
        if not domain:
            flash('Invalid domain format', 'error')
            return redirect(url_for('index'))

        # Check if at least one monitoring option is selected
        if not any([monitor_ssl, monitor_expiry, monitor_ping]):
            flash('Please select at least one monitoring option', 'error')
            return redirect(url_for('index'))

        # Get current user
        current_user = auth.get_current_user()

        # Get current organization
        current_org = current_user.get('current_organization')
        if not current_org:
            flash("You don't have access to any organizations", 'error')
            return redirect(url_for('profile'))

        # Load the current configuration
        config = load_config()

        # Initialize config sections if they don't exist
        if 'ssl_domains' not in config:
            config['ssl_domains'] = []
        if 'domain_expiry' not in config:
            config['domain_expiry'] = []
        if 'ping_hosts' not in config:
            config['ping_hosts'] = []

        # Check if domain already exists in this organization
        existing_domain = db.get_domain_by_name_and_org(domain, current_org['id'])
        if existing_domain:
            # Update existing domain with new monitoring options
            ssl_monitored = monitor_ssl or existing_domain['ssl_monitored']
            expiry_monitored = monitor_expiry or existing_domain['expiry_monitored']
            ping_monitored = monitor_ping or existing_domain['ping_monitored']

            updated = db.update_domain(
                existing_domain['id'],
                domain,
                ssl_monitored,
                expiry_monitored,
                ping_monitored
            )
            if updated:
                # Update the configuration file

                # Update SSL monitoring
                if ssl_monitored:
                    # Check if domain is already in ssl_domains
                    ssl_exists = False
                    for entry in config['ssl_domains']:
                        if entry.get('url') == domain:
                            ssl_exists = True
                            break

                    # Add to ssl_domains if not already there
                    if not ssl_exists:
                        config['ssl_domains'].append({'url': domain})
                else:
                    # Remove from SSL domains if monitoring is disabled
                    config['ssl_domains'] = [entry for entry in config['ssl_domains']
                                           if entry.get('url') != domain]

                # Update domain expiry monitoring
                if expiry_monitored:
                    # Check if domain is already in domain_expiry
                    expiry_exists = False
                    for entry in config['domain_expiry']:
                        if entry.get('name') == domain:
                            expiry_exists = True
                            break

                    # Add to domain_expiry if not already there
                    if not expiry_exists:
                        config['domain_expiry'].append({'name': domain})
                else:
                    # Remove from domain expiry if monitoring is disabled
                    config['domain_expiry'] = [entry for entry in config['domain_expiry']
                                             if entry.get('name') != domain]

                # Update ping monitoring
                if ping_monitored:
                    # Check if domain is already in ping_hosts
                    ping_exists = False
                    for entry in config['ping_hosts']:
                        if entry.get('host') == domain:
                            ping_exists = True
                            break

                    # Add to ping_hosts if not already there
                    if not ping_exists:
                        config['ping_hosts'].append({'host': domain})
                else:
                    # Remove from ping hosts if monitoring is disabled
                    config['ping_hosts'] = [entry for entry in config['ping_hosts']
                                          if entry.get('host') != domain]

                # Save the updated configuration
                save_config(config)

                flash(f'Domain {domain} monitoring options updated successfully', 'success')
            else:
                flash(f'Error updating domain {domain}', 'error')
        else:
            # Add new domain with selected monitoring options and auto-logging
            domain_id = db.add_domain(
                name=domain,
                organization_id=current_org['id'],
                ssl_monitored=monitor_ssl,
                expiry_monitored=monitor_expiry,
                ping_monitored=monitor_ping,
                user_id=current_user['id'],
                username=current_user['username'],
                ip_address=request.remote_addr,
                auto_log=True
            )
            if domain_id:
                # Update the configuration file
                config_updated = False

                # Add to SSL monitoring
                if monitor_ssl:
                    # Check if domain is already in ssl_domains
                    ssl_exists = False
                    for entry in config['ssl_domains']:
                        if entry.get('url') == domain:
                            ssl_exists = True
                            break

                    # Add to ssl_domains if not already there
                    if not ssl_exists:
                        config['ssl_domains'].append({'url': domain})
                        config_updated = True

                        # Pre-cache an error status to prevent initial slow loading
                        # This will be updated in the background
                        error_status = CertificateStatus(
                            domain=domain,
                            days_remaining=-1,
                            expiry_date=datetime.now(),
                            status='checking',  # Use 'checking' status to indicate it's being processed
                            ping_status='unknown'
                        )
                        cache_ssl_data(domain, error_status)

                        # Start a background thread to check the certificate
                        threading.Thread(
                            target=lambda d: check_certificate(d),
                            args=(domain,),
                            daemon=True
                        ).start()

                # Add to domain expiry monitoring
                if monitor_expiry:
                    # Check if domain is already in domain_expiry
                    expiry_exists = False
                    for entry in config['domain_expiry']:
                        if entry.get('name') == domain:
                            expiry_exists = True
                            break

                    # Add to domain_expiry if not already there
                    if not expiry_exists:
                        config['domain_expiry'].append({'name': domain})
                        config_updated = True

                        # Force an immediate check to populate the domain expiry cache
                        try:
                            # Clear any existing cache entry
                            db.clear_cache(f"domain_expiry_{domain}")

                            # Perform a check to populate the cache
                            check_domain_expiry(domain)
                        except Exception as e:
                            logger.error(f"Error checking domain expiry for {domain}: {str(e)}")

                # Add to ping monitoring
                if monitor_ping:
                    # Check if domain is already in ping_hosts
                    ping_exists = False
                    for entry in config['ping_hosts']:
                        if entry.get('host') == domain:
                            ping_exists = True
                            break

                    # Add to ping_hosts if not already there
                    if not ping_exists:
                        config['ping_hosts'].append({'host': domain})
                        config_updated = True

                        # Force an immediate ping check
                        try:
                            # Clear any existing cache entry
                            db.clear_cache(f"ping_{domain}")

                            # Perform a check to populate the cache
                            check_ping(domain)
                        except Exception as e:
                            logger.error(f"Error checking ping for {domain}: {str(e)}")

                # Save the updated configuration if changes were made
                if config_updated:
                    save_config(config)

                # Track what was added
                added_to = []
                if monitor_ssl:
                    added_to.append('SSL certificates')
                if monitor_expiry:
                    added_to.append('domain expiry')
                if monitor_ping:
                    added_to.append('ping monitoring')

                flash(f'Domain {domain} added to {", ".join(added_to)} successfully!', 'success')
            else:
                flash(f'Error adding domain {domain}', 'error')

        return redirect(url_for('index'))

@app.route('/bulk_add_domains_from_dashboard', methods=['POST'])
def bulk_add_domains_from_dashboard():
    """Add multiple domains to monitoring from the dashboard"""
    if request.method == 'POST':
        domains_text = request.form.get('domains', '').strip()
        monitor_ssl = 'monitor_ssl' in request.form
        monitor_expiry = 'monitor_expiry' in request.form
        monitor_ping = 'monitor_ping' in request.form

        if not domains_text:
            flash('Please enter at least one domain', 'error')
            return redirect(url_for('index'))

        # Check if at least one monitoring option is selected
        if not any([monitor_ssl, monitor_expiry, monitor_ping]):
            flash('Please select at least one monitoring option', 'error')
            return redirect(url_for('index'))

        # Split by newlines, commas, or spaces
        domains = re.split(r'[\n,\s]+', domains_text)
        domains = [d.strip() for d in domains if d.strip()]

        if not domains:
            flash('No valid domains found', 'error')
            return redirect(url_for('index'))

        # Get current user
        current_user = auth.get_current_user()

        # Get current organization
        current_org = current_user.get('current_organization')
        if not current_org:
            flash("You don't have access to any organizations", 'error')
            return redirect(url_for('profile'))

        # Load the current configuration
        config = load_config()

        # Initialize config sections if they don't exist
        if 'ssl_domains' not in config:
            config['ssl_domains'] = []
        if 'domain_expiry' not in config:
            config['domain_expiry'] = []
        if 'ping_hosts' not in config:
            config['ping_hosts'] = []

        added_count = 0
        updated_count = 0
        skipped_count = 0

        # Track domains added to config
        ssl_domains_added = []
        expiry_domains_added = []
        ping_hosts_added = []

        for domain in domains:
            # Clean the domain
            domain = clean_domain_name(domain)
            if not domain:
                skipped_count += 1
                continue

            # Check if domain already exists in this organization
            existing_domain = db.get_domain_by_name_and_org(domain, current_org['id'])
            if existing_domain:
                # Update existing domain with new monitoring options
                updated = db.update_domain(
                    existing_domain['id'],
                    domain,
                    monitor_ssl or existing_domain['ssl_monitored'],
                    monitor_expiry or existing_domain['expiry_monitored'],
                    monitor_ping or existing_domain['ping_monitored']
                )
                if updated:
                    # Log the domain update action
                    monitoring_types_str = []
                    if monitor_ssl and not existing_domain['ssl_monitored']:
                        monitoring_types_str.append("SSL")
                    if monitor_expiry and not existing_domain['expiry_monitored']:
                        monitoring_types_str.append("expiry")
                    if monitor_ping and not existing_domain['ping_monitored']:
                        monitoring_types_str.append("ping")

                    if monitoring_types_str:  # Only log if something was actually updated
                        db.log_user_action(
                            user_id=current_user['id'],
                            username=current_user['username'],
                            action_type='update',
                            resource_type='domain',
                            resource_id=existing_domain['id'],
                            resource_name=domain,
                            details=f'Updated domain to add {", ".join(monitoring_types_str)} monitoring (bulk dashboard)',
                            ip_address=request.remote_addr,
                            organization_id=current_org['id']
                        )
                    updated_count += 1

                    # Update SSL monitoring in config
                    if monitor_ssl and not existing_domain['ssl_monitored']:
                        ssl_domains_added.append(domain)

                    # Update domain expiry monitoring in config
                    if monitor_expiry and not existing_domain['expiry_monitored']:
                        expiry_domains_added.append(domain)

                    # Update ping monitoring in config
                    if monitor_ping and not existing_domain['ping_monitored']:
                        ping_hosts_added.append(domain)
                else:
                    skipped_count += 1
            else:
                # Add new domain with selected monitoring options and auto-logging
                domain_id = db.add_domain(
                    name=domain,
                    organization_id=current_org['id'],
                    ssl_monitored=monitor_ssl,
                    expiry_monitored=monitor_expiry,
                    ping_monitored=monitor_ping,
                    user_id=current_user['id'],
                    username=current_user['username'],
                    ip_address=request.remote_addr,
                    auto_log=True
                )
                if domain_id:
                    # The log entry is now created automatically by add_domain
                    # Add to monitoring types for the flash message
                    monitoring_types_str = []
                    if monitor_ssl:
                        monitoring_types_str.append("SSL")
                    if monitor_expiry:
                        monitoring_types_str.append("expiry")
                    if monitor_ping:
                        monitoring_types_str.append("ping")
                    added_count += 1

                    # Add to config lists
                    if monitor_ssl:
                        # Check if domain is already in ssl_domains
                        ssl_exists = False
                        for entry in config['ssl_domains']:
                            if entry.get('url') == domain:
                                ssl_exists = True
                                break

                        # Add to ssl_domains_added if not already in config
                        if not ssl_exists:
                            ssl_domains_added.append(domain)

                            # Force an immediate check to populate the SSL cache
                            try:
                                # Clear any existing cache entry
                                db.clear_cache(f"ssl_{domain}")

                                # Perform a check to populate the cache
                                check_certificate(domain)
                            except Exception as e:
                                logger.error(f"Error checking SSL for {domain}: {str(e)}")

                    if monitor_expiry:
                        # Check if domain is already in domain_expiry
                        expiry_exists = False
                        for entry in config['domain_expiry']:
                            if entry.get('name') == domain:
                                expiry_exists = True
                                break

                        # Add to expiry_domains_added if not already in config
                        if not expiry_exists:
                            expiry_domains_added.append(domain)

                            # Force an immediate check to populate the domain expiry cache
                            try:
                                # Clear any existing cache entry
                                db.clear_cache(f"domain_expiry_{domain}")

                                # Perform a check to populate the cache
                                check_domain_expiry(domain)
                            except Exception as e:
                                logger.error(f"Error checking domain expiry for {domain}: {str(e)}")

                    if monitor_ping:
                        # Check if domain is already in ping_hosts
                        ping_exists = False
                        for entry in config['ping_hosts']:
                            if entry.get('host') == domain:
                                ping_exists = True
                                break

                        # Add to ping_hosts_added if not already in config
                        if not ping_exists:
                            ping_hosts_added.append(domain)

                            # Force an immediate ping check
                            try:
                                # Clear any existing cache entry
                                db.clear_cache(f"ping_{domain}")

                                # Perform a check to populate the cache
                                check_ping(domain)
                            except Exception as e:
                                logger.error(f"Error checking ping for {domain}: {str(e)}")
                else:
                    skipped_count += 1

        # Update the configuration file with all the new domains

        # Add SSL domains to config
        for domain in ssl_domains_added:
            # Check if domain is already in ssl_domains
            ssl_exists = False
            for entry in config['ssl_domains']:
                if entry.get('url') == domain:
                    ssl_exists = True
                    break

            # Add to ssl_domains if not already there
            if not ssl_exists:
                config['ssl_domains'].append({'url': domain})

        # Add domain expiry domains to config
        for domain in expiry_domains_added:
            # Check if domain is already in domain_expiry
            expiry_exists = False
            for entry in config['domain_expiry']:
                if entry.get('name') == domain:
                    expiry_exists = True
                    break

            # Add to domain_expiry if not already there
            if not expiry_exists:
                config['domain_expiry'].append({'name': domain})

        # Add ping hosts to config
        for domain in ping_hosts_added:
            # Check if domain is already in ping_hosts
            ping_exists = False
            for entry in config['ping_hosts']:
                if entry.get('host') == domain:
                    ping_exists = True
                    break

            # Add to ping_hosts if not already there
            if not ping_exists:
                config['ping_hosts'].append({'host': domain})

        # Save the updated configuration if any changes were made
        if ssl_domains_added or expiry_domains_added or ping_hosts_added:
            save_config(config)

        # Create appropriate message
        monitoring_types = []
        if monitor_ssl:
            monitoring_types.append("SSL certificates")
        if monitor_expiry:
            monitoring_types.append("Domain expiry")
        if monitor_ping:
            monitoring_types.append("Ping monitoring")

        monitoring_message = ", ".join(monitoring_types)

        if added_count > 0 or updated_count > 0:
            message = f'Added {added_count} domains'
            if updated_count > 0:
                message += f', updated {updated_count} domains'
            message += f' to {monitoring_message}'
            if skipped_count > 0:
                message += f', skipped {skipped_count} invalid domains'
            flash(message, 'success')
        else:
            flash(f'No domains added or updated, skipped {skipped_count} invalid domains', 'warning')

        return redirect(url_for('index'))

@app.route('/add_ping_from_dashboard', methods=['POST'])
def add_ping_from_dashboard():
    """Add a new host to ping monitoring from the dashboard"""
    if request.method == 'POST':
        host = request.form.get('host', '').strip()

        if not host:
            flash('Please enter a host name or IP address', 'error')
            return redirect(url_for('index'))

        # Clean the host
        host = clean_domain_name(host)
        if not host:
            flash('Invalid host format', 'error')
            return redirect(url_for('index'))

        # Get current user
        current_user = auth.get_current_user()

        # Get current organization
        current_org = current_user.get('current_organization')
        if not current_org:
            flash("You don't have access to any organizations", 'error')
            return redirect(url_for('profile'))

        # Check if host already exists in this organization
        existing_domain = db.get_domain_by_name_and_org(host, current_org['id'])
        if existing_domain:
            # If domain exists and ping monitoring is not enabled, update it
            if not existing_domain['ping_monitored']:
                db.update_domain(existing_domain['id'], host, existing_domain['ssl_monitored'], existing_domain['expiry_monitored'], True)
                flash(f'Host {host} updated to include ping monitoring', 'success')
            else:
                flash(f'Host {host} is already being monitored for ping', 'error')
        else:
            # Add new domain with ping monitoring enabled and auto-logging
            domain_id = db.add_domain(
                name=host,
                organization_id=current_org['id'],
                ssl_monitored=False,
                expiry_monitored=False,
                ping_monitored=True,
                user_id=current_user['id'],
                username=current_user['username'],
                ip_address=request.remote_addr,
                auto_log=True
            )
            if domain_id:
                flash(f'Host {host} added to ping monitoring successfully', 'success')
            else:
                flash(f'Error adding host {host}', 'error')

        return redirect(url_for('index'))

@app.route('/remove_all_monitors/<domain>')
def remove_all_monitors(domain):
    """Remove all monitoring (SSL, domain expiry, ping) for a domain"""
    if not domain:
        flash('Domain name is required', 'error')
        return redirect(url_for('index'))

    domain = clean_domain_name(domain)

    # Get current user
    current_user = auth.get_current_user()

    # Get current organization
    current_org = current_user.get('current_organization')
    if not current_org:
        flash("You don't have access to any organizations", 'error')
        return redirect(url_for('profile'))

    # Load the current configuration
    config = load_config()

    # Get domain by name and organization
    existing_domain = db.get_domain_by_name_and_org(domain, current_org['id'])
    if existing_domain:
        # Delete the domain from the database
        if db.delete_domain(existing_domain['id']):
            # Log the domain deletion action
            db.log_user_action(
                user_id=current_user['id'],
                username=current_user['username'],
                action_type='delete',
                resource_type='domain',
                resource_id=existing_domain['id'],
                resource_name=domain,
                details=f'Removed domain from all monitoring (dashboard)',
                ip_address=request.remote_addr,
                organization_id=current_org['id']
            )

            # Remove from SSL domains in config
            if 'ssl_domains' in config:
                config['ssl_domains'] = [entry for entry in config['ssl_domains']
                                        if entry.get('url') != domain]

            # Remove from domain expiry in config
            if 'domain_expiry' in config:
                config['domain_expiry'] = [entry for entry in config['domain_expiry']
                                          if entry.get('name') != domain]

            # Remove from ping hosts in config
            if 'ping_hosts' in config:
                config['ping_hosts'] = [entry for entry in config['ping_hosts']
                                       if entry.get('host') != domain]

            # Save the updated configuration
            save_config(config)

            flash(f'All monitoring for {domain} has been removed', 'success')
        else:
            flash(f'Error removing monitoring for {domain}', 'error')
    else:
        flash(f'Domain {domain} not found', 'error')

    return redirect(url_for('index'))

@app.route('/notifications')
def notifications():
    """Notification settings page"""
    notification_settings = get_notification_settings()
    return render_template('notifications.html', notifications=notification_settings.__dict__)

@app.route('/save_notifications', methods=['POST'])
def save_notifications():
    """Save notification settings"""
    if request.method == 'POST':
        notification_type = request.form.get('notification_type')
        enabled = request.form.get('enabled') == 'true'

        config = load_config()
        if 'notifications' not in config:
            config['notifications'] = {}

        if notification_type == 'email':
            if 'email' not in config['notifications']:
                config['notifications']['email'] = {}

            config['notifications']['email']['enabled'] = enabled
            if enabled:
                config['notifications']['email']['smtp_server'] = request.form.get('smtp_server', '')

                # Safely convert smtp_port to integer with a default value
                try:
                    smtp_port = request.form.get('smtp_port', '')
                    config['notifications']['email']['smtp_port'] = int(smtp_port) if smtp_port.strip() else 587
                except (ValueError, TypeError):
                    logger.warning(f"Invalid smtp_port value: {request.form.get('smtp_port')}. Using default value of 587.")
                    config['notifications']['email']['smtp_port'] = 587

                config['notifications']['email']['smtp_username'] = request.form.get('smtp_username', '')
                config['notifications']['email']['from_email'] = request.form.get('from_email', '')

                # Only update password if provided
                new_password = request.form.get('smtp_password', '')
                if new_password:
                    config['notifications']['email']['smtp_password'] = new_password

                config['notifications']['email']['notification_email'] = request.form.get('notification_email', '')

                # Get warning threshold from app settings
                app_settings = get_app_settings()
                warning_threshold = app_settings.warning_threshold_days

                # Log the warning threshold for debugging
                logger.info(f"Using warning threshold from app settings for email notifications: {warning_threshold}")

                # Set warning threshold from app settings and save directly to database
                config['notifications']['email']['warning_threshold_days'] = warning_threshold

                # Save the updated notifications config to the database
                db.set_setting('notifications', config['notifications'])

            flash('Email notification settings saved successfully', 'success')

        elif notification_type == 'teams':
            if 'teams' not in config['notifications']:
                config['notifications']['teams'] = {}

            config['notifications']['teams']['enabled'] = enabled
            if enabled:
                config['notifications']['teams']['webhook_url'] = request.form.get('webhook_url', '')

            flash('Microsoft Teams notification settings saved successfully', 'success')

        elif notification_type == 'slack':
            if 'slack' not in config['notifications']:
                config['notifications']['slack'] = {}

            config['notifications']['slack']['enabled'] = enabled
            if enabled:
                config['notifications']['slack']['webhook_url'] = request.form.get('webhook_url', '')
                config['notifications']['slack']['channel'] = request.form.get('channel', '')

            flash('Slack notification settings saved successfully', 'success')

        elif notification_type == 'discord':
            if 'discord' not in config['notifications']:
                config['notifications']['discord'] = {}

            config['notifications']['discord']['enabled'] = enabled
            if enabled:
                config['notifications']['discord']['webhook_url'] = request.form.get('webhook_url', '')
                config['notifications']['discord']['username'] = request.form.get('username', 'Certifly Bot')

            flash('Discord notification settings saved successfully', 'success')

        elif notification_type == 'telegram':
            if 'telegram' not in config['notifications']:
                config['notifications']['telegram'] = {}

            config['notifications']['telegram']['enabled'] = enabled
            if enabled:
                config['notifications']['telegram']['bot_token'] = request.form.get('bot_token', '')
                config['notifications']['telegram']['chat_id'] = request.form.get('chat_id', '')

            flash('Telegram notification settings saved successfully', 'success')

        elif notification_type == 'webhook':
            if 'webhook' not in config['notifications']:
                config['notifications']['webhook'] = {}

            config['notifications']['webhook']['enabled'] = enabled
            if enabled:
                config['notifications']['webhook']['webhook_url'] = request.form.get('webhook_url', '')
                config['notifications']['webhook']['format'] = request.form.get('format', 'json')

                # Validate JSON format for custom headers and fields
                custom_headers = request.form.get('custom_headers', '{}')
                custom_fields = request.form.get('custom_fields', '{}')

                try:
                    json.loads(custom_headers)
                    config['notifications']['webhook']['custom_headers'] = custom_headers
                except json.JSONDecodeError:
                    logger.warning(f"Invalid custom_headers JSON format: {custom_headers}. Using default value.")
                    config['notifications']['webhook']['custom_headers'] = '{}'

                try:
                    json.loads(custom_fields)
                    config['notifications']['webhook']['custom_fields'] = custom_fields
                except json.JSONDecodeError:
                    logger.warning(f"Invalid custom_fields JSON format: {custom_fields}. Using default value.")
                    config['notifications']['webhook']['custom_fields'] = '{}'

            flash('Custom webhook notification settings saved successfully', 'success')

        elif notification_type == 'sms':
            if 'sms' not in config['notifications']:
                config['notifications']['sms'] = {}

            config['notifications']['sms']['enabled'] = enabled
            if enabled:
                config['notifications']['sms']['account_sid'] = request.form.get('account_sid', '')
                config['notifications']['sms']['auth_token'] = request.form.get('auth_token', '')
                config['notifications']['sms']['from_number'] = request.form.get('from_number', '')
                config['notifications']['sms']['to_number'] = request.form.get('to_number', '')

            flash('SMS notification settings saved successfully', 'success')

        save_config(config)
        return redirect(url_for('notifications'))

    return redirect(url_for('notifications'))

@app.route('/test_notification', methods=['POST'])
def test_notification():
    """Test notification settings"""
    if request.method == 'POST':
        notification_type = request.form.get('notification_type')

        valid_types = ['email', 'teams', 'slack', 'discord', 'telegram', 'webhook', 'sms']
        if notification_type not in valid_types:
            flash(f"Unknown notification type: {notification_type}", 'danger')
            return redirect(url_for('notifications'))

        notification_settings = get_notification_settings()
        settings = getattr(notification_settings, notification_type, {})

        if not settings.get('enabled', False):
            flash(f"{notification_type.capitalize()} notifications are disabled", 'warning')
            return redirect(url_for('notifications'))

        # Use the already imported notifications module
        success, message = notifications_module.send_test_notification(notification_type, settings)

        if success:
            flash(f'Test notification sent successfully: {message}', 'success')
        else:
            flash(f'Failed to send test notification: {message}', 'error')

        return redirect(url_for('notifications'))

    return redirect(url_for('notifications'))

@app.route('/check_and_send_summary', methods=['POST'])
@auth.login_required
def manual_check_and_send_summary():
    """Manually check all domains and send a summary email"""
    try:
        # Check if the user is an admin
        current_user = auth.get_current_user()
        if not current_user['is_admin']:
            flash('Only administrators can manually trigger summary notifications', 'error')
            return redirect(url_for('notifications'))

        # Run the check and send the summary
        success, message = check_all_domains_and_send_summary()

        if success:
            flash(f'Summary check completed: {message}', 'success')
        else:
            flash(f'Summary check failed: {message}', 'error')

        return redirect(url_for('notifications'))
    except Exception as e:
        logger.error(f"Error in manual summary check: {str(e)}", exc_info=True)
        flash(f'Error running summary check: {str(e)}', 'error')
        return redirect(url_for('notifications'))

def check_all_domains_and_send_summary():
    """
    Check all domains for SSL certificate, domain expiry, and ping status issues
    and send a summary email notification with all warnings and errors
    """
    try:
        logger.info("Checking all domains for issues and preparing summary...")
        config = load_config()
        notification_settings = get_notification_settings()

        # Only proceed if email notifications are enabled
        if not notification_settings.email.get('enabled', False):
            logger.info("Email notifications are not enabled. Skipping summary check.")
            return False, "Email notifications are not enabled"

        # Initialize lists to store issues
        ssl_issues = []
        domain_expiry_issues = []
        ping_issues = []

        # Check SSL certificates
        for entry in config.get('ssl_domains', []):
            domain_name = entry.get('url')
            if not domain_name:
                continue

            # Get the current certificate status
            cert_status = check_certificate(domain_name)

            # If the certificate is in warning or expired state, add to issues list
            if cert_status.status in ['warning', 'expired']:
                ssl_issues.append({
                    'domain': domain_name,
                    'status': cert_status.status,
                    'days_remaining': cert_status.days_remaining,
                    'expiry_date': cert_status.expiry_date.strftime('%Y-%m-%d')
                })

        # Check domain expiry
        for entry in config.get('domain_expiry', []):
            domain_name = entry.get('name')
            if not domain_name:
                continue

            # Get the current domain expiry status
            domain_status = check_domain_expiry(domain_name)

            # If the domain is in warning or expired state, add to issues list
            if domain_status.status in ['warning', 'expired']:
                domain_expiry_issues.append({
                    'domain': domain_name,
                    'status': domain_status.status,
                    'days_remaining': domain_status.days_remaining,
                    'expiry_date': domain_status.expiry_date.strftime('%Y-%m-%d'),
                    'registrar': domain_status.registrar
                })

        # Check ping status
        for entry in config.get('ping_hosts', []):
            domain_name = entry.get('host')
            if not domain_name:
                continue

            # Get the current ping status
            ping_result = check_ping(domain_name)

            # If the ping status is down, add to issues list
            if ping_result['status'] == 'down':
                ping_issues.append({
                    'domain': domain_name,
                    'status': 'down',
                    'last_checked': ping_result['last_checked'].strftime('%Y-%m-%d %H:%M:%S')
                })

        # If there are no issues, don't send a notification
        if not ssl_issues and not domain_expiry_issues and not ping_issues:
            logger.info("No issues found. Skipping summary notification.")
            return True, "No issues found"

        # Prepare the summary message
        subject = f"Certifly Daily Summary - {datetime.now().strftime('%Y-%m-%d')}"

        message = f"""
Certifly Daily Monitoring Summary - {datetime.now().strftime('%Y-%m-%d')}

This is an automated summary of all domain monitoring issues detected in your Certifly instance.

"""

        # Add SSL certificate issues to the message
        if ssl_issues:
            message += f"""
SSL CERTIFICATE ISSUES ({len(ssl_issues)})
----------------------------------------
"""
            for issue in ssl_issues:
                message += f"""
Domain: {issue['domain']}
Status: {issue['status'].upper()}
Days Remaining: {issue['days_remaining']}
Expiry Date: {issue['expiry_date']}
"""

        # Add domain expiry issues to the message
        if domain_expiry_issues:
            message += f"""
DOMAIN EXPIRY ISSUES ({len(domain_expiry_issues)})
----------------------------------------
"""
            for issue in domain_expiry_issues:
                message += f"""
Domain: {issue['domain']}
Status: {issue['status'].upper()}
Days Remaining: {issue['days_remaining']}
Expiry Date: {issue['expiry_date']}
Registrar: {issue['registrar']}
"""

        # Add ping issues to the message
        if ping_issues:
            message += f"""
PING MONITORING ISSUES ({len(ping_issues)})
----------------------------------------
"""
            for issue in ping_issues:
                message += f"""
Domain: {issue['domain']}
Status: {issue['status'].upper()}
Last Checked: {issue['last_checked']}
"""

        # Add footer
        message += f"""
----------------------------------------
This is an automated message from Certifly. Please do not reply to this email.
"""

        # Import the send_email_notification function from the notifications module
        from notifications import send_email_notification

        # Send the summary email
        success, result_message = send_email_notification(notification_settings.email, subject, message)

        if success:
            logger.info("Summary notification email sent successfully")
            # Log the notification
            db.log_notification_sent("summary", "summary", "daily")
            return True, "Summary notification email sent successfully"
        else:
            logger.error(f"Failed to send summary notification email: {result_message}")
            return False, f"Failed to send summary notification email: {result_message}"

    except Exception as e:
        error_message = f"Error checking domains and sending summary: {str(e)}"
        logger.error(error_message, exc_info=True)
        return False, error_message

def check_and_send_pending_notifications():
    """Check for domains with warning or expired SSL certificates and send notifications if needed"""
    try:
        logger.info("Checking for pending notifications...")
        config = load_config()
        notification_settings = get_notification_settings()

        # Only proceed if at least one notification channel is enabled
        if not any([
            notification_settings.email.get('enabled', False),
            notification_settings.teams.get('enabled', False),
            notification_settings.slack.get('enabled', False),
            notification_settings.discord.get('enabled', False),
            notification_settings.telegram.get('enabled', False),
            notification_settings.webhook.get('enabled', False),
            notification_settings.sms.get('enabled', False)
        ]):
            logger.info("No notification channels are enabled. Skipping notification check.")
            return

        # Check SSL certificates
        for entry in config.get('ssl_domains', []):
            domain_name = entry.get('url')
            if not domain_name:
                continue

            # Get the current certificate status
            cert_status = check_certificate(domain_name)

            # If the certificate is in warning or expired state, check if we need to send a notification
            if cert_status.status in ['warning', 'expired']:
                should_notify = db.should_send_notification(domain_name, 'ssl', cert_status.status)

                if should_notify:
                    logger.info(f"Sending pending notification for {domain_name} SSL certificate ({cert_status.status})")
                    notifications_sent = False

                    # Send notifications to all enabled platforms
                    if notification_settings.email.get('enabled', False):
                        success, message = notifications_module.send_certificate_expiry_notification('email', notification_settings.email, cert_status)
                        if success:
                            notifications_sent = True
                            logger.info(f"Email notification sent for {domain_name}: {message}")
                        else:
                            logger.error(f"Failed to send email notification for {domain_name}: {message}")

                    # Add similar blocks for other notification channels as needed

                    # Log the notification if any were sent successfully
                    if notifications_sent:
                        db.log_notification_sent(domain_name, 'ssl', cert_status.status)
                        logger.info(f"Logged notification for {domain_name} SSL certificate ({cert_status.status})")
                else:
                    logger.debug(f"Skipping notification for {domain_name} SSL certificate ({cert_status.status}) - already sent recently")

        logger.info("Finished checking for pending notifications")
    except Exception as e:
        logger.error(f"Error checking for pending notifications: {str(e)}", exc_info=True)



@app.route('/check_ping/<domain>')
def check_domain_ping(domain):
    """Check ping status for a domain and return JSON response"""
    ping_result = check_ping(domain)

    # Get uptime percentage
    uptime_percentage = db.calculate_uptime_percentage(domain)
    if uptime_percentage is not None:
        uptime_percentage = round(uptime_percentage, 1)

    # Get ping history for uptime segments
    ping_history = db.get_ping_history(domain, hours=12)

    # Create uptime segments (12 segments representing the last 12 hours)
    uptime_segments = []
    if ping_history:
        # Group history into 12 hourly segments
        current_time = datetime.now()

        for i in range(12):
            segment_start = current_time - timedelta(hours=12-i)
            segment_end = current_time - timedelta(hours=11-i)

            # Find all checks in this segment
            segment_checks = [check for check in ping_history
                             if check['checked_at'] and
                             segment_start <= datetime.fromtimestamp(check['checked_at']) < segment_end]

            if segment_checks:
                # If any check is down, the segment is down
                if any(check['status'] == 'down' for check in segment_checks):
                    uptime_segments.append('down')
                else:
                    uptime_segments.append('up')
            else:
                # No checks in this segment
                uptime_segments.append('unknown')
    else:
        # If no history, use current status only for the most recent segment (last hour)
        uptime_segments = ['unknown'] * 11 + [ping_result["status"]]

    # Update ping status in config for ping hosts
    config = load_config()
    ping_hosts = config.get('ping_hosts', [])
    response_history = []

    for ping_host in ping_hosts:
        if ping_host.get('host') == domain:
            # Initialize response history if it doesn't exist
            if 'response_history' not in ping_host:
                ping_host['response_history'] = []

            # Add new response time to history (keep last 20 entries)
            if ping_result["status"] == "up":
                ping_host['response_history'].append(ping_result["response_time"])
                # Keep only the last 20 entries
                ping_host['response_history'] = ping_host['response_history'][-20:]

            response_history = ping_host.get('response_history', [])

            # Save the updated config
            save_config(config)
            break

    # Return response with history and uptime data
    return {
        'domain': domain,
        'ping_status': ping_result["status"],
        'response_time': ping_result["response_time"],
        'response_history': response_history,
        'uptime_percentage': uptime_percentage,
        'uptime_segments': uptime_segments
    }

@app.route('/api/domains/<int:domain_id>', methods=['DELETE'])
def delete_domain(domain_id):
    try:
        # Get current user
        current_user = auth.get_current_user()

        # Get current organization
        current_org = current_user.get('current_organization')
        if not current_org:
            return jsonify({'success': False, 'error': 'You don\'t have access to any organizations'}), 403

        # Get domain by ID
        domain = db.get_domain(domain_id)
        if not domain:
            return jsonify({'success': False, 'error': 'Domain not found'}), 404

        # Check if domain belongs to current organization
        if domain['organization_id'] != current_org['id']:
            return jsonify({'success': False, 'error': 'You don\'t have permission to delete this domain'}), 403

        # Load the current configuration
        config = load_config()

        # Get domain name before deleting
        domain_name = domain['name']

        # Delete domain from database
        if db.delete_domain(domain_id):
            # Log the domain deletion action
            db.log_user_action(
                user_id=current_user['id'],
                username=current_user['username'],
                action_type='delete',
                resource_type='domain',
                resource_id=domain_id,
                resource_name=domain_name,
                details=f'Removed domain from all monitoring (API)',
                ip_address=request.remote_addr,
                organization_id=current_org['id']
            )

            # Remove from SSL domains in config
            if 'ssl_domains' in config:
                config['ssl_domains'] = [entry for entry in config['ssl_domains']
                                        if entry.get('url') != domain_name]

            # Remove from domain expiry in config
            if 'domain_expiry' in config:
                config['domain_expiry'] = [entry for entry in config['domain_expiry']
                                          if entry.get('name') != domain_name]

            # Remove from ping hosts in config
            if 'ping_hosts' in config:
                config['ping_hosts'] = [entry for entry in config['ping_hosts']
                                       if entry.get('host') != domain_name]

            # Save the updated configuration
            save_config(config)

            logger.info(f"Domain {domain_name} (ID: {domain_id}) deleted successfully")
            return jsonify({'success': True, 'message': f'Domain {domain_name} deleted successfully'})
        else:
            return jsonify({'success': False, 'error': 'Failed to delete domain'}), 500
    except Exception as e:
        logger.error(f"Error deleting domain {domain_id}: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

def add_cache_headers(response, max_age=0, private=True, etag=None):
    """Add cache control headers to a response"""
    if private:
        response.cache_control.private = True
    else:
        response.cache_control.public = True

    response.cache_control.max_age = max_age

    if etag:
        response.set_etag(etag)

    return response

@app.route('/api/domains/<int:domain_id>', methods=['GET'])
def get_domain(domain_id):
    try:
        # Get current user
        current_user = auth.get_current_user()

        # Get current organization
        current_org = current_user.get('current_organization')
        if not current_org:
            return jsonify({'success': False, 'error': 'You don\'t have access to any organizations'}), 403

        # Get domain by ID
        domain = db.get_domain(domain_id)
        if not domain:
            return jsonify({'success': False, 'error': 'Domain not found'}), 404

        # Check if domain belongs to current organization
        if domain['organization_id'] != current_org['id']:
            return jsonify({'success': False, 'error': 'You don\'t have permission to view this domain'}), 403

        # Determine which monitoring services are enabled for this domain
        monitors = []
        if domain['ssl_monitored']:
            monitors.append('ssl')
        if domain['expiry_monitored']:
            monitors.append('expiry')
        if domain['ping_monitored']:
            monitors.append('ping')

        # Create response data with both formats for backward compatibility
        response_data = {
            'success': True,
            'domain': {
                'id': domain['id'],
                'name': domain['name'],
                'ssl_monitored': domain['ssl_monitored'],
                'expiry_monitored': domain['expiry_monitored'],
                'ping_monitored': domain['ping_monitored'],
                'organization_id': domain['organization_id']
            },
            'data': {
                'domain': domain['name'],
                'monitors': monitors
            }
        }

        # Generate ETag based on response data
        import hashlib
        etag = hashlib.md5(str(response_data).encode()).hexdigest()

        # Create response with cache headers
        response = jsonify(response_data)
        return add_cache_headers(response, max_age=60, private=True, etag=etag)

    except Exception as e:
        logger.error(f"Error getting domain details: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/edit_domain_from_dashboard', methods=['POST'])
def edit_domain_from_dashboard():
    """Edit domain monitoring settings from the dashboard"""
    if request.method == 'POST':
        domain_id = request.form.get('domain_id')
        domain = request.form.get('domain')
        original_domain = request.form.get('original_domain')
        monitor_ssl = 'monitor_ssl' in request.form
        monitor_expiry = 'monitor_expiry' in request.form
        monitor_ping = 'monitor_ping' in request.form

        if not domain or not domain_id or not original_domain:
            flash('Domain information is missing', 'error')
            return redirect(url_for('index'))

        try:
            domain_id = int(domain_id)
        except ValueError:
            flash('Invalid domain ID', 'error')
            return redirect(url_for('index'))

        # Clean the domain names
        domain = clean_domain_name(domain)
        original_domain = clean_domain_name(original_domain)

        if not domain:
            flash('Invalid domain format', 'error')
            return redirect(url_for('index'))

        # Get current user
        current_user = auth.get_current_user()

        # Get current organization
        current_org = current_user.get('current_organization')
        if not current_org:
            flash("You don't have access to any organizations", 'error')
            return redirect(url_for('profile'))

        # Load the current configuration
        config = load_config()

        # Initialize config sections if they don't exist
        if 'ssl_domains' not in config:
            config['ssl_domains'] = []
        if 'domain_expiry' not in config:
            config['domain_expiry'] = []
        if 'ping_hosts' not in config:
            config['ping_hosts'] = []

        # Get the domain by ID
        existing_domain = db.get_domain(domain_id)
        if not existing_domain:
            flash(f'Domain with ID {domain_id} not found', 'error')
            return redirect(url_for('index'))

        # Check if the domain belongs to the current organization
        if existing_domain['organization_id'] != current_org['id']:
            flash("You don't have permission to edit this domain", 'error')
            return redirect(url_for('index'))

        # Check if the new domain name already exists (only if domain name was changed)
        if domain != original_domain:
            # Check if domain already exists in this organization
            domain_exists = db.get_domain_by_name_and_org(domain, current_org['id'])
            if domain_exists:
                flash(f'Domain {domain} is already being monitored. Please use a different name.', 'error')
                return redirect(url_for('index'))

        # Update the domain in the database
        updated = db.update_domain(
            domain_id,
            domain,  # New domain name
            monitor_ssl,
            monitor_expiry,
            monitor_ping
        )

        if updated:
            # Update the configuration file

            # If domain name was changed, remove old domain from config
            if domain != original_domain:
                # Remove old domain from SSL domains
                if 'ssl_domains' in config:
                    config['ssl_domains'] = [entry for entry in config['ssl_domains']
                                            if entry.get('url') != original_domain]

                # Remove old domain from domain expiry
                if 'domain_expiry' in config:
                    config['domain_expiry'] = [entry for entry in config['domain_expiry']
                                              if entry.get('name') != original_domain]

                # Remove old domain from ping hosts
                if 'ping_hosts' in config:
                    config['ping_hosts'] = [entry for entry in config['ping_hosts']
                                           if entry.get('host') != original_domain]

            # Update SSL monitoring
            if monitor_ssl:
                # Check if domain is already in ssl_domains
                ssl_exists = False
                for entry in config['ssl_domains']:
                    if entry.get('url') == domain:
                        ssl_exists = True
                        break

                # Add to ssl_domains if not already there
                if not ssl_exists:
                    config['ssl_domains'].append({'url': domain})
            else:
                # Remove from SSL domains if monitoring is disabled
                config['ssl_domains'] = [entry for entry in config['ssl_domains']
                                        if entry.get('url') != domain]

            # Update domain expiry monitoring
            if monitor_expiry:
                # Check if domain is already in domain_expiry
                expiry_exists = False
                for entry in config['domain_expiry']:
                    if entry.get('name') == domain:
                        expiry_exists = True
                        break

                # Add to domain_expiry if not already there
                if not expiry_exists:
                    config['domain_expiry'].append({'name': domain})
            else:
                # Remove from domain expiry if monitoring is disabled
                config['domain_expiry'] = [entry for entry in config['domain_expiry']
                                          if entry.get('name') != domain]

            # Update ping monitoring
            if monitor_ping:
                # Check if domain is already in ping_hosts
                ping_exists = False
                for entry in config['ping_hosts']:
                    if entry.get('host') == domain:
                        ping_exists = True
                        break

                # Add to ping_hosts if not already there
                if not ping_exists:
                    config['ping_hosts'].append({'host': domain})
            else:
                # Remove from ping hosts if monitoring is disabled
                config['ping_hosts'] = [entry for entry in config['ping_hosts']
                                       if entry.get('host') != domain]

            # Save the updated configuration
            save_config(config)

            # Show appropriate success message based on whether the domain name was changed
            if domain != original_domain:
                flash(f'Domain changed from {original_domain} to {domain} and monitoring settings updated successfully', 'success')
            else:
                flash(f'Domain {domain} monitoring settings updated successfully', 'success')
        else:
            flash(f'Error updating domain {domain}', 'error')

        return redirect(url_for('index'))

@app.route('/api/domains/<int:domain_id>/refresh', methods=['POST'])
def refresh_domain(domain_id):
    try:
        start_time = time.time()
        logger.debug(f"Starting refresh for domain ID {domain_id}")

        # Get current user
        current_user = auth.get_current_user()

        # Get current organization
        current_org = current_user.get('current_organization')
        if not current_org:
            return jsonify({'success': False, 'error': 'You don\'t have access to any organizations'}), 403

        # Get domain by ID
        domain = db.get_domain(domain_id)
        if not domain:
            return jsonify({'success': False, 'error': 'Domain not found'}), 404

        # Check if domain belongs to current organization
        if domain['organization_id'] != current_org['id']:
            return jsonify({'success': False, 'error': 'You don\'t have permission to refresh this domain'}), 403

        domain_to_refresh = domain['name']

        # Refresh domain data
        domain_data = {}

        # Check if domain is monitored for SSL
        if domain['ssl_monitored']:
            # Clear SSL cache for this domain to force a fresh check
            db.clear_cache(f"ssl_{domain_to_refresh}")

            cert_status = check_certificate(domain_to_refresh)
            domain_data['ssl_status'] = {
                'status': cert_status.status,
                'days_remaining': cert_status.days_remaining,
                'expiry_date': cert_status.expiry_date.strftime('%Y-%m-%d')
            }

        # Check if domain is monitored for expiry
        if domain['expiry_monitored']:
            # Clear domain expiry cache for this domain to force a fresh check
            db.clear_cache(f"domain_expiry_{domain_to_refresh}")

            domain_status = check_domain_expiry(domain_to_refresh)
            domain_data['domain_status'] = {
                'status': domain_status.status,
                'days_remaining': domain_status.days_remaining,
                'expiry_date': domain_status.expiry_date.strftime('%Y-%m-%d'),
                'registrar': domain_status.registrar
            }

        # Clear ping cache for this domain to force a fresh check
        db.clear_cache(f"ping_{domain_to_refresh}")

        # Check ping status
        ping_result = check_ping(domain_to_refresh)
        domain_data['ping_status'] = ping_result

        end_time = time.time()
        logger.info(f"Domain {domain_to_refresh} (ID: {domain_id}) refreshed successfully in {end_time - start_time:.2f} seconds")

        # Create response data
        response_data = {
            'success': True,
            'message': f'Domain {domain_to_refresh} refreshed successfully',
            'data': domain_data,
            'timestamp': int(time.time())
        }

        # Generate ETag based on response data
        import hashlib
        etag = hashlib.md5(str(response_data).encode()).hexdigest()

        # Create response with cache headers - no caching for refresh results
        response = jsonify(response_data)
        return add_cache_headers(response, max_age=0, private=True, etag=etag)
    except Exception as e:
        logger.error(f"Error refreshing domain {domain_id}: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

# Storage route removed as requested

# Template filters
@app.template_filter('datetime')
def format_datetime(value, format='%Y-%m-%d %H:%M:%S'):
    """Format a datetime object to a string"""
    if value is None:
        return ""

    # Handle numeric timestamps (stored as int or float)
    if isinstance(value, (int, float)):
        try:
            value = datetime.fromtimestamp(value)
            return value.strftime(format)
        except (ValueError, OverflowError, OSError):
            return str(value)

    # Handle string values
    if isinstance(value, str):
        try:
            # Try to parse as ISO format
            value = datetime.fromisoformat(value.replace('Z', '+00:00'))
            return value.strftime(format)
        except ValueError:
            try:
                # Try to parse as timestamp
                value = datetime.fromtimestamp(float(value))
                return value.strftime(format)
            except (ValueError, OverflowError, OSError):
                # If all parsing fails, return the string as is
                return value

    # If it's already a datetime object, format it
    try:
        return value.strftime(format)
    except AttributeError:
        # If all else fails, return as string
        return str(value)



# Organization Routes
@app.route('/organizations')
@auth.login_required
def organizations():
    """Organizations page"""
    user = auth.get_current_user()

    # Get all organizations for admin, or user's organizations for regular users
    if user['is_admin']:
        organizations = db.get_all_organizations()
    else:
        organizations = user['organizations']

    # Sort organizations: current organization first, then alphabetically by name
    current_org_id = session.get('current_organization_id')

    # Separate current organization and other organizations
    current_org = None
    other_orgs = []

    for org in organizations:
        if org['id'] == current_org_id:
            current_org = org
        else:
            other_orgs.append(org)

    # Sort other organizations alphabetically by name
    other_orgs = sorted(other_orgs, key=lambda x: x['name'].lower())

    # Combine current organization and other organizations
    sorted_organizations = []
    if current_org:
        sorted_organizations.append(current_org)
    sorted_organizations.extend(other_orgs)

    return render_template('organizations.html',
                          organizations=sorted_organizations,
                          user=user)

@app.route('/organizations/switch/<int:org_id>')
@auth.login_required
def switch_organization(org_id):
    """Switch current organization"""
    user = auth.get_current_user()

    # Check if user has access to this organization
    has_access = False
    for org in user['organizations']:
        if org['id'] == org_id:
            has_access = True
            break

    # Admin users have access to all organizations
    if user['is_admin'] and not has_access:
        org = db.get_organization(org_id)
        if org:
            has_access = True

    if has_access:
        session['current_organization_id'] = org_id
        flash(f"Switched to organization: {org['name']}", 'success')
    else:
        flash("You don't have access to this organization", 'error')

    from urllib.parse import urlparse
    referrer = request.referrer
    if referrer:
        parsed_url = urlparse(referrer)
        if parsed_url.netloc != request.host:
            referrer = None
    return redirect(referrer or url_for('index'))

@app.route('/organizations/create', methods=['GET', 'POST'])
@auth.admin_required
def create_organization():
    """Create a new organization"""
    if request.method == 'POST':
        name = request.form.get('name')
        description = request.form.get('description')

        if not name:
            flash('Organization name is required', 'error')
            return redirect(url_for('create_organization'))

        # Check if organization already exists
        existing_org = db.get_organization_by_name(name)
        if existing_org:
            flash(f'Organization "{name}" already exists', 'error')
            return redirect(url_for('create_organization'))

        # Create organization
        org_id = db.create_organization(name, description)
        if org_id:
            flash(f'Organization "{name}" created successfully', 'success')
            return redirect(url_for('organizations'))
        else:
            flash('Failed to create organization', 'error')
            return redirect(url_for('create_organization'))

    return render_template('organization_form.html',
                          title='Create Organization',
                          user=auth.get_current_user())

@app.route('/organizations/<int:org_id>/edit', methods=['GET', 'POST'])
@auth.admin_required
def edit_organization(org_id):
    """Edit an organization"""
    org = db.get_organization(org_id)
    if not org:
        flash('Organization not found', 'error')
        return redirect(url_for('organizations'))

    if request.method == 'POST':
        form_type = request.form.get('form_type', 'organization_details')

        if form_type == 'organization_details':
            name = request.form.get('name')
            description = request.form.get('description')

            if not name:
                flash('Organization name is required', 'error')
                return redirect(url_for('edit_organization', org_id=org_id))

            # Check if new name already exists for another organization
            existing_org = db.get_organization_by_name(name)
            if existing_org and existing_org['id'] != org_id:
                flash(f'Organization "{name}" already exists', 'error')
                return redirect(url_for('edit_organization', org_id=org_id))

            # Update organization
            success = db.update_organization(org_id, name, description)
            if success:
                flash(f'Organization "{name}" updated successfully', 'success')
                return redirect(url_for('edit_organization', org_id=org_id))
            else:
                flash('Failed to update organization', 'error')
                return redirect(url_for('edit_organization', org_id=org_id))

    # Get organization users, tags, and domains for the editor
    users = db.get_organization_users(org_id)
    tags = db.get_organization_tags(org_id)
    domains = db.get_domains_by_organization(org_id)

    # Get SSL status for each domain
    for domain in domains:
        if domain['ssl_monitored']:
            cert_status = check_certificate(domain['name'])
            domain['ssl_status'] = cert_status.status
        else:
            domain['ssl_status'] = None

    return render_template('organization_editor.html',
                          organization=org,
                          users=users,
                          tags=tags,
                          domains=domains,
                          user=auth.get_current_user())

@app.route('/organizations/<int:org_id>/delete', methods=['POST'])
@auth.admin_required
def delete_organization(org_id):
    """Delete an organization"""
    org = db.get_organization(org_id)
    if not org:
        flash('Organization not found', 'error')
        return redirect(url_for('organizations'))

    # Delete organization
    success = db.delete_organization(org_id)
    if success:
        # If the deleted organization was the current one, reset the current organization
        if session.get('current_organization_id') == org_id:
            session.pop('current_organization_id', None)

        flash(f'Organization "{org["name"]}" deleted successfully', 'success')
    else:
        flash('Failed to delete organization', 'error')

    return redirect(url_for('organizations'))

@app.route('/organizations/<int:org_id>/users')
@auth.organization_admin_required
def organization_users(org_id):
    """Organization users page"""
    user = auth.get_current_user()

    # Check if user has access to this organization
    has_access = user['is_admin']
    if not has_access:
        for org in user['organizations']:
            if org['id'] == org_id and org['role'] == 'admin':
                has_access = True
                break

    if not has_access:
        flash("You don't have permission to view this page", 'error')
        return redirect(url_for('index'))

    org = db.get_organization(org_id)
    if not org:
        flash('Organization not found', 'error')
        return redirect(url_for('organizations'))

    users = db.get_organization_users(org_id)

    return render_template('organization_users.html',
                          organization=org,
                          users=users,
                          user=user)

@app.route('/api/check_user_exists', methods=['GET'])
@auth.login_required
def api_check_user_exists():
    """API endpoint to check if a user exists"""
    username = request.args.get('username')
    if not username:
        return jsonify({'exists': False, 'message': 'Username is required'})

    # Get user by username
    user = db.get_user_by_username(username)
    if user:
        return jsonify({'exists': True, 'message': 'User found'})
    else:
        return jsonify({'exists': False, 'message': 'User not found'})

@app.route('/api/get_user_suggestions', methods=['GET'])
@auth.login_required
def api_get_user_suggestions():
    """API endpoint to get user suggestions"""
    query = request.args.get('query', '')
    org_id = request.args.get('org_id')

    logger.debug(f"User suggestion request - query: '{query}', org_id: {org_id}")

    if not org_id:
        logger.warning("User suggestion request missing org_id")
        return jsonify({'success': False, 'message': 'Organization ID is required'})

    try:
        # Get current user
        current_user = auth.get_current_user()
        logger.debug(f"Current user: {current_user['username']}, admin: {current_user['is_admin']}")

        # Check if user has access to this organization
        has_access = current_user['is_admin']
        if not has_access:
            for org in current_user['organizations']:
                if org['id'] == int(org_id) and org['role'] == 'admin':
                    has_access = True
                    break

        logger.debug(f"User has access to organization: {has_access}")

        if not has_access:
            logger.warning(f"Access denied for user {current_user['username']} to organization {org_id}")
            return jsonify({'success': False, 'message': 'Access denied'})

        # Get users that match the query and are not already in the organization
        users = db.get_users_by_partial_username(query)
        logger.debug(f"Found {len(users)} users matching query '{query}'")

        org_users = db.get_organization_users(int(org_id))
        logger.debug(f"Organization {org_id} has {len(org_users)} users")

        org_user_ids = [user['id'] for user in org_users]

        # Filter out users that are already in the organization
        filtered_users = [user for user in users if user['id'] not in org_user_ids]
        logger.debug(f"After filtering, {len(filtered_users)} users remain")

        # Format user data for suggestions
        suggestions = []
        for user in filtered_users:
            suggestions.append(user['username'])
            logger.debug(f"Adding suggestion: {user['username']}")

        response = {'success': True, 'suggestions': suggestions}
        logger.debug(f"Returning response: {response}")
        return jsonify(response)
    except Exception as e:
        logger.error(f"Error in get_user_suggestions: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'message': f'Server error: {str(e)}'})

@app.route('/api/get_available_users', methods=['GET'])
@auth.login_required
def api_get_available_users():
    """API endpoint to get all available users for an organization"""
    org_id = request.args.get('org_id')
    search = request.args.get('search', '')

    logger.debug(f"Available users request - org_id: {org_id}, search: '{search}'")

    if not org_id:
        logger.warning("Available users request missing org_id")
        return jsonify({'success': False, 'message': 'Organization ID is required'})

    try:
        # Get current user
        current_user = auth.get_current_user()
        logger.debug(f"Current user: {current_user['username']} (ID: {current_user['id']})")

        # Check if user has access to this organization
        has_access = current_user['is_admin']
        if not has_access:
            for org in current_user['organizations']:
                if org['id'] == int(org_id) and org['role'] == 'admin':
                    has_access = True
                    break

        if not has_access:
            logger.warning(f"Access denied for user {current_user['username']} to organization {org_id}")
            return jsonify({'success': False, 'message': 'Access denied'})

        # Get all users
        all_users = db.get_all_users()
        logger.debug(f"Total users in system: {len(all_users)}")

        # Log all users for debugging
        for user in all_users:
            logger.debug(f"User in system: {user['username']} (ID: {user['id']})")

        # Get organization users
        org_users = db.get_organization_users(int(org_id))
        logger.debug(f"Users in organization {org_id}: {len(org_users)}")

        org_user_ids = [user['id'] for user in org_users]
        logger.debug(f"Organization user IDs: {org_user_ids}")

        # Filter out users that are already in the organization
        available_users = []
        for user in all_users:
            if user['id'] not in org_user_ids:
                available_users.append({
                    'id': user['id'],
                    'username': user['username'],
                    'email': user['email'],
                    'is_admin': user['is_admin']
                })
                logger.debug(f"Adding available user: {user['username']} (ID: {user['id']})")

        # Apply search filter if provided
        if search:
            search = search.lower()
            filtered_users = []
            for user in available_users:
                if search in user['username'].lower() or search in user['email'].lower():
                    filtered_users.append(user)
                    logger.debug(f"User matches search: {user['username']}")
            available_users = filtered_users

        logger.debug(f"Found {len(available_users)} available users for organization {org_id}")
        logger.debug(f"Available users: {[user['username'] for user in available_users]}")

        # If no users are available, create a test user for demonstration
        if len(available_users) == 0 and current_user['is_admin']:
            logger.debug("No available users found, creating a test user")
            test_username = f"testuser_{int(time.time())}"
            test_email = f"{test_username}@example.com"

            # Create the test user
            test_user_id = db.add_user(test_username, "password123", test_email, is_admin=False)
            if test_user_id:
                logger.debug(f"Created test user: {test_username} with ID {test_user_id}")
                available_users.append({
                    'id': test_user_id,
                    'username': test_username,
                    'email': test_email,
                    'is_admin': False
                })

        return jsonify({
            'success': True,
            'users': available_users
        })
    except Exception as e:
        logger.error(f"Error in get_available_users: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'message': f'Server error: {str(e)}'})

@app.route('/api/debug_user_suggestions', methods=['POST'])
@auth.login_required
def api_debug_user_suggestions():
    """Debug endpoint for user suggestions"""
    try:
        data = request.json
        query = data.get('query', '')
        org_id = data.get('org_id')

        logger.debug(f"Debug user suggestions - query: '{query}', org_id: {org_id}")

        # Get current user
        current_user = auth.get_current_user()

        # Create a test user if we don't have many users
        all_users = db.get_all_users()

        if len(all_users) < 3:  # If we have fewer than 3 users
            test_username = f"testuser_{int(time.time())}"
            test_email = f"{test_username}@example.com"

            # Create the test user
            test_user_id = db.add_user(test_username, "password123", test_email, is_admin=False)
            logger.debug(f"Created test user: {test_username} with ID {test_user_id}")

            # Refresh the list of all users
            all_users = db.get_all_users()

        # Get users matching the query
        matching_users = db.get_users_by_partial_username(query)

        # Get organization users
        org_users = db.get_organization_users(int(org_id)) if org_id else []

        return jsonify({
            'success': True,
            'debug_info': {
                'query': query,
                'org_id': org_id,
                'current_user': current_user['username'],
                'total_users': len(all_users),
                'matching_users': [u['username'] for u in matching_users],
                'org_users': [u['username'] for u in org_users],
                'test_username': test_username
            }
        })
    except Exception as e:
        logger.error(f"Error in debug_user_suggestions: {str(e)}", exc_info=True)
        return jsonify({'success': False, 'message': f'Debug error: {str(e)}'})

@app.route('/organizations/<int:org_id>/users/add', methods=['GET', 'POST'])
@auth.organization_admin_required
def add_organization_user(org_id):
    """Add users to an organization"""
    user = auth.get_current_user()

    # Check if user has access to this organization
    has_access = user['is_admin']
    if not has_access:
        for org in user['organizations']:
            if org['id'] == org_id and org['role'] == 'admin':
                has_access = True
                break

    if not has_access:
        flash("You don't have permission to view this page", 'error')
        return redirect(url_for('index'))

    org = db.get_organization(org_id)
    if not org:
        flash('Organization not found', 'error')
        return redirect(url_for('organizations'))

    if request.method == 'POST':
        # Get selected users and role
        selected_users = request.form.getlist('selected_users[]')
        role = request.form.get('role', 'member')

        if not selected_users:
            flash('Please select at least one user to add', 'error')
            return redirect(url_for('add_organization_user', org_id=org_id))

        # Track success and failures
        added_users = []
        failed_users = []

        for user_id in selected_users:
            try:
                user_id = int(user_id)

                # Get user by ID
                user_to_add = db.get_user_by_id(user_id)
                if not user_to_add:
                    failed_users.append(f"User ID {user_id} not found")
                    continue

                # Check if user is already in organization
                if db.is_user_in_organization(user_id, org_id):
                    failed_users.append(f"User '{user_to_add['username']}' is already in this organization")
                    continue

                # Add user to organization
                success = db.add_user_to_organization(user_id, org_id, role)
                if success:
                    added_users.append(user_to_add['username'])
                else:
                    failed_users.append(f"Failed to add user '{user_to_add['username']}'")
            except Exception as e:
                logger.error(f"Error adding user {user_id} to organization: {str(e)}")
                failed_users.append(f"Error processing user ID {user_id}")

        # Show success and failure messages
        if added_users:
            if len(added_users) == 1:
                flash(f"User '{added_users[0]}' added to organization successfully", 'success')
            else:
                flash(f"{len(added_users)} users added to organization successfully", 'success')

        if failed_users:
            for message in failed_users:
                flash(message, 'error')

        if added_users:
            return redirect(url_for('organization_users', org_id=org_id))
        else:
            return redirect(url_for('add_organization_user', org_id=org_id))

    # Get current organization users for display
    org_users = db.get_organization_users(org_id)

    # Get all users in the system
    all_users = db.get_all_users()

    # Filter out users that are already in the organization
    org_user_ids = [user['id'] for user in org_users]
    available_users = [
        {
            'id': user['id'],
            'username': user['username'],
            'email': user['email'],
            'is_admin': user['is_admin']
        }
        for user in all_users
        if user['id'] not in org_user_ids
    ]

    logger.debug(f"Found {len(available_users)} available users for organization {org_id}")

    # If no real users are available, add some test users for demonstration
    if len(available_users) == 0:
        available_users = [
            {
                'id': 1001,
                'username': 'testuser1',
                'email': 'testuser1@example.com',
                'is_admin': False
            },
            {
                'id': 1002,
                'username': 'testuser2',
                'email': 'testuser2@example.com',
                'is_admin': False
            },
            {
                'id': 1003,
                'username': 'adminuser',
                'email': 'admin@example.com',
                'is_admin': True
            }
        ]
        logger.debug("No real users available, using test users instead")

    # For GET requests, render the simplified template
    return render_template('add_members_simple.html',
                          organization=org,
                          users=org_users,
                          available_users=available_users,
                          user=user)

@app.route('/organizations/<int:org_id>/users/add/debug', methods=['GET'])
@auth.organization_admin_required
def debug_add_organization_user(org_id):
    """Debug page for adding users to an organization"""
    user = auth.get_current_user()

    # Check if user has access to this organization
    has_access = user['is_admin']
    if not has_access:
        for org in user['organizations']:
            if org['id'] == org_id and org['role'] == 'admin':
                has_access = True
                break

    if not has_access:
        flash("You don't have permission to view this page", 'error')
        return redirect(url_for('index'))

    org = db.get_organization(org_id)
    if not org:
        flash('Organization not found', 'error')
        return redirect(url_for('organizations'))

    # Get current organization users for display
    users = db.get_organization_users(org_id)

    # Render the debug template
    return render_template('debug_info.html',
                          organization=org,
                          users=users,
                          user=user)

@app.route('/organizations/<int:org_id>/users/<int:user_id>/role', methods=['POST'])
@auth.organization_admin_required
def update_organization_user_role(org_id, user_id):
    """Update a user's role in an organization"""
    current_user = auth.get_current_user()

    # Check if user has access to this organization
    has_access = current_user['is_admin']
    if not has_access:
        for org in current_user['organizations']:
            if org['id'] == org_id and org['role'] == 'admin':
                has_access = True
                break

    if not has_access:
        flash("You don't have permission to perform this action", 'error')
        return redirect(url_for('index'))

    # Don't allow changing own role
    if user_id == current_user['id']:
        flash("You cannot change your own role", 'error')
        return redirect(url_for('organization_users', org_id=org_id))

    role = request.form.get('role')
    if not role or role not in ['admin', 'member']:
        flash('Invalid role', 'error')
        return redirect(url_for('organization_users', org_id=org_id))

    # Update user role
    success = db.update_user_organization_role(user_id, org_id, role)
    if success:
        flash('User role updated successfully', 'success')
    else:
        flash('Failed to update user role', 'error')

    return redirect(url_for('organization_users', org_id=org_id))

@app.route('/organizations/<int:org_id>/users/<int:user_id>/remove', methods=['POST'])
@auth.organization_admin_required
def remove_organization_user(org_id, user_id):
    """Remove a user from an organization"""
    current_user = auth.get_current_user()

    # Check if user has access to this organization
    has_access = current_user['is_admin']
    if not has_access:
        for org in current_user['organizations']:
            if org['id'] == org_id and org['role'] == 'admin':
                has_access = True
                break

    if not has_access:
        flash("You don't have permission to perform this action", 'error')
        return redirect(url_for('index'))

    # Don't allow removing self
    if user_id == current_user['id']:
        flash("You cannot remove yourself from the organization", 'error')
        return redirect(url_for('organization_users', org_id=org_id))

    # Remove user from organization
    success = db.remove_user_from_organization(user_id, org_id)
    if success:
        flash('User removed from organization successfully', 'success')
    else:
        flash('Failed to remove user from organization', 'error')

    return redirect(url_for('organization_users', org_id=org_id))

# Tag Routes
@app.route('/organizations/<int:org_id>/tags')
@auth.organization_access_required
def organization_tags(org_id):
    """Organization tags page"""
    user = auth.get_current_user()

    # Check if user has access to this organization
    has_access = user['is_admin']
    if not has_access:
        for org in user['organizations']:
            if org['id'] == org_id:
                has_access = True
                break

    if not has_access:
        flash("You don't have permission to view this page", 'error')
        return redirect(url_for('index'))

    org = db.get_organization(org_id)
    if not org:
        flash('Organization not found', 'error')
        return redirect(url_for('organizations'))

    tags = db.get_organization_tags(org_id)

    return render_template('organization_tags.html',
                          organization=org,
                          tags=tags,
                          user=user)

@app.route('/organizations/<int:org_id>/tags/create', methods=['GET', 'POST'])
@auth.organization_admin_required
def create_tag(org_id):
    """Create a new tag"""
    user = auth.get_current_user()

    # Check if user has access to this organization
    has_access = user['is_admin']
    if not has_access:
        for org in user['organizations']:
            if org['id'] == org_id and org['role'] == 'admin':
                has_access = True
                break

    if not has_access:
        flash("You don't have permission to view this page", 'error')
        return redirect(url_for('index'))

    org = db.get_organization(org_id)
    if not org:
        flash('Organization not found', 'error')
        return redirect(url_for('organizations'))

    if request.method == 'POST':
        name = request.form.get('name')
        color = request.form.get('color', '#6c757d')

        if not name:
            flash('Tag name is required', 'error')
            return redirect(url_for('create_tag', org_id=org_id))

        # Create tag
        tag_id = db.create_tag(org_id, name, color)
        if tag_id:
            flash(f'Tag "{name}" created successfully', 'success')
            return redirect(url_for('organization_tags', org_id=org_id))
        else:
            flash('Failed to create tag', 'error')
            return redirect(url_for('create_tag', org_id=org_id))

    return render_template('tag_form.html',
                          title='Create Tag',
                          organization=org,
                          user=user)

@app.route('/organizations/<int:org_id>/tags/<int:tag_id>/edit', methods=['GET', 'POST'])
@auth.organization_admin_required
def edit_tag(org_id, tag_id):
    """Edit a tag"""
    user = auth.get_current_user()

    # Check if user has access to this organization
    has_access = user['is_admin']
    if not has_access:
        for org in user['organizations']:
            if org['id'] == org_id and org['role'] == 'admin':
                has_access = True
                break

    if not has_access:
        flash("You don't have permission to view this page", 'error')
        return redirect(url_for('index'))

    org = db.get_organization(org_id)
    if not org:
        flash('Organization not found', 'error')
        return redirect(url_for('organizations'))

    tag = db.get_tag(tag_id)
    if not tag or tag['organization_id'] != org_id:
        flash('Tag not found', 'error')
        return redirect(url_for('organization_tags', org_id=org_id))

    if request.method == 'POST':
        name = request.form.get('name')
        color = request.form.get('color', '#6c757d')

        if not name:
            flash('Tag name is required', 'error')
            return redirect(url_for('edit_tag', org_id=org_id, tag_id=tag_id))

        # Update tag
        success = db.update_tag(tag_id, name, color)
        if success:
            flash(f'Tag "{name}" updated successfully', 'success')
            return redirect(url_for('organization_tags', org_id=org_id))
        else:
            flash('Failed to update tag', 'error')
            return redirect(url_for('edit_tag', org_id=org_id, tag_id=tag_id))

    return render_template('tag_form.html',
                          title='Edit Tag',
                          organization=org,
                          tag=tag,
                          user=user)

@app.route('/organizations/<int:org_id>/tags/<int:tag_id>/delete', methods=['POST'])
@auth.organization_admin_required
def delete_tag(org_id, tag_id):
    """Delete a tag"""
    user = auth.get_current_user()

    # Check if user has access to this organization
    has_access = user['is_admin']
    if not has_access:
        for org in user['organizations']:
            if org['id'] == org_id and org['role'] == 'admin':
                has_access = True
                break

    if not has_access:
        flash("You don't have permission to perform this action", 'error')
        return redirect(url_for('index'))

    tag = db.get_tag(tag_id)
    if not tag or tag['organization_id'] != org_id:
        flash('Tag not found', 'error')
        return redirect(url_for('organization_tags', org_id=org_id))

    # Delete tag
    success = db.delete_tag(tag_id)
    if success:
        flash(f'Tag "{tag["name"]}" deleted successfully', 'success')
    else:
        flash('Failed to delete tag', 'error')

    return redirect(url_for('organization_tags', org_id=org_id))

# User Profile Routes
@app.route('/profile', methods=['GET', 'POST'])
@auth.login_required
def profile():
    """User profile page"""
    current_user = auth.get_current_user()

    if request.method == 'POST':
        email = request.form.get('email')
        display_name = request.form.get('display_name', '')

        # Validate input
        if not email:
            flash('Email is required', 'error')
            return render_template('profile.html', user=current_user)

        # Check if email already exists (if changed)
        if email != current_user['email']:
            existing_user = db.get_user_by_email(email)
            if existing_user and existing_user['id'] != current_user['id']:
                flash('Email already in use by another account', 'error')
                return render_template('profile.html', user=current_user)

        # Update user profile
        success = db.update_user_email(
            user_id=current_user['id'],
            email=email
        )

        # Update display name in user preferences
        if success:
            # Store display name in user preferences
            set_user_preference(current_user['id'], 'display_name', display_name)
            flash('Profile updated successfully', 'success')
        else:
            flash('Error updating profile', 'error')

        # Refresh user data
        current_user = auth.get_current_user()

    # Get user preferences
    current_user['display_name'] = get_user_preference(current_user['id'], 'display_name', '')
    current_user['email_alerts'] = get_user_preference(current_user['id'], 'email_alerts', True)

    # Get profile image
    profile_image = db.get_user_profile_image(current_user['id'])
    if profile_image:
        current_user['profile_image'] = profile_image

    return render_template('profile.html', user=current_user)

@app.route('/upload_profile_image', methods=['POST'])
@auth.login_required
def upload_profile_image():
    """Upload profile image for cropping"""
    current_user = auth.get_current_user()

    # Check if the post request has the file part
    if 'profile_image' not in request.files:
        flash('No file part', 'error')
        return redirect(url_for('profile'))

    file = request.files['profile_image']

    # If user does not select file, browser also
    # submit an empty part without filename
    if file.filename == '':
        flash('No selected file', 'error')
        return redirect(url_for('profile'))

    # Check if the file is allowed
    allowed_extensions = {'png', 'jpg', 'jpeg', 'gif'}
    if not '.' in file.filename or file.filename.rsplit('.', 1)[1].lower() not in allowed_extensions:
        flash('Invalid file type. Allowed types: png, jpg, jpeg, gif', 'error')
        return redirect(url_for('profile'))

    try:
        # Generate a unique filename for the temporary image
        import uuid
        temp_filename = f"{uuid.uuid4()}.jpg"

        # Save the uploaded file to the temp directory
        from PIL import Image
        import io
        import os

        # Get the current directory
        current_dir = os.path.dirname(os.path.abspath(__file__))

        # Create the full path to save the temporary image
        temp_path = os.path.join(current_dir, 'static', 'temp', temp_filename)

        # Ensure the directory exists
        os.makedirs(os.path.dirname(temp_path), exist_ok=True)

        # Read and save the image
        img = Image.open(file)

        # Convert to RGB mode (in case it's RGBA or another mode)
        if img.mode != 'RGB':
            img = img.convert('RGB')

        # Resize the image if it's too large to improve performance
        max_size = (1200, 1200)
        if img.width > max_size[0] or img.height > max_size[1]:
            img.thumbnail(max_size, Image.LANCZOS)
            logger.info(f"Resized image to {img.width}x{img.height}")

        # Save the temporary image with high quality
        img.save(temp_path, 'JPEG', quality=95)

        # Log the temporary file path for debugging
        logger.info(f"Temporary image saved at: {temp_path}")
        logger.info(f"Redirecting to crop page with filename: {temp_filename}")

        # Verify the file was saved correctly
        if os.path.exists(temp_path) and os.path.getsize(temp_path) > 0:
            logger.info(f"File saved successfully. Size: {os.path.getsize(temp_path)} bytes")
        else:
            logger.error(f"File not saved correctly or has zero size")
            flash('Error saving uploaded image', 'error')
            return redirect(url_for('profile'))

        # Redirect to the crop page
        return redirect(url_for('crop_profile_image', filename=temp_filename))

    except Exception as e:
        logger.error(f"Error uploading profile image: {str(e)}")
        flash('Error uploading profile image', 'error')
        return redirect(url_for('profile'))

@app.route('/crop_profile_image/<filename>', methods=['GET'])
@auth.login_required
def crop_profile_image(filename):
    """Show the image cropping page"""
    logger.info(f"Rendering crop page for file: {filename}")

    # Check if the temporary file exists
    import os
    base_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static', 'temp')
    temp_path = os.path.normpath(os.path.join(base_path, filename))
    if not temp_path.startswith(base_path):
        logger.error(f"Invalid file path: {temp_path}")
        flash("Error: Invalid file path.", "error")
        return redirect(url_for('profile'))
    if os.path.exists(temp_path):
        logger.info(f"Temporary file exists at: {temp_path}")
    else:
        logger.error(f"Temporary file does not exist at: {temp_path}")
        flash("Error: The uploaded image could not be found.", "error")
        return redirect(url_for('profile'))

    # Add timestamp to prevent browser caching
    timestamp = int(time.time())
    logger.info(f"Adding timestamp {timestamp} to prevent caching")

    return render_template('crop_profile_image.html', temp_filename=filename, now=timestamp)

@app.route('/save_cropped_image', methods=['POST'])
@auth.login_required
def save_cropped_image():
    """Save the cropped profile image"""
    current_user = auth.get_current_user()

    try:
        # Get the crop data and temp filename
        crop_data = request.form.get('crop_data')
        temp_filename = request.form.get('temp_filename')

        if not crop_data or not temp_filename:
            flash('Missing crop data or filename', 'error')
            return redirect(url_for('profile'))

        # Process the cropped image
        import base64
        from PIL import Image
        import io
        import os

        # Get the current directory
        current_dir = os.path.dirname(os.path.abspath(__file__))

        # Create the full path to save the image
        full_path = os.path.join(current_dir, 'static', f"profile_images/user_{current_user['id']}.jpg")

        # Ensure the directory exists
        os.makedirs(os.path.dirname(full_path), exist_ok=True)

        # Convert base64 to image
        # Remove the data URL prefix (e.g., "data:image/jpeg;base64,")
        crop_data = crop_data.split(',')[1]

        # Decode base64 data
        image_data = base64.b64decode(crop_data)

        # Create image from binary data
        img = Image.open(io.BytesIO(image_data))

        # Save the image
        img.save(full_path, 'JPEG')

        # Update the user's profile image in the database
        filename = f"profile_images/user_{current_user['id']}.jpg"
        db.update_user_profile_image(current_user['id'], filename)

        # Delete the temporary file
        temp_path = os.path.join(current_dir, 'static', 'temp', temp_filename)
        if os.path.exists(temp_path):
            os.remove(temp_path)

        flash('Profile image updated successfully', 'success')
    except Exception as e:
        logger.error(f"Error saving cropped image: {str(e)}")
        flash('Error saving cropped image', 'error')

    return redirect(url_for('profile'))

@app.route('/change_password', methods=['POST'])
@auth.login_required
def change_password():
    """Change user password"""
    current_user = auth.get_current_user()

    current_password = request.form.get('current_password')
    new_password = request.form.get('new_password')
    confirm_password = request.form.get('confirm_password')

    # Validate input
    if not current_password or not new_password or not confirm_password:
        flash('All fields are required', 'error')
        return redirect(url_for('profile'))

    if new_password != confirm_password:
        flash('New passwords do not match', 'error')
        return redirect(url_for('profile'))

    # Verify current password
    password_parts = current_user['password_hash'].split(':')
    if len(password_parts) != 2:
        flash('Invalid account configuration. Please contact an administrator.', 'error')
        return redirect(url_for('profile'))

    stored_hash, salt = password_parts
    if not auth.verify_password(current_password, stored_hash, salt):
        flash('Current password is incorrect', 'error')
        return redirect(url_for('profile'))

    # Hash new password
    password_hash, salt = auth.hash_password(new_password)
    combined_hash = f"{password_hash}:{salt}"

    # Update password
    success = db.update_user_password(
        user_id=current_user['id'],
        password_hash=combined_hash
    )

    if success:
        # Delete all other sessions for this user
        db.delete_user_sessions(current_user['id'])

        # Create a new session for the current user
        session_token = auth.create_user_session(current_user['id'])

        # Set session cookie
        response = make_response(redirect(url_for('profile')))
        response.set_cookie(
            auth.SESSION_COOKIE_NAME,
            session_token,
            max_age=auth.SESSION_EXPIRY,
            httponly=True,
            secure=request.is_secure,
            samesite='Lax'
        )

        flash('Password changed successfully. All other sessions have been logged out.', 'success')
        return response
    else:
        flash('Error changing password', 'error')
        return redirect(url_for('profile'))

@app.route('/update_preferences', methods=['POST'])
@auth.login_required
def update_preferences():
    """Update user preferences"""
    current_user = auth.get_current_user()

    # Only handle email alerts now
    email_alerts = 'email_alerts' in request.form

    # Update preferences
    set_user_preference(current_user['id'], 'email_alerts', email_alerts)

    # Set a flash message
    flash('Preferences updated successfully', 'success')

    # Redirect back to profile page
    return redirect(url_for('profile'))

# User preferences are now handled by database.py

# Logs Routes
@app.route('/logs')
@auth.login_required
@auth.admin_required
def logs():
    """User action logs page with modern design and improved filtering"""
    user = auth.get_current_user()

    # Get query parameters for filtering
    page = int(request.args.get('page', 1))
    per_page = 25  # Reduced to show more pages for better pagination experience
    username = request.args.get('username', '')
    action_type = request.args.get('action_type', '')
    resource_type = request.args.get('resource_type', '')
    date_range = request.args.get('date_range', 'all')
    start_date_str = request.args.get('start_date', '')
    end_date_str = request.args.get('end_date', '')

    # Import pytz at the top level
    import pytz
    from datetime import timezone

    # Get user's timezone
    app_settings = get_app_settings()
    timezone_str = app_settings.timezone

    # Get the user's timezone
    try:
        user_tz = pytz.timezone(timezone_str)
    except pytz.exceptions.UnknownTimeZoneError:
        logger.warning(f"Unknown timezone {timezone_str}, using UTC")
        user_tz = pytz.UTC

    # Parse dates based on date_range or custom range
    start_date = None
    end_date = None

    # Handle predefined date ranges
    now = datetime.now(tz=user_tz)
    if date_range == 'today':
        # Today (in user's timezone)
        start_date = datetime(now.year, now.month, now.day, 0, 0, 0, tzinfo=user_tz)
        end_date = datetime(now.year, now.month, now.day, 23, 59, 59, tzinfo=user_tz)
    elif date_range == 'yesterday':
        # Yesterday (in user's timezone)
        yesterday = now - timedelta(days=1)
        start_date = datetime(yesterday.year, yesterday.month, yesterday.day, 0, 0, 0, tzinfo=user_tz)
        end_date = datetime(yesterday.year, yesterday.month, yesterday.day, 23, 59, 59, tzinfo=user_tz)
    elif date_range == 'last7days':
        # Last 7 days (in user's timezone)
        start_date = now - timedelta(days=7)
        end_date = now
    elif date_range == 'last30days':
        # Last 30 days (in user's timezone)
        start_date = now - timedelta(days=30)
        end_date = now
    elif date_range == 'thismonth':
        # This month (in user's timezone)
        start_date = datetime(now.year, now.month, 1, 0, 0, 0, tzinfo=user_tz)
        end_date = now
    elif date_range == 'lastmonth':
        # Last month (in user's timezone)
        if now.month == 1:
            last_month = datetime(now.year - 1, 12, 1, 0, 0, 0, tzinfo=user_tz)
        else:
            last_month = datetime(now.year, now.month - 1, 1, 0, 0, 0, tzinfo=user_tz)
        start_date = last_month
        # Last day of last month
        if now.month == 1:
            end_date = datetime(now.year - 1, 12, 31, 23, 59, 59, tzinfo=user_tz)
        else:
            # Get the last day of the previous month
            last_day = (datetime(now.year, now.month, 1) - timedelta(days=1)).day
            end_date = datetime(now.year, now.month - 1, last_day, 23, 59, 59, tzinfo=user_tz)
    elif date_range == 'custom' and (start_date_str or end_date_str):
        # Custom date range
        if start_date_str:
            try:
                # Handle different date formats
                if 'T' in start_date_str:
                    # HTML datetime-local format: YYYY-MM-DDThh:mm
                    local_dt = datetime.strptime(start_date_str, '%Y-%m-%dT%H:%M')
                else:
                    # Try ISO format
                    local_dt = datetime.fromisoformat(start_date_str)

                # Localize the datetime to user's timezone
                start_date = user_tz.localize(local_dt)
                logger.info(f"Parsed custom start date: {start_date}")
            except ValueError as e:
                logger.error(f"Error parsing start date: {e}")
                flash('Invalid start date format', 'error')

        if end_date_str:
            try:
                # Handle different date formats
                if 'T' in end_date_str:
                    # HTML datetime-local format: YYYY-MM-DDThh:mm
                    local_dt = datetime.strptime(end_date_str, '%Y-%m-%dT%H:%M')
                else:
                    # Try ISO format
                    local_dt = datetime.fromisoformat(end_date_str)

                # Set end date to end of day if no time specified
                if local_dt.hour == 0 and local_dt.minute == 0 and local_dt.second == 0:
                    local_dt = local_dt.replace(hour=23, minute=59, second=59)

                # Localize the datetime to user's timezone
                end_date = user_tz.localize(local_dt)
                logger.info(f"Parsed custom end date: {end_date}")
            except ValueError as e:
                logger.error(f"Error parsing end date: {e}")
                flash('Invalid end date format', 'error')

    # Convert timezone-aware datetimes to UTC for database filtering
    if start_date:
        start_date = start_date.astimezone(pytz.UTC).replace(tzinfo=None)
        logger.info(f"Using start date (UTC): {start_date}")

    if end_date:
        end_date = end_date.astimezone(pytz.UTC).replace(tzinfo=None)
        logger.info(f"Using end date (UTC): {end_date}")

    # Get user ID if username filter is provided
    user_id = None
    if username:
        user_data = db.get_user_by_username(username)
        if user_data:
            user_id = user_data['id']
            logger.info(f"Filtering by user ID: {user_id} (username: {username})")
        else:
            logger.warning(f"Username not found: {username}")

    # For admin users, show all logs regardless of organization
    # For regular users, only show logs for their current organization
    organization_id = None
    if not user.get('is_admin'):
        organization_id = user.get('current_organization', {}).get('id')
        logger.info(f"Non-admin user, filtering by organization ID: {organization_id}")

    # Get logs with filters
    logs = db.get_user_action_logs(
        limit=per_page,
        offset=(page - 1) * per_page,
        user_id=user_id,
        action_type=action_type if action_type else None,
        resource_type=resource_type if resource_type else None,
        organization_id=organization_id,
        start_date=start_date,
        end_date=end_date
    )

    # Get total count for pagination
    total_logs = db.get_user_action_logs_count(
        user_id=user_id,
        action_type=action_type if action_type else None,
        resource_type=resource_type if resource_type else None,
        organization_id=organization_id,
        start_date=start_date,
        end_date=end_date
    )

    logger.info(f"Found {total_logs} logs matching filters")

    # Calculate total pages
    total_pages = (total_logs + per_page - 1) // per_page if total_logs > 0 else 1

    # Get unique action types and resource types for filter dropdowns
    action_types = db.get_unique_action_types()
    resource_types = db.get_unique_resource_types()

    # Get all users for username filter dropdown
    all_users = db.get_all_users()

    return render_template('logs.html',
                          logs=logs,
                          total_logs=total_logs,
                          page=page,
                          total_pages=total_pages,
                          user=user,
                          filter_username=username,
                          filter_action_type=action_type,
                          filter_resource_type=resource_type,
                          filter_date_range=date_range,
                          filter_start_date=start_date_str,
                          filter_end_date=end_date_str,
                          action_types=action_types,
                          resource_types=resource_types,
                          all_users=all_users,
                          timezone=timezone_str)

@app.route('/export_logs')
@auth.login_required
@auth.admin_required
def export_logs():
    """Export logs as CSV"""
    user = auth.get_current_user()

    # Get query parameters for filtering
    username = request.args.get('username')
    action_type = request.args.get('action_type')
    resource_type = request.args.get('resource_type')
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')

    # Parse dates if provided
    start_date = None
    end_date = None

    # Import pytz at the top level
    import pytz
    from datetime import timezone

    # Get user's timezone
    app_settings = get_app_settings()
    timezone_str = app_settings.timezone

    # Get the user's timezone
    try:
        user_tz = pytz.timezone(timezone_str)
    except pytz.exceptions.UnknownTimeZoneError:
        logger.warning(f"Unknown timezone {timezone_str}, using UTC")
        user_tz = pytz.UTC

    if start_date_str:
        try:
            logger.info(f"Parsing start date for export: {start_date_str}")
            # Handle different date formats
            if 'T' in start_date_str:
                # HTML datetime-local format: YYYY-MM-DDThh:mm
                local_dt = datetime.strptime(start_date_str, '%Y-%m-%dT%H:%M')
            else:
                # Try ISO format
                local_dt = datetime.fromisoformat(start_date_str)

            # Localize the datetime to user's timezone
            local_dt = user_tz.localize(local_dt)
            # Convert to UTC for database filtering
            start_date = local_dt.astimezone(pytz.UTC).replace(tzinfo=None)

            logger.info(f"Parsed start date for export: {start_date} (original: {local_dt})")
        except ValueError as e:
            logger.error(f"Error parsing start date for export: {e}")
            flash('Invalid start date format', 'error')
            return redirect(url_for('logs'))

    if end_date_str:
        try:
            logger.info(f"Parsing end date for export: {end_date_str}")
            # Handle different date formats
            if 'T' in end_date_str:
                # HTML datetime-local format: YYYY-MM-DDThh:mm
                local_dt = datetime.strptime(end_date_str, '%Y-%m-%dT%H:%M')
            else:
                # Try ISO format
                local_dt = datetime.fromisoformat(end_date_str)

            # Set end date to end of day
            local_dt = local_dt.replace(hour=23, minute=59, second=59)

            # Localize the datetime to user's timezone
            local_dt = user_tz.localize(local_dt)
            # Convert to UTC for database filtering
            end_date = local_dt.astimezone(pytz.UTC).replace(tzinfo=None)

            logger.info(f"Parsed end date for export: {end_date} (original: {local_dt})")
        except ValueError as e:
            logger.error(f"Error parsing end date for export: {e}")
            flash('Invalid end date format', 'error')
            return redirect(url_for('logs'))

    # Get user ID if username filter is provided
    user_id = None
    if username:
        user_data = db.get_user_by_username(username)
        if user_data:
            user_id = user_data['id']

    # For admin users, show all logs regardless of organization
    # For regular users, only show logs for their current organization
    organization_id = None
    if not user.get('is_admin'):
        organization_id = user.get('current_organization', {}).get('id')

    # Log the filter parameters for debugging
    logger.info(f"Exporting logs with filters: user_id={user_id}, action_type={action_type}, "
                f"resource_type={resource_type}, organization_id={organization_id}, "
                f"start_date={start_date}, end_date={end_date}")

    # Get all logs with filters (no pagination)
    logs = db.get_user_action_logs(
        limit=10000,  # Set a high limit to get all logs
        offset=0,
        user_id=user_id,
        action_type=action_type,
        resource_type=resource_type,
        organization_id=organization_id,
        start_date=start_date,
        end_date=end_date
    )

    logger.info(f"Exporting {len(logs)} logs")

    # Create CSV file in memory
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow(['Time', 'User', 'Action', 'Resource Type', 'Resource Name', 'Details', 'IP Address'])

    # Write data
    for log in logs:
        # Format the timestamp with the user's timezone
        timestamp = ''
        if log['created_at']:
            try:
                # Use the datetime filter to format the timestamp with the user's timezone
                timestamp = format_datetime(log['created_at'], '%Y-%m-%d %H:%M:%S')
            except Exception as e:
                logger.error(f"Error formatting timestamp for CSV: {e}")
                # Fallback to UTC
                timestamp = datetime.fromtimestamp(log['created_at']).strftime('%Y-%m-%d %H:%M:%S')

        writer.writerow([
            timestamp,
            log['username'],
            log['action_type'],
            log['resource_type'],
            log['resource_name'] or '',
            log['details'] or '',
            log['ip_address'] or ''
        ])

    # Prepare response
    output.seek(0)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return Response(
        output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': f'attachment;filename=certifly_logs_{timestamp}.csv'}
    )

# User Administration Routes
@app.route('/user_admin')
@auth.login_required
@auth.admin_required
def user_admin():
    """User administration page"""
    users = db.get_all_users()

    # Get organizations for each user
    for user in users:
        user['organizations'] = db.get_user_organizations(user['id'])

    # Sort users alphabetically by username by default
    sort_by = request.args.get('sort', 'username')
    sort_order = request.args.get('order', 'asc')

    # Define sorting functions for different fields
    sort_functions = {
        'username': lambda x: x['username'].lower(),
        'email': lambda x: x['email'].lower(),
        'is_admin': lambda x: (0 if x['is_admin'] else 1),  # Admins first
        'is_active': lambda x: (0 if x['is_active'] else 1),  # Active users first
        'created_at': lambda x: x['created_at'] if x['created_at'] else 0
    }

    # Apply sorting
    if sort_by in sort_functions:
        users = sorted(users, key=sort_functions[sort_by], reverse=(sort_order == 'desc'))

    organizations = db.get_all_organizations()
    return render_template('user_admin.html',
                          users=users,
                          organizations=organizations,
                          sort_by=sort_by,
                          sort_order=sort_order)

@app.route('/bulk_activate_users', methods=['POST'])
@auth.login_required
@auth.admin_required
def bulk_activate_users():
    """Activate multiple users"""
    # Check if user is admin
    current_user = auth.get_current_user()
    if not current_user.get('is_admin'):
        flash('You do not have permission to perform this action', 'error')
        return redirect(url_for('user_admin'))

    user_ids = request.form.getlist('user_ids')

    if not user_ids:
        flash('No users selected', 'error')
        return redirect(url_for('user_admin'))

    # Activate each user
    success_count = 0
    error_count = 0

    for user_id in user_ids:
        try:
            # Get user
            user = db.get_user_by_id(user_id)
            if not user:
                error_count += 1
                continue

            # Skip if user is already active
            if user.get('is_active'):
                continue

            # Activate user
            db.update_user_status(user_id, True)

            # Log the user activation action
            db.log_user_action(
                user_id=current_user['id'],
                username=current_user['username'],
                action_type='update',
                resource_type='user',
                resource_id=user_id,
                resource_name=user['username'],
                details=f'Admin activated user {user["username"]}',
                ip_address=request.remote_addr
            )

            success_count += 1
        except Exception as e:
            logger.error(f"Error activating user {user_id}: {str(e)}")
            error_count += 1

    if success_count > 0:
        flash(f'Successfully activated {success_count} user(s)', 'success')

    if error_count > 0:
        flash(f'Failed to activate {error_count} user(s)', 'error')

    return redirect(url_for('user_admin'))

@app.route('/bulk_deactivate_users', methods=['POST'])
@auth.login_required
@auth.admin_required
def bulk_deactivate_users():
    """Deactivate multiple users"""
    # Check if user is admin
    current_user = auth.get_current_user()
    if not current_user.get('is_admin'):
        flash('You do not have permission to perform this action', 'error')
        return redirect(url_for('user_admin'))

    user_ids = request.form.getlist('user_ids')

    if not user_ids:
        flash('No users selected', 'error')
        return redirect(url_for('user_admin'))

    # Deactivate each user
    success_count = 0
    error_count = 0

    for user_id in user_ids:
        try:
            # Get user
            user = db.get_user_by_id(user_id)
            if not user:
                error_count += 1
                continue

            # Skip if user is already inactive
            if not user.get('is_active'):
                continue

            # Skip if user is current user
            if str(user_id) == str(current_user.get('id')):
                flash('You cannot deactivate your own account', 'warning')
                continue

            # Deactivate user
            db.update_user_status(user_id, False)

            # Log the user deactivation action
            db.log_user_action(
                user_id=current_user['id'],
                username=current_user['username'],
                action_type='update',
                resource_type='user',
                resource_id=user_id,
                resource_name=user['username'],
                details=f'Admin deactivated user {user["username"]}',
                ip_address=request.remote_addr
            )

            success_count += 1
        except Exception as e:
            logger.error(f"Error deactivating user {user_id}: {str(e)}")
            error_count += 1

    if success_count > 0:
        flash(f'Successfully deactivated {success_count} user(s)', 'success')

    if error_count > 0:
        flash(f'Failed to deactivate {error_count} user(s)', 'error')

    return redirect(url_for('user_admin'))

@app.route('/bulk_delete_users', methods=['POST'])
@auth.login_required
@auth.admin_required
def bulk_delete_users():
    """Delete multiple users"""
    # Check if user is admin
    current_user = auth.get_current_user()
    if not current_user.get('is_admin'):
        flash('You do not have permission to perform this action', 'error')
        return redirect(url_for('user_admin'))

    user_ids = request.form.getlist('user_ids')

    if not user_ids:
        flash('No users selected', 'error')
        return redirect(url_for('user_admin'))

    # Delete each user
    success_count = 0
    error_count = 0

    for user_id in user_ids:
        try:
            # Get user
            user = db.get_user_by_id(user_id)
            if not user:
                error_count += 1
                continue

            # Skip if user is current user
            if str(user_id) == str(current_user.get('id')):
                flash('You cannot delete your own account', 'warning')
                continue

            # Delete user
            db.delete_user(user_id)

            # Log the user deletion action
            db.log_user_action(
                user_id=current_user['id'],
                username=current_user['username'],
                action_type='delete',
                resource_type='user',
                resource_id=user_id,
                resource_name=user['username'],
                details=f'Admin bulk deleted user {user["username"]} with email {user["email"]}',
                ip_address=request.remote_addr
            )

            success_count += 1
        except Exception as e:
            logger.error(f"Error deleting user {user_id}: {str(e)}")
            error_count += 1

    if success_count > 0:
        flash(f'Successfully deleted {success_count} user(s)', 'success')

    if error_count > 0:
        flash(f'Failed to delete {error_count} user(s)', 'error')

    return redirect(url_for('user_admin'))

@app.route('/user_admin/add', methods=['POST'])
@auth.login_required
@auth.admin_required
def user_admin_add():
    """Add a new user"""
    username = request.form.get('username')
    email = request.form.get('email')
    password = request.form.get('password')
    confirm_password = request.form.get('confirm_password')
    is_admin = request.form.get('is_admin') == '1'

    # Validate input
    if not username or not email or not password or not confirm_password:
        flash('Please fill in all fields', 'error')
        return redirect(url_for('user_admin'))

    if password != confirm_password:
        flash('Passwords do not match', 'error')
        return redirect(url_for('user_admin'))

    # Check if username or email already exists
    if db.get_user_by_username(username):
        flash('Username already exists', 'error')
        return redirect(url_for('user_admin'))

    if db.get_user_by_email(email):
        flash('Email already exists', 'error')
        return redirect(url_for('user_admin'))

    # Hash password
    password_hash, salt = auth.hash_password(password)
    combined_hash = f"{password_hash}:{salt}"

    # Create user
    user_id = db.create_user(username, email, combined_hash, is_admin=is_admin)
    if not user_id:
        flash('Error creating user', 'error')
        return redirect(url_for('user_admin'))

    # Get current admin user
    current_user = auth.get_current_user()

    # Log the user creation action
    db.log_user_action(
        user_id=current_user['id'],
        username=current_user['username'],
        action_type='create',
        resource_type='user',
        resource_id=user_id,
        resource_name=username,
        details=f'Admin created user with email {email}, admin privileges: {is_admin}',
        ip_address=request.remote_addr
    )

    flash(f'User {username} created successfully', 'success')
    return redirect(url_for('user_admin'))

@app.route('/user_admin/edit', methods=['POST'])
@auth.login_required
@auth.admin_required
def user_admin_edit():
    """Edit a user"""
    user_id = request.form.get('user_id')
    username = request.form.get('username')
    email = request.form.get('email')
    is_admin = request.form.get('is_admin') == '1'
    is_active = request.form.get('is_active') == '1'

    # Validate input
    if not user_id or not username or not email:
        flash('Please fill in all fields', 'error')
        return redirect(url_for('user_admin'))

    # Get the user
    user = db.get_user_by_id(user_id)
    if not user:
        flash('User not found', 'error')
        return redirect(url_for('user_admin'))

    # Check if username or email already exists (if changed)
    if username != user['username'] and db.get_user_by_username(username):
        flash('Username already exists', 'error')
        return redirect(url_for('user_admin'))

    if email != user['email'] and db.get_user_by_email(email):
        flash('Email already exists', 'error')
        return redirect(url_for('user_admin'))

    # Update user
    success = db.update_user(
        user_id=user_id,
        username=username,
        email=email,
        is_admin=is_admin,
        is_active=is_active
    )

    if not success:
        flash('Error updating user', 'error')
        return redirect(url_for('user_admin'))

    # Get current admin user
    current_user = auth.get_current_user()

    # Log the user update action
    changes = []
    if username != user['username']:
        changes.append(f"username from {user['username']} to {username}")
    if email != user['email']:
        changes.append(f"email from {user['email']} to {email}")
    if is_admin != user['is_admin']:
        changes.append(f"admin status from {user['is_admin']} to {is_admin}")
    if is_active != user['is_active']:
        changes.append(f"active status from {user['is_active']} to {is_active}")

    change_details = ", ".join(changes) if changes else "no changes"

    db.log_user_action(
        user_id=current_user['id'],
        username=current_user['username'],
        action_type='update',
        resource_type='user',
        resource_id=user_id,
        resource_name=username,
        details=f'Admin updated user: {change_details}',
        ip_address=request.remote_addr
    )

    flash(f'User {username} updated successfully', 'success')
    return redirect(url_for('user_admin'))

@app.route('/user_admin/reset_password', methods=['POST'])
@auth.login_required
@auth.admin_required
def user_admin_reset_password():
    """Reset a user's password"""
    user_id = request.form.get('user_id')
    new_password = request.form.get('new_password')
    confirm_new_password = request.form.get('confirm_new_password')

    # Validate input
    if not user_id or not new_password or not confirm_new_password:
        flash('Please fill in all fields', 'error')
        return redirect(url_for('user_admin'))

    if new_password != confirm_new_password:
        flash('Passwords do not match', 'error')
        return redirect(url_for('user_admin'))

    # Get the user
    user = db.get_user_by_id(user_id)
    if not user:
        flash('User not found', 'error')
        return redirect(url_for('user_admin'))

    # Hash password
    password_hash, salt = auth.hash_password(new_password)
    combined_hash = f"{password_hash}:{salt}"

    # Update user
    success = db.update_user(
        user_id=user_id,
        password_hash=combined_hash
    )

    if not success:
        flash('Error resetting password', 'error')
        return redirect(url_for('user_admin'))

    # Invalidate all sessions for this user
    db.delete_user_sessions(user_id)

    # Get current admin user
    current_user = auth.get_current_user()

    # Log the password reset action
    db.log_user_action(
        user_id=current_user['id'],
        username=current_user['username'],
        action_type='update',
        resource_type='user',
        resource_id=user_id,
        resource_name=user["username"],
        details=f'Admin reset password for user {user["username"]}',
        ip_address=request.remote_addr
    )

    flash(f'Password for {user["username"]} reset successfully', 'success')
    return redirect(url_for('user_admin'))

@app.route('/user_admin/delete', methods=['POST'])
@auth.login_required
@auth.admin_required
def user_admin_delete():
    """Delete a user"""
    user_id = request.form.get('user_id')

    # Validate input
    if not user_id:
        flash('Invalid request', 'error')
        return redirect(url_for('user_admin'))

    # Get the user
    user = db.get_user_by_id(user_id)
    if not user:
        flash('User not found', 'error')
        return redirect(url_for('user_admin'))

    # Check if user is trying to delete themselves
    current_user = auth.get_current_user()
    if str(current_user['id']) == str(user_id):
        flash('You cannot delete your own account', 'error')
        return redirect(url_for('user_admin'))

    # Delete user
    success = db.delete_user(user_id)
    if not success:
        flash('Error deleting user', 'error')
        return redirect(url_for('user_admin'))

    # Log the user deletion action
    db.log_user_action(
        user_id=current_user['id'],
        username=current_user['username'],
        action_type='delete',
        resource_type='user',
        resource_id=user_id,
        resource_name=user["username"],
        details=f'Admin deleted user {user["username"]} with email {user["email"]}',
        ip_address=request.remote_addr
    )

    flash(f'User {user["username"]} deleted successfully', 'success')
    return redirect(url_for('user_admin'))

# API Routes for User Organization Management
@app.route('/api/users/<int:user_id>/organizations')
@auth.login_required
def api_get_user_organizations(user_id):
    """API endpoint to get a user's organizations"""

    # Check if user has permission to view this user's organizations
    current_user = auth.get_current_user()

    # System admins can view any user's organizations
    has_access = current_user['is_admin']

    # Users can view their own organizations
    if not has_access and current_user['id'] == user_id:
        has_access = True

    if not has_access:
        return jsonify({'success': False, 'message': 'You do not have permission to view this user\'s organizations'}), 403
    # Get the user
    user = db.get_user_by_id(user_id)
    if not user:
        return jsonify({'success': False, 'message': 'User not found'}), 404

    # Get user's organizations
    organizations = db.get_user_organizations(user_id)

    return jsonify({
        'success': True,
        'organizations': organizations
    })

@app.route('/api/users/add_to_organization', methods=['POST'])
@auth.login_required
def api_add_user_to_organization():
    """API endpoint to add a user to an organization"""
    data = request.json
    user_id = data.get('user_id')
    org_id = data.get('org_id')
    role = data.get('role', 'member')

    # Check if user has permission to add users to this organization
    current_user = auth.get_current_user()
    has_access = current_user['is_admin']

    if not has_access and org_id:
        # Check if user is an admin of this organization
        for org in current_user['organizations']:
            if org['id'] == int(org_id) and org['role'] == 'admin':
                has_access = True
                break

    if not has_access:
        return jsonify({'success': False, 'message': 'You do not have permission to add users to this organization'}), 403

    # Validate input
    if not user_id or not org_id:
        return jsonify({'success': False, 'message': 'Missing required parameters'}), 400

    # Check if user exists
    user = db.get_user_by_id(user_id)
    if not user:
        return jsonify({'success': False, 'message': 'User not found'}), 404

    # Check if organization exists
    org = db.get_organization(org_id)
    if not org:
        return jsonify({'success': False, 'message': 'Organization not found'}), 404

    # Check if user is already in organization
    if db.is_user_in_organization(user_id, org_id):
        return jsonify({'success': False, 'message': 'User is already a member of this organization'}), 400

    # Add user to organization
    result = db.add_user_to_organization(user_id, org_id, role)
    if not result:
        return jsonify({'success': False, 'message': 'Error adding user to organization'}), 500

    return jsonify({
        'success': True,
        'message': f'User added to {org["name"]} as {role}'
    })

@app.route('/api/users/update_organization_role', methods=['POST'])
@auth.login_required
def api_update_user_organization_role():
    """API endpoint to update a user's role in an organization"""

    # Check if user has permission to update roles in this organization
    data = request.json
    org_id = data.get('org_id')

    current_user = auth.get_current_user()
    has_access = current_user['is_admin']

    if not has_access and org_id:
        # Check if user is an admin of this organization
        for org in current_user['organizations']:
            if org['id'] == int(org_id) and org['role'] == 'admin':
                has_access = True
                break

    if not has_access:
        return jsonify({'success': False, 'message': 'You do not have permission to update roles in this organization'}), 403

    user_id = data.get('user_id')
    # org_id is already retrieved above
    role = data.get('role')

    # Validate input
    if not user_id or not org_id or not role:
        return jsonify({'success': False, 'message': 'Missing required parameters'}), 400

    # Check if user exists
    user = db.get_user_by_id(user_id)
    if not user:
        return jsonify({'success': False, 'message': 'User not found'}), 404

    # Check if organization exists
    org = db.get_organization(org_id)
    if not org:
        return jsonify({'success': False, 'message': 'Organization not found'}), 404

    # Check if user is in organization
    if not db.is_user_in_organization(user_id, org_id):
        return jsonify({'success': False, 'message': 'User is not a member of this organization'}), 400

    # Update user's role
    success = db.update_user_organization_role(user_id, org_id, role)
    if not success:
        return jsonify({'success': False, 'message': 'Error updating user role'}), 500

    return jsonify({
        'success': True,
        'message': f'User role updated to {role} in {org["name"]}'
    })

@app.route('/api/users/remove_from_organization', methods=['POST'])
@auth.login_required
def api_remove_user_from_organization():
    """API endpoint to remove a user from an organization"""

    # Check if user has permission to remove users from this organization
    data = request.json
    org_id = data.get('org_id')

    current_user = auth.get_current_user()
    has_access = current_user['is_admin']

    if not has_access and org_id:
        # Check if user is an admin of this organization
        for org in current_user['organizations']:
            if org['id'] == int(org_id) and org['role'] == 'admin':
                has_access = True
                break

    if not has_access:
        return jsonify({'success': False, 'message': 'You do not have permission to remove users from this organization'}), 403

    user_id = data.get('user_id')
    # org_id is already retrieved above

    # Validate input
    if not user_id or not org_id:
        return jsonify({'success': False, 'message': 'Missing required parameters'}), 400

    # Check if user exists
    user = db.get_user_by_id(user_id)
    if not user:
        return jsonify({'success': False, 'message': 'User not found'}), 404

    # Check if organization exists
    org = db.get_organization(org_id)
    if not org:
        return jsonify({'success': False, 'message': 'Organization not found'}), 404

    # Check if user is in organization
    if not db.is_user_in_organization(user_id, org_id):
        return jsonify({'success': False, 'message': 'User is not a member of this organization'}), 400

    # Remove user from organization
    success = db.remove_user_from_organization(user_id, org_id)
    if not success:
        return jsonify({'success': False, 'message': 'Error removing user from organization'}), 500

    return jsonify({
        'success': True,
        'message': f'User removed from {org["name"]}'
    })

@app.route('/api/organizations/<int:org_id>/domains', methods=['GET'])
@auth.login_required
def api_get_organization_domains(org_id):
    """API endpoint to get domains for an organization"""
    try:
        # Get current user
        current_user = auth.get_current_user()

        # Check if user has access to this organization
        has_access = current_user['is_admin']
        if not has_access:
            for org in current_user['organizations']:
                if org['id'] == org_id:
                    has_access = True
                    break

        if not has_access:
            return jsonify({'success': False, 'message': 'You do not have permission to view domains for this organization'}), 403

        # Get domains for the organization
        domains = db.get_domains_by_organization(org_id)

        # Format the response
        formatted_domains = []
        for domain in domains:
            formatted_domains.append({
                'id': domain['id'],
                'name': domain['name'],
                'ssl_monitored': domain['ssl_monitored'],
                'expiry_monitored': domain['expiry_monitored'],
                'ping_monitored': domain['ping_monitored']
            })

        return jsonify({
            'success': True,
            'domains': formatted_domains
        })
    except Exception as e:
        logger.error(f"Error getting domains for organization {org_id}: {str(e)}")
        return jsonify({'success': False, 'message': f'Server error: {str(e)}'}), 500

# Background task for periodic ping checks - disabled in favor of manual refresh
def run_periodic_ping_checks():
    """This function is now disabled as auto background checks have been removed"""
    logger.info("Periodic ping checks are disabled")

    # Just sleep indefinitely to keep the thread alive but not doing anything
    while True:
        time.sleep(3600)  # Sleep for an hour

# Function to add test ping data for a domain
def add_test_ping_data(domain_name, num_entries=24):
    """Add test ping data for a domain to populate the response time chart"""
    logger.info(f"Adding {num_entries} test ping entries for {domain_name}")

    # Get the domain
    domain = db.get_domain_by_name(domain_name)
    if not domain:
        logger.error(f"Domain {domain_name} not found!")
        return False

    # Generate test data
    success_count = 0

    for i in range(num_entries):
        # Generate a random response time between 10ms and 200ms
        import random
        response_time = random.randint(10, 200)

        # Most entries should be 'up', but add some 'down' entries randomly
        status = 'up' if random.random() > 0.1 else 'down'

        # Use the timestamp for the database record
        # The record_ping_status function will create the database entry with this timestamp

        # Record the ping status
        success = db.record_ping_status(domain_name, status, response_time)
        if success:
            success_count += 1

    logger.info(f"Successfully added {success_count} ping entries for {domain_name}")
    return True

# Start the background task in a separate thread
def start_background_tasks():
    """Start all background tasks in separate threads"""
    # Start the ping check thread (currently disabled)
    ping_thread = threading.Thread(target=run_periodic_ping_checks, daemon=True)
    ping_thread.start()

    # Start the notification checker thread
    def notification_checker():
        """Background thread to periodically check for and send notifications"""
        # Wait 60 seconds after startup before first check
        time.sleep(60)

        # Track when the last daily summary was sent
        last_summary_date = None

        while True:
            try:
                # Check for pending notifications
                check_and_send_pending_notifications()

                # Check if we need to send a daily summary
                current_datetime = datetime.now()
                current_date = current_datetime.date()
                current_hour = current_datetime.hour

                # Send the summary at 8:00 AM (between 8:00 and 8:59)
                summary_hour = 8

                # Only send if it's a new day and we're in the right hour
                if (last_summary_date is None or current_date > last_summary_date) and current_hour == summary_hour:
                    try:
                        logger.info(f"Running daily summary check at {current_datetime.strftime('%Y-%m-%d %H:%M:%S')}...")
                        success, message = check_all_domains_and_send_summary()
                        if success:
                            last_summary_date = current_date
                            logger.info(f"Daily summary check completed: {message}")
                        else:
                            logger.error(f"Daily summary check failed: {message}")
                    except Exception as e:
                        logger.error(f"Error in daily summary check: {str(e)}", exc_info=True)

                # Sleep for 1 hour before checking again
                time.sleep(3600)  # 3600 seconds = 1 hour
            except Exception as e:
                logger.error(f"Error in notification checker thread: {str(e)}", exc_info=True)
                # Sleep for 5 minutes before trying again after an error
                time.sleep(300)

    notification_thread = threading.Thread(target=notification_checker, daemon=True)
    notification_thread.start()

    logger.info("Background tasks started")

if __name__ == '__main__':
    # Create database tables if they don't exist
    db.create_tables()

    # Check if we need to create a default admin user
    if not db.get_users():
        # Create default admin user
        default_username = os.getenv('DEFAULT_ADMIN_USERNAME', 'admin')
        default_email = os.getenv('DEFAULT_ADMIN_EMAIL', 'admin@example.com')
        default_password = os.getenv('DEFAULT_ADMIN_PASSWORD', 'admin')

        db.create_user(default_username, default_email, default_password, is_admin=True)
        logger.info(f"Created default admin user: {default_username}")

    # Start background tasks
    start_background_tasks()

    # Start the application
    port = int(os.getenv('PORT', 5000))
    host = os.getenv('HOST', '0.0.0.0')

    if os.getenv('FLASK_ENV') == 'production':
        # Production settings - never use debug mode
        try:
            from waitress import serve
            logger.info(f"Starting production server with Waitress on {host}:{port}")
            serve(app, host=host, port=port)
        except ImportError:
            logger.info(f"Waitress not installed, using Flask production server on {host}:{port}")
            app.run(debug=False, host=host, port=port)
    else:
        # Development settings - only use debug mode in a controlled local environment
        # Debug mode should NEVER be enabled on a publicly accessible server
        debug_mode = os.getenv('FLASK_DEBUG', 'false').lower() == 'true'

        if debug_mode:
            logger.warning("Running with debug=True. This should NEVER be used in production environments!")
            app.run(debug=True, host=host, port=port)
        else:
            logger.info(f"Starting development server without debug mode on {host}:{port}")
            app.run(debug=False, host=host, port=port)
